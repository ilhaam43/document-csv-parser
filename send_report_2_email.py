"""Capture and email the Report 2 ongoing-tracking PIVOT sheet."""

from __future__ import annotations

import argparse
from pathlib import Path

from send_report_1_email import excel_range_to_png, send_email_images

REPORT_2_IMAGE_IDS = (
    "report-2-ongoing-table",
    "report-2-ongoing-legend",
    "report-2-after-table",
    "report-2-after-legend",
    "report-2-aging-table",
)


def report_2_ranges(workbook_path: Path) -> list[tuple[str, str]]:
    from openpyxl import load_workbook

    workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    sheet = workbook["PIVOT"]
    rows: dict[str, int] = {}
    for row in sheet.iter_rows():
        for cell in row:
            value = str(cell.value or "").strip().upper()
            if "TARGET ORDER ON GOING PROCESS" in value:
                rows.setdefault("ongoing", cell.row)
            elif "DELAY COMPLETION (SHOULD BE COMPLETE BEFORE" in value:
                rows.setdefault("delay", cell.row)
            elif "TARGET AFTER" in value:
                rows.setdefault("after", cell.row)
            elif "ALL TARGET ORDER" in value:
                rows.setdefault("all_target", cell.row)
            elif "ORDER AGING START FROM PRE-INSTALLATION STATUS" in value:
                rows.setdefault("aging", cell.row)

    missing = [name for name in ("ongoing", "delay", "after", "all_target", "aging") if name not in rows]
    if missing:
        raise RuntimeError(f"Could not find Report 2 PIVOT headers: {', '.join(missing)}")
    aging_cells = [
        cell
        for row in sheet.iter_rows(min_col=19, max_col=38, max_row=rows["after"] - 1)
        for cell in row
        if cell.value is not None
    ]
    aging_end = max(
        (cell.row for cell in aging_cells),
        default=15,
    )
    aging_column = max((cell.column for cell in aging_cells), default=25)
    from openpyxl.utils import get_column_letter
    return [
        (f"A{rows['ongoing']}:N{rows['after'] - 1}", REPORT_2_IMAGE_IDS[0]),
        ("O12:O14", REPORT_2_IMAGE_IDS[1]),
        (f"A{rows['after']}:N{rows['all_target'] - 1}", REPORT_2_IMAGE_IDS[2]),
        ("T232:T234", REPORT_2_IMAGE_IDS[3]),
        (f"S{rows['aging']}:{get_column_letter(aging_column)}{aging_end}", REPORT_2_IMAGE_IDS[4]),
    ]


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, required=True)
    parser.add_argument("--image", type=Path)
    parser.add_argument("--endpoint", default=None)
    parser.add_argument("--to", default="ilhaam.akmal@lintasarta.co.id")
    parser.add_argument("--subject", default="Daily Tracking Report 2 Ongoing")
    parser.add_argument("--message", default='<p>Daily Tracking Report 2 Ongoing</p><p><img src="cid:report-2-pivot" alt="Report 2 ongoing tracking PIVOT"></p>')
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    workbook = args.workbook.resolve()
    if not workbook.exists():
        parser.error(f"Workbook not found: {workbook}")
    image = (args.image or workbook.with_name(f"{workbook.stem} - report-2-pivot.png")).resolve()
    ranges = report_2_ranges(workbook)
    output_dir = image.parent
    images = []
    for index, (cell_range, content_id) in enumerate(ranges, start=1):
        image_path = output_dir / f"{workbook.stem} - report-2-pivot-{index}.png"
        excel_range_to_png(workbook, image_path, "PIVOT", cell_range, scale=3.0 if "legend" in content_id else 1.0)
        images.append((image_path, content_id))
        print(f"Created image: {image_path} ({cell_range})")
    if args.dry_run:
        print("Dry run: SMTP request skipped.")
        return 0

    endpoint = args.endpoint or "http://10.34.144.197/secm-portal/smtp/api_send_email"
    send_email_images(endpoint, args.to, args.subject, args.message, images, timeout=60)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

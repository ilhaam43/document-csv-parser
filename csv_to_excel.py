#!/usr/bin/env python3
"""Parse and clean CSV files, then export them to Excel workbooks."""

from __future__ import annotations

import argparse
import csv
import os
import re
import shutil
import sys
import tempfile
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Iterable

import pandas as pd
from charset_normalizer import from_path
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


COMMON_ENCODINGS = ("utf-8-sig", "utf-8", "cp1252", "latin1")
NULL_LIKE_VALUES = {"", "na", "n/a", "null", "none", "nan", "-"}
EXCEL_MAX_SHEET_NAME_LENGTH = 31
DEFAULT_INPUT_DIR = "input-today"
DEFAULT_OUTPUT_DIR = "output-today"
DEFAULT_VLOOKUP_DIR = "vlookup-yesterday"
TARGET_HEADER_INSERT_AFTER = "Quo"
TARGET_SOURCE_HEADER = "Target + On Hold Duration"
MIN_VALID_TARGET_DATE = date(2000, 1, 1)
UNKNOWN_YEAR_FAB_UPLOAD = "1900"
MIN_REPORTING_YEAR_FAB_UPLOAD = 2023
YEAR_FAB_UPLOAD_BLANK_LIKE_VALUES = {"", UNKNOWN_YEAR_FAB_UPLOAD, "nan", "none", "n/a", "(blank)", "null"}
QUO_EXCLUDE_TERM = "XBOT"
DEPT_SD_HEADER = "Dept SD"
PM_HEADER = "PM"
QUO_HEADER = "Quo"
STATUS_HEADER = "Status"
PROCESS_HEADER = "Process"
PROCESS_ADJUSTMENT_HEADER = "Process Adjustment"
PRODUCT_CATEGORY_HEADER = "Product Category"
GROUP_SALES_HEADER = "group_sales"
DIVISION_SALES_HEADER = "Division_sales"
SEGMENT_SALES_HEADER = "segment_sales"
SALES_HEADER = "Sales"
SERVICE_DELIVERY_DIV_HEADER = "Service Delivery Div"
LOOKUP_SERVICE_DELIVERY_DIV_HEADER = "Service Delivery Div."
LOOKUP_GROUP_SALES_HEADER = "Group Sales"
LOOKUP_DIVISION_SALES_HEADER = "Division Sales"
LOOKUP_SEGMENT_SALES_HEADER = "Segment Sales"
PIVOT_TEMPLATE_SERVICE_DELIVERY_DIV_HEADER = "Service Delivery Div. "
PHASE_HEADER_PREFIX = "Phase"
FAB_UPLOAD_HEADER = "FAB Upload"
YEAR_FAB_UPLOAD_HEADER = "YEAR FAB UPLOAD"
RFS_COMMMIT_HEADER = "RFS Commmit"
AGING_OF_RFS_HEADER = "Aging Of RFS"
STATUS_ORDER_HEADER = "STATUS ORDER"
TARGET_SO_COMPLETE_DATE_HEADERS = ("Target SO Complete Date", "Target SO Completion Date")
EXCEL_DATE_DISPLAY_FORMAT = "yyyy-mm-dd"
EXCEL_TABLE_STYLE_NAME = "TableStyleMedium2"
EXCEL_HEADER_FONT_COLOR = "FFFFFF"
EXCEL_RED_FONT_COLOR = "FF0000"
EXCEL_RED_HEADER_FILL = "FF0000"
EXCEL_YELLOW_HEADER_FILL = "FFFF00"
EXCEL_HEADER_ROW_HEIGHT_POINTS = 22.5
EXCEL_ON_PROGRESS_TAB_COLOR = "ADD8E6"
EXCEL_PIVOT_TAB_COLOR = "FFFF00"
VLOOKUP_SHEET_NAME = "ALL ORDER"
ON_PROGRESS_SHEET_PREFIX = "ALL ORDER ON PROGRESS"
PIVOT_SHEET_NAME = "PIVOT"
LOOKUP_REQUIRED_HEADERS = (
    PM_HEADER,
    DEPT_SD_HEADER,
    LOOKUP_SERVICE_DELIVERY_DIV_HEADER,
    QUO_HEADER,
    PHASE_HEADER_PREFIX,
    PROCESS_ADJUSTMENT_HEADER,
    PRODUCT_CATEGORY_HEADER,
    SALES_HEADER,
    LOOKUP_GROUP_SALES_HEADER,
    LOOKUP_DIVISION_SALES_HEADER,
    LOOKUP_SEGMENT_SALES_HEADER,
)


def is_excluded_year_fab_upload_value(value: object) -> bool:
    normalized = str(value).strip()
    if normalized.lower() in YEAR_FAB_UPLOAD_BLANK_LIKE_VALUES:
        return True

    if re.fullmatch(r"\d{4}(?:\.0+)?", normalized):
        return int(float(normalized)) < MIN_REPORTING_YEAR_FAB_UPLOAD

    return False


@dataclass(frozen=True)
class ConvertOptions:
    output: Path
    delimiter: str | None
    encoding: str | None
    normalize_headers: bool
    keep_empty: bool
    drop_empty_columns: bool
    dedupe: bool
    infer_types: bool
    combine: bool
    refresh_template: bool = True


@dataclass(frozen=True)
class LookupMappings:
    dept_sd: dict[str, str]
    service_delivery_div: dict[str, str]
    status: dict[str, str]
    phase: dict[str, str]
    process_adjustment: dict[str, str]
    product_category: dict[str, str]
    group_sales: dict[str, str]
    division_sales: dict[str, str]
    segment_sales: dict[str, str]
    phase_header: str


def detect_encoding(path: Path) -> str:
    """Detect file encoding with a conservative fallback list."""
    result = from_path(path).best()
    if result and result.encoding:
        return result.encoding

    for encoding in COMMON_ENCODINGS:
        try:
            path.read_text(encoding=encoding)
            return encoding
        except UnicodeDecodeError:
            continue

    return "latin1"


def detect_delimiter(path: Path, encoding: str) -> str:
    sample = path.read_text(encoding=encoding, errors="replace")[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        return dialect.delimiter
    except csv.Error:
        return ","


def normalize_column_name(value: object, index: int) -> str:
    raw = str(value).replace("\ufeff", "").replace("_", " ").strip()
    raw = re.sub(r"^#+", "", raw).strip()
    raw = re.sub(r"\s+", " ", raw)
    if not raw:
        return f"Column {index + 1}"

    words = []
    for word in raw.split(" "):
        if word.lower() == "otc":
            words.append("OTC")
        elif word.isupper() or re.fullmatch(r"[A-Z0-9#()/-]+", word):
            words.append(word)
        else:
            words.append(word[:1].upper() + word[1:])

    return " ".join(words)


def make_unique_columns(columns: Iterable[object], normalize: bool) -> list[str]:
    names: list[str] = []
    seen: dict[str, int] = {}

    for index, column in enumerate(columns):
        name = normalize_column_name(column, index) if normalize else str(column).strip()
        name = name or f"Column {index + 1}"

        count = seen.get(name, 0)
        seen[name] = count + 1
        names.append(name if count == 0 else f"{name} ({count + 1})")

    return names


def rename_output_headers(df: pd.DataFrame) -> pd.DataFrame:
    return df.rename(
        columns={
            "Group Sales": GROUP_SALES_HEADER,
            "Division Sales": DIVISION_SALES_HEADER,
            "Segment Sales": SEGMENT_SALES_HEADER,
        }
    )


def escape_excel_formula_text(value: object) -> object:
    if isinstance(value, str) and value.startswith("="):
        return f"'{value}"
    return value


def escape_excel_formulas(df: pd.DataFrame) -> pd.DataFrame:
    escaped = df.copy()
    escaped.columns = [escape_excel_formula_text(column) for column in escaped.columns]

    for column in escaped.columns:
        if pd.api.types.is_object_dtype(escaped[column]) or pd.api.types.is_string_dtype(escaped[column]):
            escaped[column] = escaped[column].map(escape_excel_formula_text, na_action="ignore")

    return escaped


def current_target_header() -> str:
    today = date.today()
    return f"TARGET  Detemined as 1 {today.strftime('%B %Y')}"


def current_target_values(source: pd.Series) -> pd.Series:
    today = date.today()
    month_name = today.strftime("%B")
    target_month = pd.Period(today, freq="M")
    source_dates = pd.to_datetime(source, errors="coerce", format="mixed")
    valid_source_dates = source_dates.where(source_dates.dt.date >= MIN_VALID_TARGET_DATE)
    source_months = valid_source_dates.dt.to_period("M")
    values = pd.Series("Target Not Yet Inputted", index=source.index, dtype="string")

    values[source_months < target_month] = f"Before {month_name}"
    values[source_months == target_month] = f"Target {month_name}"
    values[source_months > target_month] = f"After {month_name}"

    return values


def add_target_header(df: pd.DataFrame) -> pd.DataFrame:
    if TARGET_HEADER_INSERT_AFTER not in df.columns or TARGET_SOURCE_HEADER not in df.columns:
        return df

    result = df.copy()
    target_header = current_target_header()
    if target_header in result.columns:
        result[target_header] = current_target_values(result[TARGET_SOURCE_HEADER])
        return result

    insert_at = result.columns.get_loc(TARGET_HEADER_INSERT_AFTER) + 1
    result.insert(insert_at, target_header, current_target_values(result[TARGET_SOURCE_HEADER]))
    return result


def drop_excluded_quo_rows(df: pd.DataFrame) -> pd.DataFrame:
    if TARGET_HEADER_INSERT_AFTER not in df.columns:
        return df

    quo_values = df[TARGET_HEADER_INSERT_AFTER].astype("string")
    return df.loc[~quo_values.str.contains(QUO_EXCLUDE_TERM, case=False, na=False)]


def add_year_fab_upload_column(df: pd.DataFrame) -> pd.DataFrame:
    if FAB_UPLOAD_HEADER not in df.columns:
        return df

    result = df.copy()
    fab_upload_dates = pd.to_datetime(result[FAB_UPLOAD_HEADER], errors="coerce", format="mixed")
    year_values = pd.Series(UNKNOWN_YEAR_FAB_UPLOAD, index=result.index, dtype="string")
    has_date = fab_upload_dates.notna()
    year_values.loc[has_date] = fab_upload_dates.loc[has_date].dt.strftime("%Y")

    if YEAR_FAB_UPLOAD_HEADER in result.columns:
        result[YEAR_FAB_UPLOAD_HEADER] = year_values
        return result

    insert_at = result.columns.get_loc(FAB_UPLOAD_HEADER) + 1
    result.insert(insert_at, YEAR_FAB_UPLOAD_HEADER, year_values)
    return result


def reference_date_from_csv_path(csv_path: Path) -> date:
    match = re.search(r"(\d{8})", csv_path.stem)
    if match:
        return pd.to_datetime(match.group(1), format="%Y%m%d").date()

    return date.today()


def output_filename_from_csv_path(csv_path: Path) -> str:
    reference_date = reference_date_from_csv_path(csv_path)
    return f"Daily Tracking {reference_date.day} {reference_date.strftime('%B %Y')}.xlsx"


def add_aging_of_rfs_column(df: pd.DataFrame, reference_date: date) -> pd.DataFrame:
    if RFS_COMMMIT_HEADER not in df.columns:
        return df

    result = df.copy()
    rfs_commit_dates = pd.to_datetime(result[RFS_COMMMIT_HEADER], errors="coerce", format="mixed")
    aging_values = pd.Series(pd.NA, index=result.index, dtype="Int64")
    has_date = rfs_commit_dates.notna()
    aging_values.loc[has_date] = (pd.Timestamp(reference_date) - rfs_commit_dates.loc[has_date]).dt.days

    if AGING_OF_RFS_HEADER in result.columns:
        result[AGING_OF_RFS_HEADER] = aging_values
        return result

    insert_at = result.columns.get_loc(RFS_COMMMIT_HEADER) + 1
    result.insert(insert_at, AGING_OF_RFS_HEADER, aging_values)
    return result


def add_status_order_column(df: pd.DataFrame) -> pd.DataFrame:
    if AGING_OF_RFS_HEADER not in df.columns:
        return df

    result = df.copy()
    aging_values = pd.to_numeric(result[AGING_OF_RFS_HEADER], errors="coerce")
    status_order_values = pd.Series(pd.NA, index=result.index, dtype="string")
    has_aging = aging_values.notna()
    status_order_values.loc[has_aging] = "Delay"
    status_order_values.loc[aging_values < 0] = "Potential Delay"
    status_order_values.loc[aging_values < -5] = "On Track"

    if STATUS_ORDER_HEADER in result.columns:
        result[STATUS_ORDER_HEADER] = status_order_values
        return result

    insert_at = result.columns.get_loc(AGING_OF_RFS_HEADER) + 1
    result.insert(insert_at, STATUS_ORDER_HEADER, status_order_values)
    return result


def clean_dataframe(df: pd.DataFrame, options: ConvertOptions, csv_path: Path) -> pd.DataFrame:
    df = df.copy()
    df.columns = make_unique_columns(df.columns, options.normalize_headers)
    df = rename_output_headers(df)
    df = add_target_header(df)

    for column in df.columns:
        if pd.api.types.is_object_dtype(df[column]) or pd.api.types.is_string_dtype(df[column]):
            cleaned = (
                df[column]
                .astype("string")
                .str.replace("\ufeff", "", regex=False)
                .str.strip()
                .str.replace(r"\s+", " ", regex=True)
            )
            df[column] = cleaned.mask(cleaned.str.lower().isin(NULL_LIKE_VALUES), pd.NA)

    df = drop_excluded_quo_rows(df)
    df = add_year_fab_upload_column(df)
    df = add_aging_of_rfs_column(df, reference_date_from_csv_path(csv_path))
    df = add_status_order_column(df)

    if not options.keep_empty:
        df = df.dropna(how="all")

    if options.drop_empty_columns:
        df = df.dropna(axis="columns", how="all")

    if options.dedupe:
        df = df.drop_duplicates()

    if options.infer_types:
        df = infer_column_types(df)

    return escape_excel_formulas(df).reset_index(drop=True)


def infer_column_types(df: pd.DataFrame) -> pd.DataFrame:
    converted = df.copy()

    for column in converted.columns:
        series = converted[column]
        non_empty = series.dropna()
        if non_empty.empty:
            continue

        numeric_candidate = (
            non_empty.astype("string")
            .str.replace(",", "", regex=False)
            .str.replace("%", "", regex=False)
        )
        numeric = pd.to_numeric(numeric_candidate, errors="coerce")
        if numeric.notna().mean() >= 0.9:
            converted[column] = pd.to_numeric(
                series.astype("string").str.replace(",", "", regex=False).str.replace("%", "", regex=False),
                errors="coerce",
            )
            continue

        parsed_dates = pd.to_datetime(non_empty, errors="coerce", format="mixed")
        if parsed_dates.notna().mean() >= 0.9:
            converted[column] = pd.to_datetime(series, errors="coerce", format="mixed")

    return converted


def read_csv(path: Path, options: ConvertOptions) -> pd.DataFrame:
    encoding = options.encoding or detect_encoding(path)
    delimiter = options.delimiter or detect_delimiter(path, encoding)

    with path.open("r", encoding=encoding, errors="replace", newline="") as handle:
        rows = list(csv.reader(handle, delimiter=delimiter, skipinitialspace=True))

    first_data_row = next((index for index, row in enumerate(rows) if any(cell.strip() for cell in row)), None)
    if first_data_row is None:
        raise ValueError(f"CSV file is empty: {path}")

    header = rows[first_data_row]
    data_rows = rows[first_data_row + 1 :]
    width = max([len(header), *(len(row) for row in data_rows)] or [len(header)])

    if len(header) < width:
        header = [*header, *(f"Extra Column {index + 1}" for index in range(len(header), width))]

    normalized_rows = [
        [*row, *([""] * (width - len(row)))] if len(row) < width else row[:width]
        for row in data_rows
    ]

    return pd.DataFrame(normalized_rows, columns=header, dtype="string")


def sanitize_sheet_name(path: Path, used: set[str]) -> str:
    base = re.sub(r"[\[\]\:\*\?\/\\]", "_", path.stem).strip() or "Sheet"
    base = base[:EXCEL_MAX_SHEET_NAME_LENGTH]
    return make_unique_sheet_name(base, used)


def make_unique_sheet_name(base: str, used: set[str]) -> str:
    sheet_name = base[:EXCEL_MAX_SHEET_NAME_LENGTH] or "Sheet"
    suffix = 2

    while sheet_name in used:
        suffix_text = f"_{suffix}"
        sheet_name = f"{base[: EXCEL_MAX_SHEET_NAME_LENGTH - len(suffix_text)]}{suffix_text}"
        suffix += 1

    used.add(sheet_name)
    return sheet_name


def on_progress_sheet_name(csv_path: Path, used: set[str] | None = None) -> str:
    reference_date = reference_date_from_csv_path(csv_path)
    preferred = f"{ON_PROGRESS_SHEET_PREFIX} {reference_date.day} {reference_date.strftime('%b')}"

    if used is None:
        return preferred[:EXCEL_MAX_SHEET_NAME_LENGTH]

    return make_unique_sheet_name(preferred, used)


def resolve_vlookup_workbook() -> Path:
    lookup_dir = Path(__file__).resolve().parent / DEFAULT_VLOOKUP_DIR
    candidates = sorted(
        path for path in lookup_dir.glob("*.xlsx") if path.is_file() and not path.name.startswith("~$")
    )
    if not candidates:
        raise FileNotFoundError(f"No lookup workbook found in: {lookup_dir}")

    return max(candidates, key=lambda path: path.stat().st_mtime)


def phase_lookup_header(lookup_workbook: Path) -> str:
    match = re.search(r"Daily Tracking\s+(.+)$", lookup_workbook.stem, flags=re.IGNORECASE)
    if match:
        return f"{PHASE_HEADER_PREFIX} {match.group(1).strip()}"

    return PHASE_HEADER_PREFIX


def remember_first_mapping(mapping: dict[str, str], key: object, value: object) -> None:
    key_text = str(key).strip()
    if key_text and key_text not in mapping and not pd.isna(value):
        mapping[key_text] = str(value).strip()


def load_lookup_mappings(lookup_workbook: Path) -> LookupMappings:
    lookup_df = pd.read_excel(lookup_workbook, sheet_name=VLOOKUP_SHEET_NAME, dtype="string")
    lookup_df.columns = make_unique_columns(lookup_df.columns, normalize=True)

    missing_headers = [header for header in LOOKUP_REQUIRED_HEADERS if header not in lookup_df.columns]
    if missing_headers:
        raise ValueError(
            f"Lookup workbook sheet '{VLOOKUP_SHEET_NAME}' is missing required columns: "
            f"{', '.join(missing_headers)}"
        )

    dept_sd_mapping: dict[str, str] = {}
    service_delivery_div_mapping: dict[str, str] = {}
    status_mapping: dict[str, str] = {}
    phase_mapping: dict[str, str] = {}
    process_adjustment_mapping: dict[str, str] = {}
    product_category_mapping: dict[str, str] = {}
    group_sales_mapping: dict[str, str] = {}
    division_sales_mapping: dict[str, str] = {}
    segment_sales_mapping: dict[str, str] = {}
    mapping_headers = list(LOOKUP_REQUIRED_HEADERS)
    if STATUS_HEADER in lookup_df.columns:
        mapping_headers.append(STATUS_HEADER)

    for _, row in lookup_df[mapping_headers].dropna(subset=[PM_HEADER]).iterrows():
        pm_value = str(row[PM_HEADER]).strip()
        if not pm_value:
            continue

        remember_first_mapping(dept_sd_mapping, pm_value, row[DEPT_SD_HEADER])
        remember_first_mapping(service_delivery_div_mapping, pm_value, row[LOOKUP_SERVICE_DELIVERY_DIV_HEADER])

        quo_value = str(row[QUO_HEADER]).strip()
        if STATUS_HEADER in lookup_df.columns:
            remember_first_mapping(status_mapping, quo_value, row[STATUS_HEADER])
        remember_first_mapping(phase_mapping, quo_value, row[PHASE_HEADER_PREFIX])
        remember_first_mapping(process_adjustment_mapping, quo_value, row[PROCESS_ADJUSTMENT_HEADER])
        remember_first_mapping(product_category_mapping, quo_value, row[PRODUCT_CATEGORY_HEADER])

        sales_value = str(row[SALES_HEADER]).strip()
        remember_first_mapping(group_sales_mapping, sales_value, row[LOOKUP_GROUP_SALES_HEADER])
        remember_first_mapping(division_sales_mapping, sales_value, row[LOOKUP_DIVISION_SALES_HEADER])
        remember_first_mapping(segment_sales_mapping, sales_value, row[LOOKUP_SEGMENT_SALES_HEADER])

    return LookupMappings(
        dept_sd=dept_sd_mapping,
        service_delivery_div=service_delivery_div_mapping,
        status=status_mapping,
        phase=phase_mapping,
        process_adjustment=process_adjustment_mapping,
        product_category=product_category_mapping,
        group_sales=group_sales_mapping,
        division_sales=division_sales_mapping,
        segment_sales=segment_sales_mapping,
        phase_header=phase_lookup_header(lookup_workbook),
    )


def phase_from_status(status_values: pd.Series) -> pd.Series:
    normalized_status = status_values.astype("string").str.strip().map(normalize_header_key)
    phase_mapping = {
        "allocationiepcapacity": "03-Allocation",
        "cancel": "Cancel",
        "commercialissue": "04-Pre Installation",
        "depositbucket": "07-UAT On Hold",
        "installation": "06-Installation",
        "lastmileinstallation": "05-Customer Preparation",
        "new": "00-New",
        "onholdallocation": "03-Allocation",
        "onholdcustomerpreparation": "05-Customer Preparation",
        "onholdinstallation": "06-Installation",
        "onholdpreinstallation": "04-Pre Installation",
        "onholdsurvey": "02-Survey",
        "presalesverificationtechnical": "01-Presales",
        "proposetocancelquote": "Cancel",
        "proposetocancelso": "Cancel",
        "scheduledforsurvey": "02-Survey",
        "socomplete": "SO Complete",
        "survey": "02-Survey",
        "uatonhold": "07-UAT On Hold",
        "waitingso": "04-Pre Installation",
    }
    return pd.Series(normalized_status.map(phase_mapping), index=status_values.index, dtype="string")


def apply_current_status_phase_overrides(
    result: pd.DataFrame,
    phase_header: str,
    previous_status: pd.Series | None = None,
) -> None:
    if STATUS_HEADER not in result.columns or phase_header not in result.columns:
        return

    normalized_status = result[STATUS_HEADER].astype("string").str.strip().map(normalize_header_key)
    current_phase = result[phase_header].astype("string").str.strip()
    missing_phase = current_phase.isna() | current_phase.eq("")
    derived_phase = phase_from_status(result[STATUS_HEADER])
    has_derived_phase = derived_phase.notna()
    preserve_historical_so_complete = pd.Series(False, index=result.index)
    if previous_status is not None:
        previous_normalized_status = previous_status.astype("string").str.strip().map(normalize_header_key)
        preserve_historical_so_complete = (
            (normalized_status == "socomplete")
            & (previous_normalized_status == "socomplete")
            & ~missing_phase
        )

    override_phase = has_derived_phase & ~preserve_historical_so_complete
    result.loc[override_phase, phase_header] = derived_phase.loc[override_phase]


def apply_current_month_process_adjustment_overrides(result: pd.DataFrame, reference_date: date) -> None:
    required_columns = {PROCESS_HEADER, PROCESS_ADJUSTMENT_HEADER, FAB_UPLOAD_HEADER}
    if not required_columns.issubset(result.columns):
        return

    process_values = result[PROCESS_HEADER].astype("string").str.strip().map(normalize_header_key)
    process_adjustment_values = result[PROCESS_ADJUSTMENT_HEADER].astype("string").str.strip().fillna("")
    fab_upload_dates = pd.to_datetime(result[FAB_UPLOAD_HEADER], errors="coerce", format="mixed")
    current_report_month = pd.Period(reference_date, freq="M")

    missing_process_adjustment = process_adjustment_values.eq("")
    current_month_fab_upload = fab_upload_dates.dt.to_period("M").eq(current_report_month)
    inferred_new_registration = process_values.eq("newregistration")

    result.loc[
        missing_process_adjustment & current_month_fab_upload & inferred_new_registration,
        PROCESS_ADJUSTMENT_HEADER,
    ] = "New Reg"


def apply_lookup_values(df: pd.DataFrame, mappings: LookupMappings) -> pd.DataFrame:
    if PM_HEADER not in df.columns and QUO_HEADER not in df.columns:
        return df

    result = df.copy()
    pm_values = result[PM_HEADER].astype("string").str.strip() if PM_HEADER in result.columns else None

    if DEPT_SD_HEADER in result.columns and pm_values is not None:
        looked_up_dept_sd = pm_values.map(mappings.dept_sd)
        result[DEPT_SD_HEADER] = pd.Series(looked_up_dept_sd, index=result.index, dtype="string")

    if pm_values is not None:
        looked_up_service_delivery_div = pd.Series(pm_values.map(mappings.service_delivery_div), index=result.index, dtype="string")
        if SERVICE_DELIVERY_DIV_HEADER in result.columns:
            result[SERVICE_DELIVERY_DIV_HEADER] = looked_up_service_delivery_div
        else:
            insert_at = result.columns.get_loc(DEPT_SD_HEADER) + 1 if DEPT_SD_HEADER in result.columns else len(result.columns)
            result.insert(insert_at, SERVICE_DELIVERY_DIV_HEADER, looked_up_service_delivery_div)

    if SALES_HEADER in result.columns:
        looked_up_group_sales = pd.Series(
            result[SALES_HEADER].astype("string").str.strip().map(mappings.group_sales),
            index=result.index,
            dtype="string",
        )
        if GROUP_SALES_HEADER in result.columns:
            result[GROUP_SALES_HEADER] = looked_up_group_sales

        looked_up_division_sales = pd.Series(
            result[SALES_HEADER].astype("string").str.strip().map(mappings.division_sales),
            index=result.index,
            dtype="string",
        )
        if DIVISION_SALES_HEADER in result.columns:
            result[DIVISION_SALES_HEADER] = looked_up_division_sales
        else:
            insert_at = result.columns.get_loc(GROUP_SALES_HEADER) + 1 if GROUP_SALES_HEADER in result.columns else len(result.columns)
            result.insert(insert_at, DIVISION_SALES_HEADER, looked_up_division_sales)

        looked_up_segment_sales = pd.Series(
            result[SALES_HEADER].astype("string").str.strip().map(mappings.segment_sales),
            index=result.index,
            dtype="string",
        )
        if SEGMENT_SALES_HEADER in result.columns:
            result[SEGMENT_SALES_HEADER] = looked_up_segment_sales

    if QUO_HEADER in result.columns:
        quo_values = result[QUO_HEADER].astype("string").str.strip()
        looked_up_phase = pd.Series(quo_values.map(mappings.phase), index=result.index, dtype="string")
        looked_up_status = pd.Series(quo_values.map(mappings.status), index=result.index, dtype="string")
        if mappings.phase_header in result.columns:
            result[mappings.phase_header] = looked_up_phase
        else:
            insert_at = result.columns.get_loc(STATUS_HEADER) + 1 if STATUS_HEADER in result.columns else len(result.columns)
            result.insert(insert_at, mappings.phase_header, looked_up_phase)

        looked_up_process_adjustment = pd.Series(
            quo_values.map(mappings.process_adjustment),
            index=result.index,
            dtype="string",
        )
        if PROCESS_ADJUSTMENT_HEADER in result.columns:
            result[PROCESS_ADJUSTMENT_HEADER] = looked_up_process_adjustment
        else:
            insert_at = result.columns.get_loc(PROCESS_HEADER) + 1 if PROCESS_HEADER in result.columns else len(result.columns)
            result.insert(insert_at, PROCESS_ADJUSTMENT_HEADER, looked_up_process_adjustment)

        apply_current_status_phase_overrides(result, mappings.phase_header, looked_up_status)

        if PRODUCT_CATEGORY_HEADER in result.columns:
            result[PRODUCT_CATEGORY_HEADER] = pd.Series(
                quo_values.map(mappings.product_category),
                index=result.index,
                dtype="string",
            )

    return result


def write_excel_sheet(
    writer: pd.ExcelWriter,
    df: pd.DataFrame,
    sheet_name: str,
    mappings: LookupMappings,
) -> pd.DataFrame:
    df = apply_lookup_values(df, mappings)
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    worksheet = writer.sheets[sheet_name]
    apply_target_so_complete_date_filter_format(worksheet, df)
    apply_worksheet_presentation(worksheet)
    return df


def pivot_template_target_header(csv_path: Path) -> str:
    reference_date = reference_date_from_csv_path(csv_path)
    return f"TARGET  Detemined as 1 {reference_date.strftime('%b')} {reference_date.strftime('%y')}"


def add_pivot_template_compatibility_columns(
    df: pd.DataFrame,
    csv_path: Path,
    target_header: str,
    phase_header: str,
) -> tuple[pd.DataFrame, set[str]]:
    """Add hidden source fields required by the copied PIVOT sheet."""
    result = df.copy()
    hidden_columns: set[str] = set()

    short_target_header = pivot_template_target_header(csv_path)
    if target_header and target_header in result.columns and short_target_header not in result.columns:
        result[short_target_header] = result[target_header]
        hidden_columns.add(short_target_header)

    if SERVICE_DELIVERY_DIV_HEADER in result.columns:
        result[PIVOT_TEMPLATE_SERVICE_DELIVERY_DIV_HEADER] = result[SERVICE_DELIVERY_DIV_HEADER]
        hidden_columns.add(PIVOT_TEMPLATE_SERVICE_DELIVERY_DIV_HEADER)

    if phase_header in result.columns:
        if PHASE_HEADER_PREFIX not in result.columns:
            hidden_columns.add(PHASE_HEADER_PREFIX)
        result[PHASE_HEADER_PREFIX] = result[phase_header]

    return result, hidden_columns


def hide_worksheet_columns(worksheet, hidden_columns: set[str]) -> None:
    if not hidden_columns:
        return

    for cell in worksheet[1]:
        if cell.value in hidden_columns:
            worksheet.column_dimensions[get_column_letter(cell.column)].hidden = True


def coerce_numeric_columns_for_pivots(df: pd.DataFrame) -> pd.DataFrame:
    result = df.copy()
    for column in ("MRC",):
        if column not in result.columns:
            continue

        cleaned = (
            result[column]
            .astype("string")
            .str.strip()
            .str.replace(",", "", regex=False)
            .str.replace(" ", "", regex=False)
        )
        result[column] = pd.to_numeric(cleaned, errors="coerce")

    return result


def write_all_order_sheet_for_template(
    writer: pd.ExcelWriter,
    df: pd.DataFrame,
    csv_path: Path,
    mappings: LookupMappings,
) -> tuple[pd.DataFrame, set[str]]:
    output_df = apply_lookup_values(df, mappings)
    apply_current_month_process_adjustment_overrides(output_df, reference_date_from_csv_path(csv_path))
    output_df = coerce_numeric_columns_for_pivots(output_df)
    target_header = next((column for column in output_df.columns if str(column).startswith("TARGET")), "")
    output_df, hidden_columns = add_pivot_template_compatibility_columns(
        output_df,
        csv_path,
        target_header,
        mappings.phase_header,
    )
    output_df.to_excel(writer, sheet_name=VLOOKUP_SHEET_NAME, index=False)
    worksheet = writer.sheets[VLOOKUP_SHEET_NAME]
    apply_target_so_complete_date_filter_format(worksheet, output_df)
    apply_worksheet_presentation(worksheet)
    hide_worksheet_columns(worksheet, hidden_columns)
    return output_df, hidden_columns


def write_on_progress_sheet(writer: pd.ExcelWriter, df: pd.DataFrame, csv_path: Path, sheet_name: str) -> None:
    target_header = next((column for column in df.columns if str(column).startswith("TARGET")), None)
    if target_header is None or QUO_HEADER not in df.columns or PHASE_HEADER_PREFIX not in df.columns:
        return

    reference_date = reference_date_from_csv_path(csv_path)
    month_name = reference_date.strftime("%B")
    target_month_label = f"Target {month_name}"
    source_month_label = f"This {month_name}"
    filtered_all = df.copy()
    if YEAR_FAB_UPLOAD_HEADER in filtered_all.columns:
        year_values = filtered_all[YEAR_FAB_UPLOAD_HEADER].astype("string").fillna("")
        filtered_all = filtered_all[~year_values.map(is_excluded_year_fab_upload_value)]

    filtered_table_2 = filtered_all.copy()
    if PHASE_HEADER_PREFIX in filtered_table_2.columns:
        phase_values = filtered_table_2[PHASE_HEADER_PREFIX].astype("string").str.strip().str.lower()
        filtered_table_2 = filtered_table_2[~phase_values.isin({"cancel", "so complete"})]

    pivot_all = pd.pivot_table(
        filtered_all,
        index=target_header,
        values=QUO_HEADER,
        aggfunc="count",
        fill_value=0,
    )
    counts_all = {str(index): int(row[QUO_HEADER]) for index, row in pivot_all.iterrows()}
    if source_month_label in counts_all:
        counts_all[target_month_label] = counts_all.pop(source_month_label)

    pivot_table_2 = pd.pivot_table(
        filtered_table_2,
        index=target_header,
        values=QUO_HEADER,
        aggfunc="count",
        fill_value=0,
    )
    counts_table_2 = {str(index): int(row[QUO_HEADER]) for index, row in pivot_table_2.iterrows()}
    if source_month_label in counts_table_2:
        counts_table_2[target_month_label] = counts_table_2.pop(source_month_label)

    worksheet = writer.book.create_sheet(title=sheet_name)
    writer.sheets[sheet_name] = worksheet
    worksheet.sheet_properties.tabColor = EXCEL_ON_PROGRESS_TAB_COLOR
    write_on_progress_filters(worksheet, start_col=1, phase_value="(All)")
    write_on_progress_filters(worksheet, start_col=4, phase_value="(Multiple Items)")
    write_on_progress_table(
        worksheet,
        start_row=5,
        start_col=1,
        labels=[target_month_label, f"After {month_name}"],
        counts=counts_all,
    )
    write_on_progress_table(
        worksheet,
        start_row=5,
        start_col=4,
        labels=[f"Before {month_name}", "Target Not Yet Inputted"],
        counts=counts_table_2,
    )

    autofit_worksheet_columns(worksheet)


def write_on_progress_filters(worksheet, start_col: int, phase_value: str) -> None:
    worksheet.cell(row=1, column=start_col, value=PHASE_HEADER_PREFIX)
    worksheet.cell(row=1, column=start_col + 1, value=phase_value)
    worksheet.cell(row=2, column=start_col, value=PROCESS_ADJUSTMENT_HEADER)
    worksheet.cell(row=2, column=start_col + 1, value="(All)")
    worksheet.cell(row=3, column=start_col, value=YEAR_FAB_UPLOAD_HEADER)
    worksheet.cell(row=3, column=start_col + 1, value="(Multiple Items)")


def write_on_progress_table(
    worksheet,
    start_row: int,
    start_col: int,
    labels: list[str],
    counts: dict[str, int],
) -> None:
    worksheet.cell(row=start_row, column=start_col, value="Row Labels")
    worksheet.cell(row=start_row, column=start_col + 1, value="Count of quo")
    for cell in worksheet[start_row]:
        if cell.column in {start_col, start_col + 1}:
            cell.font = Font(bold=True)

    row_index = start_row + 1
    subtotal = 0
    for label in labels:
        if label not in counts:
            continue

        worksheet.cell(row=row_index, column=start_col, value=label)
        worksheet.cell(row=row_index, column=start_col + 1, value=counts[label])
        subtotal += counts[label]
        row_index += 1

    worksheet.cell(row=row_index, column=start_col, value="Grand Total")
    worksheet.cell(row=row_index, column=start_col + 1, value=subtotal)


def add_pivot_tables_via_com(output_path: Path, on_progress_name: str, target_header: str) -> None:
    """Post-process the saved workbook via the Excel COM API to add real PivotTables."""
    try:
        import win32com.client  # type: ignore
    except ImportError:
        print("[warn] pywin32 not installed – skipping real PivotTable creation.", file=sys.stderr)
        return

    xl = None
    wb = None
    try:
        xl = win32com.client.DispatchEx("Excel.Application")
        xl.Visible = False
        xl.DisplayAlerts = False

        wb = xl.Workbooks.Open(str(output_path.resolve()))
        data_sheet = None
        progress_sheet = None
        for sheet in wb.Sheets:
            if sheet.Name == VLOOKUP_SHEET_NAME:
                data_sheet = sheet
            elif sheet.Name == on_progress_name:
                progress_sheet = sheet

        if data_sheet is None or progress_sheet is None:
            print("[warn] Could not find required sheets for PivotTable creation.", file=sys.stderr)
            return

        # Clear the existing static summary that openpyxl wrote
        progress_sheet.Cells.Clear()

        # Build the PivotCache from the ALL ORDER sheet data range
        data_range = data_sheet.UsedRange
        pivot_cache = wb.PivotCaches().Create(
            SourceType=1,  # xlDatabase
            SourceData=data_range,
        )

        # Constants
        xlRowField   = 1
        xlPageField  = 3
        xlCount      = -4112
        xlTabularRow = 0

        date_match = re.search(r"(\d{1,2}) ([A-Za-z]{3})$", on_progress_name)
        month_short = date_match.group(2) if date_match else ""
        month_full = datetime.strptime(month_short, "%b").strftime("%B") if month_short else ""
        source_month_label = f"This {month_full}" if month_full else ""
        target_month_label = f"Target {month_full}" if month_full else ""

        def _build_pivot(
            top_left_cell,
            table_name: str,
            visible_row_items: list[str] | None = None,
            exclude_phase_values: set[str] | None = None,
        ) -> None:
            table_range = progress_sheet.Range(top_left_cell)
            pt = pivot_cache.CreatePivotTable(
                TableDestination=table_range,
                TableName=table_name,
            )
            pt.ManualUpdate = True

            # --- Filter: YEAR FAB UPLOAD (all except out-of-reporting-scope values) ---
            try:
                year_field = pt.PivotFields(YEAR_FAB_UPLOAD_HEADER)
                year_field.Orientation = xlPageField
                year_field.EnableMultiplePageItems = True
                for item in year_field.PivotItems():
                    if is_excluded_year_fab_upload_value(item.Value):
                        item.Visible = False
            except Exception:
                pass

            # --- Filter: Process Adjustment (All) ---
            try:
                pt.PivotFields(PROCESS_ADJUSTMENT_HEADER).Orientation = xlPageField
            except Exception:
                pass

            # --- Filter: Phase (All or multiple items with exclusions) ---
            try:
                phase_field = pt.PivotFields(PHASE_HEADER_PREFIX)
                phase_field.Orientation = xlPageField
                if exclude_phase_values:
                    phase_field.EnableMultiplePageItems = True
                    excluded_lower = {value.lower() for value in exclude_phase_values}
                    for item in phase_field.PivotItems():
                        if str(item.Value).strip().lower() in excluded_lower:
                            item.Visible = False
            except Exception:
                pass

            # --- Row field: TARGET Determined as 1 <Month Year> ---
            try:
                rf = pt.PivotFields(target_header)
                rf.Orientation = xlRowField
                rf.LayoutForm = xlTabularRow
                if visible_row_items:
                    visible_lower = {value.lower() for value in visible_row_items}
                    for item in rf.PivotItems():
                        item_name = str(item.Name).strip()
                        compare_name = target_month_label if item_name == source_month_label else item_name
                        item.Visible = compare_name.lower() in visible_lower
                    for position, label in enumerate(visible_row_items, start=1):
                        for item in rf.PivotItems():
                            item_name = str(item.Name).strip()
                            compare_name = target_month_label if item_name == source_month_label else item_name
                            if compare_name.lower() == label.lower():
                                item.Position = position
                                break
            except Exception:
                pass

            # --- Value field: Count of quo ---
            try:
                df_field = pt.AddDataField(
                    pt.PivotFields(QUO_HEADER),
                    "Count of quo",
                    xlCount,
                )
                df_field.NumberFormat = "0"
            except Exception:
                pass

            try:
                pt.TableStyle2 = "PivotStyleMedium2"
            except Exception:
                pass

            pt.ManualUpdate = False
            pt.Update()

            if source_month_label and target_month_label:
                try:
                    rf = pt.PivotFields(target_header)
                    for item in rf.PivotItems():
                        if str(item.Name).strip() == source_month_label:
                            item.Caption = target_month_label
                except Exception:
                    pass

        _build_pivot(
            "A1",
            f"PivotTable_{on_progress_name}_1",
            visible_row_items=[target_month_label, f"After {month_full}"] if month_full else None,
        )
        _build_pivot(
            "D1",
            f"PivotTable_{on_progress_name}_2",
            visible_row_items=[f"Before {month_full}", "Target Not Yet Inputted"] if month_full else None,
            exclude_phase_values={"cancel", "so complete"},
        )

        wb.Save()
        print(f"[info] Added PivotTables to '{on_progress_name}' in {output_path.name}")
    except Exception as exc:
        print(f"[warn] Could not add PivotTables via COM: {exc}", file=sys.stderr)
    finally:
        if wb is not None:
            try:
                wb.Close(SaveChanges=False)
            except Exception:
                pass
        if xl is not None:
            try:
                xl.Quit()
            except Exception:
                pass


def add_pivot_sheet_via_com(
    output_path: Path,
    source_workbook_path: Path,
    csv_path: Path,
    target_header: str,
    phase_header: str,
) -> None:
    """Build the PIVOT sheet via Excel COM using ALL ORDER as the live data source."""
    try:
        import win32com.client  # type: ignore
    except ImportError:
        print("[warn] pywin32 not installed – skipping PIVOT sheet creation.", file=sys.stderr)
        return

    xl = None
    target_wb = None
    source_wb = None
    try:
        xl = win32com.client.DispatchEx("Excel.Application")
        xl.Visible = False
        xl.DisplayAlerts = False

        target_wb = xl.Workbooks.Open(str(output_path.resolve()))
        source_wb = xl.Workbooks.Open(str(source_workbook_path.resolve()))

        data_sheet = None
        for sheet in target_wb.Sheets:
            if sheet.Name == VLOOKUP_SHEET_NAME:
                data_sheet = sheet
                break

        source_pivot_sheet = None
        for sheet in source_wb.Sheets:
            if sheet.Name == PIVOT_SHEET_NAME:
                source_pivot_sheet = sheet
                break

        if data_sheet is None or source_pivot_sheet is None:
            print("[warn] Could not find ALL ORDER or source PIVOT sheet.", file=sys.stderr)
            return

        pivot_sheet = None
        for sheet in target_wb.Sheets:
            if sheet.Name == PIVOT_SHEET_NAME:
                pivot_sheet = sheet
                break
        if pivot_sheet is None:
            pivot_sheet = target_wb.Worksheets.Add(After=target_wb.Sheets(target_wb.Sheets.Count))
            pivot_sheet.Name = PIVOT_SHEET_NAME
        try:
            pivot_sheet.Cells.Clear()
        except Exception:
            pass
        try:
            pivot_sheet.Tab.Color = int(EXCEL_PIVOT_TAB_COLOR, 16)
        except Exception:
            pass

        used_rows = source_pivot_sheet.UsedRange.Rows.Count
        used_columns = source_pivot_sheet.UsedRange.Columns.Count
        for row_index in range(1, used_rows + 1):
            try:
                pivot_sheet.Rows(row_index).RowHeight = source_pivot_sheet.Rows(row_index).RowHeight
            except Exception:
                pass
        for column_index in range(1, used_columns + 1):
            try:
                pivot_sheet.Columns(column_index).ColumnWidth = source_pivot_sheet.Columns(column_index).ColumnWidth
            except Exception:
                pass

        reference_date = reference_date_from_csv_path(csv_path)
        month_name = reference_date.strftime("%B")
        month_name_upper = month_name.upper()
        month_short = reference_date.strftime("%b")
        year_value = reference_date.year
        year_short = reference_date.strftime("%y")
        target_month_label = f"Target {month_name}"
        before_month_label = f"Before {month_name}"
        after_month_label = f"After {month_name}"
        not_inputted_label = "Target Not Yet Inputted"
        short_target_header = f"TARGET  Detemined as 1 {month_short} {year_short}"

        pivot_sheet.Range("A1").Value = (
            f" TARGET COMPLETE {month_name_upper} {year_value} "
            f"(Based on Dashboard 1 {month_name_upper} {year_value})"
        )
        pivot_sheet.Range("A3").Value = "New Registration "
        pivot_sheet.Range("A109").Value = "Non New Registration "
        pivot_sheet.Range("A178").Value = (
            f"STATUS ORDER WITH AGING FROM RFS COMMIT DATE {month_name_upper} {year_short}"
        )
        pivot_sheet.Range("A368").Value = f"TARGET AFTER {month_name_upper} {year_value}"
        pivot_sheet.Range("E439").Value = "TARGET NOT INPUTTED YET"

        pivot_ranges = [
            "A5:C22",
            "D4:L22",
            "A110:C124",
            "D109:K124",
            "A180:F202",
            "A279:C297",
            "D279:I297",
            "A374:C387",
            "D373:J387",
            "E441:L461",
        ]
        for cell_range in pivot_ranges:
            try:
                pivot_sheet.Range(cell_range).Clear()
            except Exception:
                pass

        headers = [data_sheet.Cells(1, column).Value for column in range(1, data_sheet.UsedRange.Columns.Count + 1)]
        try:
            mrc_column = headers.index("MRC") + 1
        except ValueError:
            mrc_column = None
        if mrc_column is not None:
            for row_index in range(2, data_sheet.UsedRange.Rows.Count + 1):
                cell = data_sheet.Cells(row_index, mrc_column)
                value = cell.Value
                if isinstance(value, str):
                    stripped = value.strip()
                    if re.fullmatch(r"-?\d+(?:\.\d+)?", stripped):
                        try:
                            cell.Value = float(stripped) if "." in stripped else int(stripped)
                        except Exception:
                            pass

        pivot_cache = target_wb.PivotCaches().Create(
            SourceType=1,  # xlDatabase
            SourceData=data_sheet.UsedRange,
        )

        xlRowField = 1
        xlColumnField = 2
        xlPageField = 3
        xlHidden = 0
        xlCount = -4112
        xlSum = -4157
        xlPercent = 2
        def _normalize_item(value) -> str:
            return str(value).strip().lower()

        def _configure_page_field(
            pt,
            field_name: str,
            position: int,
            include_values: set[str] | None = None,
            exclude_values: set[str] | None = None,
            caption: str | None = None,
        ) -> None:
            field = pt.PivotFields(field_name)
            field.Orientation = xlPageField
            field.Position = position
            if caption is not None:
                try:
                    field.Caption = caption
                except Exception:
                    pass
            if include_values is None and exclude_values is None:
                return

            include_normalized = {_normalize_item(value) for value in include_values or set()}
            exclude_normalized = {_normalize_item(value) for value in exclude_values or set()}
            field.EnableMultiplePageItems = True
            try:
                for item in field.PivotItems():
                    item.Visible = True
            except Exception:
                pass

            visible_count = 0
            for item in field.PivotItems():
                item_value = _normalize_item(getattr(item, "Value", item.Name))
                visible = True
                if include_normalized:
                    visible = item_value in include_normalized
                if exclude_normalized and item_value in exclude_normalized:
                    visible = False
                if field_name == YEAR_FAB_UPLOAD_HEADER and is_excluded_year_fab_upload_value(
                    getattr(item, "Value", item.Name)
                ):
                    visible = False
                try:
                    item.Visible = visible
                    if visible:
                        visible_count += 1
                except Exception:
                    pass

            if visible_count == 0:
                try:
                    field.ClearAllFilters()
                except Exception:
                    pass

        def _add_data_fields(pt, include_sum_mrc: bool) -> None:
            if include_sum_mrc:
                try:
                    sum_field = pt.AddDataField(pt.PivotFields("MRC"), "Sum of MRC", xlSum)
                    sum_field.NumberFormat = "0"
                except Exception:
                    pass
            try:
                count_field = pt.AddDataField(pt.PivotFields(QUO_HEADER), "Count of quo", xlCount)
                count_field.NumberFormat = "0"
            except Exception:
                pass
            if include_sum_mrc:
                try:
                    pt.DataPivotField.Orientation = xlColumnField
                    pt.DataPivotField.Position = 1
                except Exception:
                    pass

        def _add_completion_field(pt) -> None:
            try:
                formula = (
                    "=('SO Complete'+'Cancel')/"
                    f"'{QUO_HEADER}'"
                )
                target_name = "Percentage of Completion (SO Complete, Cancel & Change Target)"
                try:
                    pt.CalculatedFields(target_name).Delete()
                except Exception:
                    pass
                pt.CalculatedFields().Add(target_name, formula, True)
                percent_field = pt.AddDataField(
                    pt.PivotFields(target_name),
                    target_name,
                    xlSum,
                )
                percent_field.NumberFormat = "0.00%"
            except Exception:
                pass

        def _build_pivot(
            table_name: str,
            destination: str,
            row_fields: list[str],
            column_field: str | None,
            page_fields: list[tuple[str, int, set[str] | None, set[str] | None, str | None]],
            include_sum_mrc: bool = False,
            row_grand: bool = True,
        ) -> None:
            pt = pivot_cache.CreatePivotTable(
                TableDestination=pivot_sheet.Range(destination),
                TableName=table_name,
            )
            pt.ManualUpdate = True
            try:
                pt.CompactLayoutRowHeader = "Div./Dept."
                pt.CompactLayoutColumnHeader = "Column Labels"
                pt.RowGrand = row_grand
                pt.ColumnGrand = True
            except Exception:
                pass
            for field_name in row_fields:
                try:
                    rf = pt.PivotFields(field_name)
                    rf.Orientation = xlRowField
                    rf.Position = row_fields.index(field_name) + 1
                except Exception:
                    pass
            if column_field:
                try:
                    cf = pt.PivotFields(column_field)
                    cf.Orientation = xlColumnField
                    cf.Position = 1
                    if column_field == phase_header:
                        cf.Caption = PHASE_HEADER_PREFIX
                except Exception:
                    pass
            for field_name, position, include_values, exclude_values, caption in page_fields:
                try:
                    _configure_page_field(pt, field_name, position, include_values, exclude_values, caption)
                except Exception:
                    pass
            _add_data_fields(pt, include_sum_mrc)
            try:
                pt.TableStyle2 = "PivotStyleLight14"
            except Exception:
                pass
            pt.ManualUpdate = False
            pt.Update()
            return pt

        year_excludes = YEAR_FAB_UPLOAD_BLANK_LIKE_VALUES
        pt1 = _build_pivot(
            "PivotTable_PIVOT_1",
            "A185",
            [SERVICE_DELIVERY_DIV_HEADER, DEPT_SD_HEADER, PM_HEADER],
            STATUS_ORDER_HEADER,
            [
                (YEAR_FAB_UPLOAD_HEADER, 1, None, year_excludes, YEAR_FAB_UPLOAD_HEADER),
                (PROCESS_ADJUSTMENT_HEADER, 2, {"new reg"}, None, PROCESS_ADJUSTMENT_HEADER),
                (
                    phase_header,
                    3,
                    {"00-new", "01-presales", "02-survey", "03-allocation", "04-pre installation", "05-customer preparation", "06-installation"},
                    None,
                    PHASE_HEADER_PREFIX,
                ),
                (RFS_COMMMIT_HEADER, 4, None, None, RFS_COMMMIT_HEADER),
            ],
            row_grand=True,
        )
        pt2 = _build_pivot(
            "PivotTable_PIVOT_2",
            "D282",
            [SERVICE_DELIVERY_DIV_HEADER, DEPT_SD_HEADER],
            phase_header,
            [
                (YEAR_FAB_UPLOAD_HEADER, 1, None, year_excludes, YEAR_FAB_UPLOAD_HEADER),
                (target_header, 2, {before_month_label.lower()}, None, short_target_header),
            ],
            row_grand=False,
        )
        pt3 = _build_pivot(
            "PivotTable_PIVOT_3",
            "A9",
            [SERVICE_DELIVERY_DIV_HEADER, DEPT_SD_HEADER],
            None,
            [
                (YEAR_FAB_UPLOAD_HEADER, 1, None, year_excludes, YEAR_FAB_UPLOAD_HEADER),
                (PROCESS_ADJUSTMENT_HEADER, 2, {"new reg"}, None, PROCESS_ADJUSTMENT_HEADER),
                (target_header, 3, {target_month_label.lower()}, None, short_target_header),
            ],
            include_sum_mrc=True,
            row_grand=True,
        )
        pt4 = _build_pivot(
            "PivotTable_PIVOT_4",
            "D8",
            [SERVICE_DELIVERY_DIV_HEADER, DEPT_SD_HEADER],
            phase_header,
            [
                (YEAR_FAB_UPLOAD_HEADER, 1, None, year_excludes, " "),
                (PROCESS_ADJUSTMENT_HEADER, 2, {"new reg"}, None, PROCESS_ADJUSTMENT_HEADER),
                (target_header, 3, {target_month_label.lower()}, None, short_target_header),
            ],
            row_grand=False,
        )
        pt5 = _build_pivot(
            "PivotTable_PIVOT_5",
            "A377",
            [SERVICE_DELIVERY_DIV_HEADER, DEPT_SD_HEADER],
            None,
            [
                (YEAR_FAB_UPLOAD_HEADER, 1, None, year_excludes, YEAR_FAB_UPLOAD_HEADER),
                (target_header, 2, {after_month_label.lower()}, None, short_target_header),
            ],
            include_sum_mrc=True,
            row_grand=True,
        )
        pt6 = _build_pivot(
            "PivotTable_PIVOT_6",
            "E444",
            [SERVICE_DELIVERY_DIV_HEADER, DEPT_SD_HEADER],
            phase_header,
            [
                (YEAR_FAB_UPLOAD_HEADER, 1, None, year_excludes, YEAR_FAB_UPLOAD_HEADER),
                (target_header, 2, {not_inputted_label.lower()}, None, short_target_header),
            ],
            row_grand=True,
        )
        pt7 = _build_pivot(
            "PivotTable_PIVOT_7",
            "A114",
            [SERVICE_DELIVERY_DIV_HEADER, DEPT_SD_HEADER],
            None,
            [
                (YEAR_FAB_UPLOAD_HEADER, 1, None, year_excludes, YEAR_FAB_UPLOAD_HEADER),
                (PROCESS_ADJUSTMENT_HEADER, 2, {"non new reg"}, None, PROCESS_ADJUSTMENT_HEADER),
                (target_header, 3, {target_month_label.lower()}, None, short_target_header),
            ],
            include_sum_mrc=True,
            row_grand=True,
        )
        pt8 = _build_pivot(
            "PivotTable_PIVOT_8",
            "A283",
            [SERVICE_DELIVERY_DIV_HEADER, DEPT_SD_HEADER],
            None,
            [
                (
                    phase_header,
                    1,
                    {"00-new", "01-presales", "02-survey", "03-allocation", "04-pre installation", "05-customer preparation", "06-installation", "07-uat on hold"},
                    None,
                    PHASE_HEADER_PREFIX,
                ),
                (YEAR_FAB_UPLOAD_HEADER, 2, None, year_excludes, YEAR_FAB_UPLOAD_HEADER),
                (target_header, 3, {before_month_label.lower()}, None, short_target_header),
            ],
            include_sum_mrc=True,
            row_grand=True,
        )
        pt9 = _build_pivot(
            "PivotTable_PIVOT_9",
            "D113",
            [SERVICE_DELIVERY_DIV_HEADER, DEPT_SD_HEADER],
            phase_header,
            [
                (YEAR_FAB_UPLOAD_HEADER, 1, None, year_excludes, YEAR_FAB_UPLOAD_HEADER),
                (PROCESS_ADJUSTMENT_HEADER, 2, {"non new reg"}, None, PROCESS_ADJUSTMENT_HEADER),
                (target_header, 3, {target_month_label.lower()}, None, short_target_header),
            ],
            row_grand=False,
        )
        pt10 = _build_pivot(
            "PivotTable_PIVOT_10",
            "D376",
            [SERVICE_DELIVERY_DIV_HEADER, DEPT_SD_HEADER],
            phase_header,
            [
                (YEAR_FAB_UPLOAD_HEADER, 1, None, year_excludes, YEAR_FAB_UPLOAD_HEADER),
                (target_header, 2, {after_month_label.lower()}, None, short_target_header),
            ],
            row_grand=False,
        )

        pivot_sheet.Range("A1").Value = (
            f" TARGET COMPLETE {month_name_upper} {year_value} "
            f"(Based on Dashboard 1 {month_name_upper} {year_value})"
        )
        pivot_sheet.Range("A3").Value = "New Registration "
        pivot_sheet.Range("A109").Value = "Non New Registration "
        pivot_sheet.Range("A178").Value = (
            f"STATUS ORDER WITH AGING FROM RFS COMMIT DATE {month_name_upper} {year_short}"
        )
        pivot_sheet.Range("A368").Value = f"TARGET AFTER {month_name_upper} {year_value}"
        pivot_sheet.Range("E439").Value = "TARGET NOT INPUTTED YET"

        def _column_letter(column_number: int) -> str:
            result = ""
            while column_number:
                column_number, remainder = divmod(column_number - 1, 26)
                result = chr(65 + remainder) + result
            return result

        def _header_column(pt, header_row: int, header_text: str) -> int | None:
            start_col = pt.TableRange2.Column
            end_col = start_col + pt.TableRange2.Columns.Count - 1
            target_normalized = str(header_text).strip().lower()
            for column_index in range(start_col, end_col + 1):
                value = pivot_sheet.Cells(header_row, column_index).Value
                if str(value).strip().lower() == target_normalized:
                    return column_index
            return None

        def _write_side_percentage(pt, page_field_count: int, left_count_column: str, mode: str) -> None:
            header_row = pt.TableRange2.Row + page_field_count + 2
            data_row_start = header_row + 1
            data_row_end = pt.TableRange2.Row + pt.TableRange2.Rows.Count - 1
            formula_col = pt.TableRange2.Column + pt.TableRange2.Columns.Count
            formula_letter = _column_letter(formula_col)
            cancel_col = _header_column(pt, header_row, "Cancel")
            so_col = _header_column(pt, header_row, "SO Complete")

            pivot_sheet.Cells(header_row, formula_col).Value = "Percentage of Completion (SO Complete, Cancel & Change Target)"
            for row_index in range(data_row_start, data_row_end + 1):
                if so_col is None:
                    continue
                so_letter = _column_letter(so_col)
                cancel_letter = _column_letter(cancel_col) if cancel_col is not None else None
                if mode == "sum_cancel_so":
                    if cancel_letter is not None:
                        formula = f"=({cancel_letter}{row_index}+{so_letter}{row_index})/{left_count_column}{row_index}"
                    else:
                        formula = f"={so_letter}{row_index}/{left_count_column}{row_index}"
                elif mode == "source_non_new":
                    if cancel_letter is not None:
                        formula = f"=({cancel_letter}{row_index}+{so_letter}{row_index}/{left_count_column}{row_index})"
                    else:
                        formula = f"={so_letter}{row_index}/{left_count_column}{row_index}"
                else:
                    formula = f"={so_letter}{row_index}/{left_count_column}{row_index}"
                pivot_sheet.Range(f"{formula_letter}{row_index}").Formula = formula
                pivot_sheet.Range(f"{formula_letter}{row_index}").NumberFormat = "0,00%"

        _write_side_percentage(pt4, 3, "C", "sum_cancel_so")
        _write_side_percentage(pt9, 3, "C", "source_non_new")
        _write_side_percentage(pt10, 2, "C", "so_only")

        target_wb.Save()
        print(f"[info] Added '{PIVOT_SHEET_NAME}' sheet to {output_path.name}")
    except Exception as exc:
        print(f"[warn] Could not add PIVOT sheet via COM: {exc}", file=sys.stderr)
    finally:
        if source_wb is not None:
            try:
                source_wb.Close(SaveChanges=False)
            except Exception:
                pass
        if target_wb is not None:
            try:
                target_wb.Close(SaveChanges=False)
            except Exception:
                pass
        if xl is not None:
            try:
                xl.Quit()
            except Exception:
                pass


def update_template_workbook_via_com(
    output_path: Path,
    generated_workbook_path: Path,
    progress_sheet_name: str,
    target_header: str,
    hidden_all_order_columns: set[str],
) -> None:
    """Replace data sheets in a copied template workbook and refresh its existing pivots."""
    try:
        import win32com.client  # type: ignore
    except ImportError:
        raise RuntimeError("pywin32 is required to update the template workbook and refresh pivots.")

    xl = None
    target_wb = None
    generated_wb = None
    try:
        xl = win32com.client.DispatchEx("Excel.Application")
        xl.Visible = False
        xl.DisplayAlerts = False

        target_wb = xl.Workbooks.Open(str(output_path.resolve()))
        generated_wb = xl.Workbooks.Open(str(generated_workbook_path.resolve()))

        def _sheet_by_name(workbook, sheet_name: str):
            for sheet in workbook.Sheets:
                if sheet.Name == sheet_name:
                    return sheet
            return None

        def _copy_used_range(source_sheet, target_sheet, preserve_first_table: bool = False) -> tuple[int, int]:
            xl_paste_all = -4104
            xl_paste_column_widths = 8
            source_used = source_sheet.UsedRange
            row_count = source_used.Rows.Count
            column_count = source_used.Columns.Count
            preserved_table_name = None

            if preserve_first_table:
                try:
                    if target_sheet.ListObjects.Count:
                        table = target_sheet.ListObjects(1)
                        preserved_table_name = table.Name
                        table.Resize(target_sheet.Range(target_sheet.Cells(1, 1), target_sheet.Cells(row_count, column_count)))
                except Exception:
                    preserved_table_name = None

            if not preserved_table_name:
                try:
                    while target_sheet.ListObjects.Count:
                        target_sheet.ListObjects(1).Delete()
                except Exception:
                    pass

            target_sheet.Cells.Clear()
            source_used.Copy()
            target_sheet.Range("A1").PasteSpecial(Paste=xl_paste_all)
            target_sheet.Range("A1").PasteSpecial(Paste=xl_paste_column_widths)
            xl.CutCopyMode = False

            if preserved_table_name and target_sheet.ListObjects.Count:
                try:
                    table = target_sheet.ListObjects(1)
                    table.Name = preserved_table_name
                    table.TableStyle = EXCEL_TABLE_STYLE_NAME
                except Exception:
                    pass

            try:
                target_sheet.Rows(1).RowHeight = source_sheet.Rows(1).RowHeight
            except Exception:
                pass

            for column_index in range(1, column_count + 1):
                try:
                    target_sheet.Columns(column_index).Hidden = source_sheet.Columns(column_index).Hidden
                except Exception:
                    pass

            return row_count, column_count

        def _ensure_table(sheet, row_count: int, column_count: int) -> None:
            if row_count < 1 or column_count < 1:
                return

            data_range = sheet.Range(sheet.Cells(1, 1), sheet.Cells(row_count, column_count))
            if sheet.ListObjects.Count:
                try:
                    table = sheet.ListObjects(1)
                    table.Resize(data_range)
                    table.TableStyle = EXCEL_TABLE_STYLE_NAME
                    return
                except Exception:
                    pass

            try:
                table = sheet.ListObjects.Add(1, data_range, None, 1)
                table.Name = "Table_ALL_ORDER"
                table.TableStyle = EXCEL_TABLE_STYLE_NAME
            except Exception:
                try:
                    if sheet.ListObjects.Count:
                        sheet.ListObjects(1).TableStyle = EXCEL_TABLE_STYLE_NAME
                except Exception:
                    pass

        def _hide_columns_by_header(sheet, row_count: int, column_count: int) -> None:
            if not hidden_all_order_columns:
                return

            for column_index in range(1, column_count + 1):
                header_value = sheet.Cells(1, column_index).Value
                if header_value in hidden_all_order_columns:
                    try:
                        sheet.Columns(column_index).Hidden = True
                    except Exception:
                        pass

        def _delete_on_progress_sheets() -> None:
            for sheet_index in range(target_wb.Sheets.Count, 0, -1):
                sheet = target_wb.Sheets(sheet_index)
                if str(sheet.Name).startswith(ON_PROGRESS_SHEET_PREFIX):
                    sheet.Delete()

        def _normalized_pivot_item_name(value: object) -> str:
            return str(value).strip().lower()

        def _hide_pivot_items(pivot_table, field_name: str, hidden_values: set[str]) -> None:
            hidden_normalized = {_normalized_pivot_item_name(value) for value in hidden_values}
            try:
                field = pivot_table.PivotFields(field_name)
                for item in field.PivotItems():
                    if _normalized_pivot_item_name(item.Name) in hidden_normalized:
                        try:
                            item.Visible = False
                        except Exception:
                            pass
            except Exception:
                pass

        def _restore_template_specific_pivot_filters(pivot_table) -> None:
            if str(pivot_table.Name) == "PivotTable13":
                _hide_pivot_items(
                    pivot_table,
                    PIVOT_TEMPLATE_SERVICE_DELIVERY_DIV_HEADER,
                    {"ICT Delivery Ops."},
                )

        def _column_letter(column_number: int) -> str:
            result = ""
            while column_number:
                column_number, remainder = divmod(column_number - 1, 26)
                result = chr(65 + remainder) + result
            return result

        def _pivot_header_column(pivot_table, header_row: int, header_text: str) -> int | None:
            start_col = pivot_table.TableRange2.Column
            end_col = start_col + pivot_table.TableRange2.Columns.Count - 1
            normalized_header = str(header_text).strip().lower()
            for column_index in range(start_col, end_col + 1):
                value = pivot_sheet.Cells(header_row, column_index).Value
                if str(value).strip().lower() == normalized_header:
                    return column_index
            return None

        percentage_completion_header = "Percentage of Completion (SO Complete, Cancel & Change Target)"
        target_after_percentage_format_source = None
        target_after_percentage_format_sheet = None
        target_after_percentage_column_width = None

        def _capture_target_after_percentage_format():
            nonlocal target_after_percentage_format_sheet, target_after_percentage_column_width
            if pivot_sheet is None:
                return None

            header_row = 377
            normalized_header = percentage_completion_header.lower()
            for column_index in range(4, 15):
                value = pivot_sheet.Cells(header_row, column_index).Value
                if str(value).strip().lower() == normalized_header:
                    target_after_percentage_column_width = pivot_sheet.Columns(column_index).ColumnWidth
                    source_range = pivot_sheet.Range(
                        pivot_sheet.Cells(header_row, column_index),
                        pivot_sheet.Cells(387, column_index),
                    )
                    try:
                        target_after_percentage_format_sheet = target_wb.Worksheets.Add()
                        target_after_percentage_format_sheet.Name = "__codex_pct_fmt"
                        target_after_percentage_format_sheet.Visible = 0
                        source_range.Copy()
                        target_after_percentage_format_sheet.Range("A1").PasteSpecial(Paste=-4122)  # xlPasteFormats
                        xl.CutCopyMode = False
                        return target_after_percentage_format_sheet.Range("A1:A11")
                    except Exception:
                        return source_range
            return None

        def _repair_target_after_completion_formula(pivot_table) -> None:
            if pivot_table.TableRange2.Row != 373:
                return

            page_field_count = 2
            header_row = pivot_table.TableRange2.Row + page_field_count + 2
            data_row_start = header_row + 1
            data_row_end = pivot_table.TableRange2.Row + pivot_table.TableRange2.Rows.Count - 1
            formula_col = pivot_table.TableRange2.Column + pivot_table.TableRange2.Columns.Count
            formula_letter = _column_letter(formula_col)
            cancel_col = _pivot_header_column(pivot_table, header_row, "Cancel")
            so_col = _pivot_header_column(pivot_table, header_row, "SO Complete")
            if so_col is None:
                return

            so_letter = _column_letter(so_col)
            cancel_letter = _column_letter(cancel_col) if cancel_col is not None else None
            pivot_sheet.Cells(header_row, formula_col).Value = percentage_completion_header
            for row_index in range(data_row_start, data_row_end + 1):
                if cancel_letter is not None:
                    formula = f"=({so_letter}{row_index}+{cancel_letter}{row_index}/C{row_index})"
                else:
                    formula = f"={so_letter}{row_index}/C{row_index}"
                pivot_sheet.Range(f"{formula_letter}{row_index}").Formula = formula
                pivot_sheet.Range(f"{formula_letter}{row_index}").NumberFormat = "0.00%"

            if target_after_percentage_format_source is not None:
                try:
                    target_after_percentage_format_source.Copy()
                    pivot_sheet.Range(
                        pivot_sheet.Cells(header_row, formula_col),
                        pivot_sheet.Cells(data_row_end, formula_col),
                    ).PasteSpecial(Paste=-4122)  # xlPasteFormats
                    xl.CutCopyMode = False
                    if target_after_percentage_column_width is not None:
                        pivot_sheet.Columns(formula_col).ColumnWidth = target_after_percentage_column_width
                except Exception:
                    pass

        data_sheet = _sheet_by_name(target_wb, VLOOKUP_SHEET_NAME)
        generated_data_sheet = _sheet_by_name(generated_wb, VLOOKUP_SHEET_NAME)
        generated_progress_sheet = _sheet_by_name(generated_wb, progress_sheet_name)
        pivot_sheet = _sheet_by_name(target_wb, PIVOT_SHEET_NAME)

        if data_sheet is None or generated_data_sheet is None or pivot_sheet is None:
            raise RuntimeError("Could not find required sheets for template update.")

        target_after_percentage_format_source = _capture_target_after_percentage_format()

        row_count, column_count = _copy_used_range(generated_data_sheet, data_sheet, preserve_first_table=True)
        _ensure_table(data_sheet, row_count, column_count)
        _hide_columns_by_header(data_sheet, row_count, column_count)

        _delete_on_progress_sheets()
        progress_sheet = None
        if generated_progress_sheet is not None:
            progress_sheet = target_wb.Worksheets.Add(None, data_sheet)
            progress_sheet.Name = progress_sheet_name
            _copy_used_range(generated_progress_sheet, progress_sheet)

        def _add_on_progress_pivots(shared_pivot_cache) -> None:
            if progress_sheet is None or not target_header:
                return

            xl_count = -4112
            xl_page_field = 3
            xl_row_field = 1
            xl_tabular_row = 1
            progress_sheet.Cells.Clear()
            date_match = re.search(r"1\s+([A-Za-z]+)\s+(\d{4})", target_header)
            month_full = date_match.group(1) if date_match else ""
            target_month_label = f"Target {month_full}" if month_full else ""
            source_month_label = f"This {month_full}" if month_full else ""

            def _build_progress_pivot(
                top_left_cell,
                table_name: str,
                visible_row_items: list[str] | None = None,
                exclude_phase_values: set[str] | None = None,
            ) -> None:
                pt = shared_pivot_cache.CreatePivotTable(
                    TableDestination=progress_sheet.Range(top_left_cell),
                    TableName=table_name,
                )
                pt.ManualUpdate = True

                try:
                    year_field = pt.PivotFields(YEAR_FAB_UPLOAD_HEADER)
                    year_field.Orientation = xl_page_field
                    year_field.EnableMultiplePageItems = True
                    for item in year_field.PivotItems():
                        if is_excluded_year_fab_upload_value(item.Value):
                            item.Visible = False
                except Exception:
                    pass

                try:
                    pt.PivotFields(PROCESS_ADJUSTMENT_HEADER).Orientation = xl_page_field
                except Exception:
                    pass

                try:
                    phase_field = pt.PivotFields(PHASE_HEADER_PREFIX)
                    phase_field.Orientation = xl_page_field
                    if exclude_phase_values:
                        phase_field.EnableMultiplePageItems = True
                        excluded_lower = {value.lower() for value in exclude_phase_values}
                        for item in phase_field.PivotItems():
                            if str(item.Value).strip().lower() in excluded_lower:
                                item.Visible = False
                except Exception:
                    pass

                try:
                    row_field = pt.PivotFields(target_header)
                    row_field.Orientation = xl_row_field
                    row_field.LayoutForm = xl_tabular_row
                    if visible_row_items:
                        visible_lower = {value.lower() for value in visible_row_items}
                        for item in row_field.PivotItems():
                            item_name = str(item.Name).strip()
                            compare_name = target_month_label if item_name == source_month_label else item_name
                            item.Visible = compare_name.lower() in visible_lower
                        for position, label in enumerate(visible_row_items, start=1):
                            for item in row_field.PivotItems():
                                item_name = str(item.Name).strip()
                                compare_name = target_month_label if item_name == source_month_label else item_name
                                if compare_name.lower() == label.lower():
                                    item.Position = position
                                    break
                except Exception:
                    pass

                try:
                    count_field = pt.AddDataField(pt.PivotFields(QUO_HEADER), "Count of quo", xl_count)
                    count_field.NumberFormat = "0"
                except Exception:
                    pass

                try:
                    pt.TableStyle2 = "PivotStyleMedium2"
                except Exception:
                    pass

                pt.ManualUpdate = False
                pt.RefreshTable()

                if source_month_label and target_month_label:
                    try:
                        row_field = pt.PivotFields(target_header)
                        for item in row_field.PivotItems():
                            if str(item.Name).strip() == source_month_label:
                                item.Caption = target_month_label
                    except Exception:
                        pass

            _build_progress_pivot(
                "A1",
                f"PivotTable_{progress_sheet_name}_1",
                visible_row_items=[target_month_label, f"After {month_full}"] if month_full else None,
            )
            _build_progress_pivot(
                "D1",
                f"PivotTable_{progress_sheet_name}_2",
                visible_row_items=[f"Before {month_full}", "Target Not Yet Inputted"] if month_full else None,
                exclude_phase_values={"cancel", "so complete"},
            )

        generated_wb.Close(SaveChanges=False)
        generated_wb = None

        data_range = data_sheet.Range(data_sheet.Cells(1, 1), data_sheet.Cells(row_count, column_count))
        source_data = f"'{VLOOKUP_SHEET_NAME}'!{data_range.Address}"
        refreshed_count = 0
        pivot_tables = pivot_sheet.PivotTables()
        for pivot_index in range(1, pivot_tables.Count + 1):
            pivot_table = pivot_tables(pivot_index)
            try:
                pivot_table.SaveData = True
                pivot_table.RefreshTable()
                _restore_template_specific_pivot_filters(pivot_table)
                pivot_table.RefreshTable()
                _repair_target_after_completion_formula(pivot_table)
                refreshed_count += 1
            except Exception as exc:
                print(
                    f"[warn] Could not refresh PivotTable '{pivot_table.Name}' on '{pivot_sheet.Name}': {exc}",
                    file=sys.stderr,
                )

        if progress_sheet is not None:
            on_progress_cache = pivot_sheet.PivotTables()(1).PivotCache()
            _add_on_progress_pivots(on_progress_cache)

        if target_after_percentage_format_sheet is not None:
            try:
                target_after_percentage_format_sheet.Delete()
            except Exception:
                pass

        target_wb.Save()
        print(f"[info] Refreshed {refreshed_count} template PivotTables in {output_path.name}")
    except Exception as exc:
        print(f"[warn] Could not update template workbook via COM: {exc}", file=sys.stderr)
        raise
    finally:
        if generated_wb is not None:
            try:
                generated_wb.Close(SaveChanges=False)
            except Exception:
                pass
        if target_wb is not None:
            try:
                target_wb.Close(SaveChanges=False)
            except Exception:
                pass
        if xl is not None:
            try:
                xl.Quit()
            except Exception:
                pass


def apply_target_so_complete_date_filter_format(worksheet, df: pd.DataFrame) -> None:
    header_positions = {
        str(cell.value).strip(): cell.column
        for cell in worksheet[1]
        if cell.value is not None and str(cell.value).strip()
    }
    target_header = next(
        (header for header in TARGET_SO_COMPLETE_DATE_HEADERS if header in df.columns and header in header_positions),
        None,
    )
    if target_header is None:
        return

    target_dates = pd.to_datetime(df[target_header], errors="coerce", format="mixed")
    target_column = header_positions[target_header]

    for row_index, target_date in enumerate(target_dates, start=2):
        if pd.isna(target_date):
            continue

        cell = worksheet.cell(row=row_index, column=target_column)
        cell.value = target_date.to_pydatetime()
        cell.number_format = EXCEL_DATE_DISPLAY_FORMAT


def apply_worksheet_presentation(worksheet) -> None:
    add_excel_table(worksheet)
    autofit_worksheet_columns(worksheet)
    style_header_row(worksheet)


def add_excel_table(worksheet) -> None:
    table_name = re.sub(r"[^A-Za-z0-9_]", "_", f"Table_{worksheet.title}") or "Table_1"
    if table_name[0].isdigit():
        table_name = f"Table_{table_name}"

    table = Table(displayName=table_name, ref=worksheet.dimensions)
    table.tableStyleInfo = TableStyleInfo(
        name=EXCEL_TABLE_STYLE_NAME,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)


def autofit_worksheet_columns(worksheet) -> None:
    for column_cells in worksheet.iter_cols():
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            value = cell.value
            if value is None:
                continue

            if isinstance(value, datetime):
                rendered = value.strftime("%Y-%m-%d")
            elif isinstance(value, date):
                rendered = value.strftime("%Y-%m-%d")
            else:
                rendered = str(value)

            max_length = max(max_length, len(rendered))

        worksheet.column_dimensions[column_letter].width = max_length + 2


def style_header_row(worksheet) -> None:
    worksheet.row_dimensions[1].height = EXCEL_HEADER_ROW_HEIGHT_POINTS
    for cell in worksheet[1]:
        fill_color = header_fill_color(cell.value)
        font_color = EXCEL_RED_FONT_COLOR if fill_color == EXCEL_YELLOW_HEADER_FILL else EXCEL_HEADER_FONT_COLOR
        cell.font = Font(color=font_color, bold=True)
        if fill_color is not None:
            cell.fill = PatternFill(fill_type="solid", fgColor=fill_color)


def header_fill_color(header_value: object) -> str | None:
    normalized = normalize_header_key(header_value)
    if not normalized:
        return None

    if (
        normalized in {
            normalize_header_key(DEPT_SD_HEADER),
            normalize_header_key(SERVICE_DELIVERY_DIV_HEADER),
            normalize_header_key(STATUS_ORDER_HEADER),
            normalize_header_key(GROUP_SALES_HEADER),
            normalize_header_key(DIVISION_SALES_HEADER),
            normalize_header_key(SEGMENT_SALES_HEADER),
            normalize_header_key(SALES_HEADER),
            normalize_header_key("Target SO Completion Date"),
            normalize_header_key("Target SO Complete Date"),
        }
        or normalized.startswith("targetdeterminedas")
        or normalized.startswith("targetdeteminedas")
    ):
        return EXCEL_RED_HEADER_FILL

    if (
        normalized in {
            normalize_header_key(PROCESS_ADJUSTMENT_HEADER),
            normalize_header_key(AGING_OF_RFS_HEADER),
            normalize_header_key(YEAR_FAB_UPLOAD_HEADER),
        }
        or normalized.startswith("phase") and normalized != normalize_header_key(PHASE_HEADER_PREFIX)
    ):
        return EXCEL_YELLOW_HEADER_FILL

    return None


def normalize_header_key(header_value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(header_value).lower())


def first_target_header(df: pd.DataFrame) -> str:
    return next((column for column in df.columns if str(column).startswith("TARGET")), "")


def env_flag_enabled(name: str) -> bool:
    return os.getenv(name, "").strip().lower() in {"1", "true", "yes"}


def resolve_csv_files(input_path: Path) -> list[Path]:
    if input_path.is_file():
        return [input_path]

    if not input_path.is_dir():
        raise FileNotFoundError(f"Input path does not exist: {input_path}")

    files = sorted(input_path.glob("*.csv"))
    if not files:
        raise FileNotFoundError(f"No CSV files found in: {input_path}")

    return files


def default_output_path(input_path: Path, combine: bool) -> Path:
    if input_path.is_file():
        return input_path.parent / output_filename_from_csv_path(input_path)

    output_dir = input_path.parent / DEFAULT_OUTPUT_DIR

    if combine:
        return output_dir / "combined.xlsx"

    return output_dir


def resolve_output_path(input_path: Path, csv_files: list[Path], args: argparse.Namespace) -> Path:
    output = args.output or default_output_path(input_path, args.combine)

    if input_path.is_file() and (output.exists() and output.is_dir()):
        return output / output_filename_from_csv_path(input_path)

    if input_path.is_dir() and args.combine and output.suffix.lower() != ".xlsx":
        return output / "combined.xlsx"

    if input_path.is_dir() and not args.combine and output.suffix.lower() == ".xlsx":
        raise ValueError("Use --combine when writing multiple CSV files to one .xlsx output file.")

    if len(csv_files) == 1 and input_path.is_dir() and args.combine and output.suffix.lower() != ".xlsx":
        return output / "combined.xlsx"

    return output


def build_options(args: argparse.Namespace, output_path: Path) -> ConvertOptions:
    return ConvertOptions(
        output=output_path,
        delimiter=args.delimiter,
        encoding=args.encoding,
        normalize_headers=not args.no_normalize_headers,
        keep_empty=args.keep_empty,
        drop_empty_columns=args.drop_empty_columns,
        dedupe=args.dedupe,
        infer_types=args.infer_types,
        combine=args.combine,
        refresh_template=not args.skip_template_refresh,
    )


def convert_csv_files(input_path: Path, csv_files: list[Path], output_path: Path, options: ConvertOptions) -> None:
    if len(csv_files) == 1 and input_path.is_file():
        convert_one(csv_files[0], output_path, options)
        return

    convert_many(csv_files, options)


def convert_one(csv_path: Path, output_path: Path, options: ConvertOptions) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    df = clean_dataframe(read_csv(csv_path, options), options, csv_path)
    lookup_workbook = resolve_vlookup_workbook()
    mappings = load_lookup_mappings(lookup_workbook)
    progress_sheet_name = on_progress_sheet_name(csv_path)

    with tempfile.TemporaryDirectory(prefix="csv_to_excel_", dir=output_path.parent) as temp_dir:
        generated_workbook_path = Path(temp_dir) / "today-data.xlsx"
        with pd.ExcelWriter(generated_workbook_path, engine="openpyxl") as writer:
            output_df, hidden_columns = write_all_order_sheet_for_template(
                writer,
                df,
                csv_path,
                mappings,
            )
            write_on_progress_sheet(writer, output_df, csv_path, progress_sheet_name)

        target_header = first_target_header(output_df)
        if options.refresh_template:
            shutil.copy2(lookup_workbook, output_path)
            update_template_workbook_via_com(
                output_path,
                generated_workbook_path,
                progress_sheet_name,
                target_header,
                hidden_columns,
            )
        else:
            shutil.copy2(generated_workbook_path, output_path)

    print(f"Wrote {output_path}")


def convert_many(csv_files: list[Path], options: ConvertOptions) -> None:
    mappings = load_lookup_mappings(resolve_vlookup_workbook())
    if options.combine:
        options.output.parent.mkdir(parents=True, exist_ok=True)
        used_sheet_names: set[str] = set()
        with pd.ExcelWriter(options.output, engine="openpyxl") as writer:
            for csv_path in csv_files:
                df = clean_dataframe(read_csv(csv_path, options), options, csv_path)
                sheet_name = sanitize_sheet_name(csv_path, used_sheet_names)
                output_df = write_excel_sheet(
                    writer,
                    df,
                    sheet_name,
                    mappings,
                )
                ps_name = on_progress_sheet_name(csv_path, used_sheet_names)
                target_hdr = first_target_header(output_df)
                write_on_progress_sheet(writer, output_df, csv_path, ps_name)
        print(f"Wrote {options.output}")
        if target_hdr:
            add_pivot_tables_via_com(options.output, ps_name, target_hdr)
        return

    options.output.mkdir(parents=True, exist_ok=True)
    for csv_path in csv_files:
        convert_one(csv_path, options.output / output_filename_from_csv_path(csv_path), options)


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Clean CSV data and export it to Excel (.xlsx).",
    )
    parser.add_argument(
        "input",
        nargs="?",
        type=Path,
        default=Path(DEFAULT_INPUT_DIR),
        help=f"CSV file or directory containing CSV files. Defaults to .\\{DEFAULT_INPUT_DIR}.",
    )
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        help=f"Output .xlsx file or output directory. Defaults to .\\{DEFAULT_OUTPUT_DIR} for directory input.",
    )
    parser.add_argument("--delimiter", help="CSV delimiter. Auto-detected when omitted.")
    parser.add_argument("--encoding", help="CSV encoding. Auto-detected when omitted.")
    parser.add_argument(
        "--keep-empty",
        action="store_true",
        help="Keep fully empty rows. Empty columns are kept by default.",
    )
    parser.add_argument(
        "--drop-empty-columns",
        action="store_true",
        help="Drop columns where all values are empty.",
    )
    parser.add_argument(
        "--no-normalize-headers",
        action="store_true",
        help="Keep original column names instead of converting underscores to readable headers.",
    )
    parser.add_argument(
        "--dedupe",
        action="store_true",
        help="Remove duplicate rows after cleaning.",
    )
    parser.add_argument(
        "--infer-types",
        action="store_true",
        help="Infer numeric and date columns. By default, values are preserved as text.",
    )
    parser.add_argument(
        "--combine",
        action="store_true",
        help="When input is a directory, write all CSV files to one workbook with one sheet per file.",
    )
    parser.add_argument(
        "--skip-template-refresh",
        action="store_true",
        default=env_flag_enabled("CSV_TO_EXCEL_SKIP_TEMPLATE_REFRESH"),
        help="Skip Windows Excel COM template refresh and write the portable generated workbook.",
    )
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv or sys.argv[1:])
    input_path = args.input.resolve()

    try:
        csv_files = resolve_csv_files(input_path)
        output_path = resolve_output_path(input_path, csv_files, args).resolve()
        options = build_options(args, output_path)
        convert_csv_files(input_path, csv_files, output_path, options)
    except Exception as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 1

    return 0


if __name__ == "__main__":
    raise SystemExit(main())

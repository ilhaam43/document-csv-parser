"""Capture and email the Report 3 iPhone-tracking PIVOT sheet."""

from __future__ import annotations

import argparse
from pathlib import Path

from send_report_1_email import excel_range_to_png, send_email


def report_3_target_complete_range(workbook_path: Path) -> str:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    sheet = load_workbook(workbook_path, read_only=True, data_only=False)["PIVOT"]
    target_row = None
    next_section_row = sheet.max_row + 1
    legend_columns = []
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            value = str(cell.value or "").strip().upper()
            if "TARGET COMPLETE" in value and target_row is None:
                target_row = cell.row
            if target_row is not None and cell.row > target_row:
                if "DELAY COMPLETION" in value or "TARGET AFTER" in value:
                    next_section_row = min(next_section_row, cell.row)
                if value.startswith(("GREEN", "YELLOW", "RED")):
                    legend_columns.append(cell.column)
    if target_row is None:
        raise RuntimeError("Could not find TARGET COMPLETE on the Report 3 PIVOT sheet.")
    end_column = max((cell.column for row in sheet.iter_rows(min_row=target_row, max_row=next_section_row - 1) for cell in row if cell.value is not None), default=13)
    if legend_columns:
        end_column = min(end_column, min(legend_columns) - 1)
    return f"A{target_row}:{get_column_letter(max(1, end_column))}{next_section_row - 1}"


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, required=True)
    parser.add_argument("--image", type=Path)
    parser.add_argument("--endpoint", default=None)
    parser.add_argument("--to", default="ilhaam.akmal@lintasarta.co.id")
    parser.add_argument("--subject", default="Daily Tracking Report 3 iPhone")
    parser.add_argument("--message", default='<p>Daily Tracking Report 3 iPhone</p><p><img src="cid:report-3-pivot" alt="Report 3 iPhone PIVOT"></p>')
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()
    workbook = args.workbook.resolve()
    if not workbook.exists():
        parser.error(f"Workbook not found: {workbook}")
    image = (args.image or workbook.with_name(f"{workbook.stem} - report-3-pivot.png")).resolve()
    cell_range = report_3_target_complete_range(workbook)
    excel_range_to_png(workbook, image, "PIVOT", cell_range)
    print(f"Created image: {image} ({cell_range})")
    if args.dry_run:
        print("Dry run: SMTP request skipped.")
        return 0
    send_email(args.endpoint or "http://10.34.144.197/secm-portal/smtp/api_send_email", args.to, args.subject, args.message, image, timeout=60, content_id="report-3-pivot")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

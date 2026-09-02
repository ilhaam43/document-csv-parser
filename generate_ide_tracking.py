#!/usr/bin/env python3
"""Generate a Daily Tracking IDE workbook from the current IDE dashboard."""

from __future__ import annotations

import argparse
import calendar
import numbers
import os
import re
import shutil
import sys
import tempfile
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Iterable

import pandas as pd
from dateutil import parser as date_parser
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

from excel_pivot_layout import normalize_layout_text, resize_dynamic_pivot_section_banners


DEFAULT_INPUT_DIR = "input-ide"
DEFAULT_REFERENCE_DIR = "vlookup-ide"
DEFAULT_COLLABS_DIR = "collabs-ide"
DEFAULT_OUTPUT_DIR = "output-ide"

RAW_SHEET_NAME = "ICT ORDER 2026"
RAW_SOURCE_SHEET_NAME = "source"
ALL_ORDER_SHEET_NAME = "ALL ORDER"
PIVOT_SHEET_NAME = "PIVOT"
TABLE_NAME = "Table47"
TABLE_STYLE_NAME = "TableStyleMedium2"

QUOTE_ID_HEADER = "QUOTE ID"
ORDER_TYPE_HEADER = "ORDER TYPE"
OTC_HEADER = "OTC"
MRC_HEADER = "MRC"
FAB_UPLOAD_HEADER = "FAB UPLOAD DATE"
YEAR_FAB_UPLOAD_HEADER = "YEAR FAB UPLOAD"
NEW_RFS_INITIAL_HEADER = "NEW RFS INITIAL"
ACTUAL_RFS_DATE_HEADER = "ACTUAL RFS DATE"
AGING_OF_RFS_HEADER = "Aging Of RFS"
STATUS_ORDER_HEADER = "STATUS ORDER"
AOS_SENT_DATE_HEADER = "AOS SENT DATE"
ON_TIME_DELIVERY_HEADER = "ON TIME DELIVERY"
STATUS_DELIVERY_HEADER = "STATUS DELIVERY"
PRE_INSTALLATION_START_HEADER = "Pre-Installation Start Date"
PRE_INSTALLATION_AGING_HEADER = "Pre-Installation Aging"
PRE_INSTALLATION_RANGE_HEADER = "Range Aging On Pre-Installation"
DEPT_HEADER = "DEPT."
PM_HEADER = "PM"
APM_HEADER = "APM"

DATE_HEADERS = (
    "QUOTE DATE",
    "FAB SIGN DATE",
    FAB_UPLOAD_HEADER,
    NEW_RFS_INITIAL_HEADER,
    "CASH DATE",
    "SO DATE",
    "SERVICE READY",
    ACTUAL_RFS_DATE_HEADER,
    AOS_SENT_DATE_HEADER,
    "AOS SIGN DATE",
    "SO COMPLETION DATE",
    PRE_INSTALLATION_START_HEADER,
)

# These dashboard fields are converted by the manual Excel workflow under a
# United States regional setting. Other date fields retain day-first parsing.
MONTH_FIRST_TEXT_DATE_HEADERS = (
    FAB_UPLOAD_HEADER,
    NEW_RFS_INITIAL_HEADER,
    ACTUAL_RFS_DATE_HEADER,
)
FIRST_MATCH_CURRENT_DATE_HEADERS = (
    ACTUAL_RFS_DATE_HEADER,
)
MONTH_OPENING_CARRY_DATE_HEADERS = (
    NEW_RFS_INITIAL_HEADER,
    ACTUAL_RFS_DATE_HEADER,
)
RAW_SOURCE_QUOTE_ID_COLUMN = 4  # D
RAW_SOURCE_NEW_RFS_COLUMN = 38  # AL

DATE_ONLY_FORMAT = "mm/dd/yy"
DATE_TIME_FORMAT = "mm/dd/yy h:mm AM/PM"
YEAR_TEXT_FORMAT = "@"
FONT_NAME = "Indosat Regular"
FONT_SIZE = 11
MAX_COLUMN_WIDTH = 45
MIN_COLUMN_WIDTH = 10
PIVOT_PERCENTAGE_MIN_WIDTH = 24
PIVOT_SECTION_TITLE_MARKERS = (
    "target complete",
    "order aging start from pre-installation status",
    "ontime delivery",
    "status order with aging from rfs commit date",
    "delay completion",
    "target after",
    "target not inputted",
)
PIVOT_SECTION_SIDE_HEADER_MARKERS = ("percentage of completion",)
UNKNOWN_YEAR = "1900"
EXCEL_ZERO_DATE_SERIAL = 0
EXCEL_NA_ERROR = "#N/A"
MIN_VALID_DATE_YEAR = 1900
MAX_VALID_DATE_YEAR = 2100
MISSING_TEXT = {"", "nan", "none", "null", "n/a", "#n/a", "nat"}
MAX_OPERATIONAL_FUTURE_YEARS = 5
REPORT_MONTH_NUMBERS = {
    **{month_name.casefold(): month_number for month_number, month_name in enumerate(calendar.month_name) if month_name},
    **{month_name.casefold(): month_number for month_number, month_name in enumerate(calendar.month_abbr) if month_name},
    "sept": 9,
    "januari": 1,
    "februari": 2,
    "maret": 3,
    "april": 4,
    "mei": 5,
    "juni": 6,
    "juli": 7,
    "agustus": 8,
    "september": 9,
    "oktober": 10,
    "november": 11,
    "desember": 12,
    "jan": 1,
    "feb": 2,
    "mar": 3,
    "apr": 4,
    "mei": 5,
    "jun": 6,
    "jul": 7,
    "agu": 8,
    "agt": 8,
    "sep": 9,
    "okt": 10,
    "nov": 11,
    "des": 12,
}
PREVIOUS_FALLBACK_HEADERS = (
    DEPT_HEADER,
    PM_HEADER,
    APM_HEADER,
    STATUS_DELIVERY_HEADER,
)

STATUS_NORMALIZATION = {
    "pre-installation": "04-Pre-Installation",
    "04-pre-installation": "04-Pre-Installation",
    "installation": "06-Installation",
    "06-installation": "06-Installation",
    "waiting aos": "07-Waiting AOS/UAT On Hold",
    "07-waiting aos/uat on hold": "07-Waiting AOS/UAT On Hold",
    "so complete": "SO Complete",
    "cancel": "Cancel",
}


@dataclass(frozen=True)
class InputFiles:
    raw: Path
    previous: Path
    collabs: Path


@dataclass(frozen=True)
class BuildResult:
    dataframe: pd.DataFrame
    invalid_dates: dict[str, list[str]]
    collabs_fallback_count: int
    zero_fallback_count: int
    duplicate_quote_ids: tuple[str, ...]
    duplicate_rows_preserved: int
    date_fallback_count: int
    field_fallback_count: int


@dataclass(frozen=True)
class PivotPercentageBlock:
    pivot_name: str
    header_row_offset: int
    original_header_row: int
    original_end_row: int
    scratch_start_row: int
    formula: str
    formula_row: int
    pivot_header_columns: dict[int, str]


@dataclass(frozen=True)
class TargetPivotPageFilter:
    pivot_name: str
    position: int
    selection_kinds: tuple[str, ...]
    show_all: bool


@dataclass(frozen=True)
class CompletionThresholds:
    week: int
    green_percent: int
    yellow_percent: int
    red_percent: int


def is_missing(value: object) -> bool:
    if value is None or value is pd.NA:
        return True
    try:
        if pd.isna(value):
            return True
    except (TypeError, ValueError):
        pass
    return isinstance(value, str) and value.strip().lower() in MISSING_TEXT


def is_excel_zero_date(value: object) -> bool:
    """Return whether a value represents Excel's serial-zero date sentinel."""
    if isinstance(value, time):
        return value == time.min
    if isinstance(value, numbers.Real) and not isinstance(value, bool):
        return float(value) == EXCEL_ZERO_DATE_SERIAL
    if isinstance(value, (pd.Timestamp, datetime, date)):
        return (value.year, value.month, value.day) == (1899, 12, 30)
    return isinstance(value, str) and value.strip() in {"0", "0.0", "00:00:00"}


def normalize_quote_id(value: object) -> str:
    if is_missing(value):
        return ""
    text = str(value).strip()
    return text[:-2] if text.endswith(".0") and text[:-2].isdigit() else text


def normalize_header(value: object) -> str:
    return re.sub(r"\s+", " ", str(value).replace("\ufeff", "").strip())


def normalize_status(value: object) -> object:
    if is_missing(value):
        return pd.NA
    text = str(value).strip()
    return STATUS_NORMALIZATION.get(text.lower(), text)


def parse_report_date(value: str) -> date:
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError as exc:
        raise argparse.ArgumentTypeError("Report date must use YYYY-MM-DD.") from exc


def date_from_filename(path: Path) -> date | None:
    numeric_match = re.search(r"(?<!\d)(20\d{6})(?!\d)", path.stem)
    if numeric_match:
        try:
            return datetime.strptime(numeric_match.group(1), "%Y%m%d").date()
        except ValueError:
            pass

    named_match = re.search(
        r"(?<!\d)(\d{1,2})\s+([A-Za-z]+)\s+(20\d{2})(?!\d)",
        path.stem,
        flags=re.IGNORECASE,
    )
    if named_match:
        day_text, month_text, year_text = named_match.groups()
        month_number = REPORT_MONTH_NUMBERS.get(month_text.casefold())
        if month_number is not None:
            try:
                return date(int(year_text), month_number, int(day_text))
            except ValueError:
                pass

    separated_numeric_match = re.search(
        r"(?<!\d)(\d{1,2})[\s._-]+(\d{1,2})[\s._-]+((?:20)?\d{2})(?!\d)",
        path.stem,
    )
    if separated_numeric_match:
        day_text, month_text, year_text = separated_numeric_match.groups()
        if len(year_text) == 2:
            year_text = f"20{year_text}"
        try:
            return date(int(year_text), int(month_text), int(day_text))
        except ValueError:
            pass

    return None


def determine_report_date(explicit_date: date | None, raw: Path) -> date:
    if explicit_date is not None:
        return explicit_date

    raw_date = date_from_filename(raw)
    if raw_date is not None and raw.stem.casefold().startswith("ide dashboard"):
        return raw_date

    raise ValueError(
        "Raw workbook filename must include the report date, for example: "
        "'IDE DASHBOARD 15 August 2026.xlsx'."
    )


def resolve_single_file(path: Path, suffix: str, description: str) -> Path:
    path = path.resolve()
    if path.is_file():
        if path.suffix.lower() != suffix:
            raise ValueError(f"{description} must be a {suffix} file: {path}")
        return path
    if not path.is_dir():
        raise FileNotFoundError(f"{description} path does not exist: {path}")

    candidates = sorted(
        item
        for item in path.glob(f"*{suffix}")
        if item.is_file() and not item.name.startswith("~$")
    )
    if not candidates:
        raise FileNotFoundError(f"No {suffix} file found for {description} in: {path}")
    if len(candidates) > 1:
        names = ", ".join(item.name for item in candidates)
        raise ValueError(
            f"Multiple files found for {description}. Keep exactly one file in {path}: {names}"
        )
    return candidates[0]


def resolve_inputs(args: argparse.Namespace) -> InputFiles:
    return InputFiles(
        raw=resolve_single_file(args.input, ".xlsx", "IDE raw workbook"),
        previous=resolve_single_file(args.reference, ".xlsx", "previous IDE workbook"),
        collabs=resolve_single_file(args.collabs, ".csv", "collabs fallback"),
    )


def output_filename(report_date: date) -> str:
    month_name = calendar.month_name[report_date.month]
    return f"Daily Tracking IDE {report_date.day} {month_name} {report_date.year}.xlsx"


def resolve_output_path(output: Path, report_date: date) -> Path:
    output = output.resolve()
    if output.suffix.lower() == ".xlsx":
        return output
    return output / output_filename(report_date)


def target_header(report_date: date) -> str:
    return f"TARGET  Detemined as 1 {calendar.month_abbr[report_date.month]} {report_date:%y}"


def previous_status_header(previous: Path, report_date: date) -> str:
    previous_date = date_from_filename(previous) or (report_date - timedelta(days=1))
    return (
        f"STATUS DELIVERY {previous_date.day} "
        f"{calendar.month_name[previous_date.month]} {previous_date.year}"
    )


def read_raw_dataframe(raw_path: Path) -> pd.DataFrame:
    raw = pd.read_excel(
        raw_path,
        sheet_name=RAW_SHEET_NAME,
        header=0,
        skiprows=[1],
        dtype=object,
    )
    raw.columns = [normalize_header(column) for column in raw.columns]
    if QUOTE_ID_HEADER not in raw.columns:
        raise ValueError(f"Raw sheet is missing required column: {QUOTE_ID_HEADER}")

    raw_workbook = load_workbook(raw_path, read_only=True, data_only=True, keep_links=False)
    try:
        raw_worksheet = raw_workbook[RAW_SHEET_NAME]
        for dataframe_row, worksheet_row in enumerate(
            raw_worksheet.iter_rows(min_row=3, max_col=len(raw.columns))
        ):
            if dataframe_row >= len(raw):
                break
            for column_index, cell in enumerate(worksheet_row):
                if cell.data_type == "e":
                    raw.iat[dataframe_row, column_index] = str(cell.value)

        if NEW_RFS_INITIAL_HEADER in raw.columns and RAW_SOURCE_SHEET_NAME in raw_workbook.sheetnames:
            quote_ids = raw[QUOTE_ID_HEADER].map(normalize_quote_id)
            string_new_rfs = raw[NEW_RFS_INITIAL_HEADER].map(
                lambda value: isinstance(value, str) and not is_missing(value)
            )
            required_quote_ids = set(quote_ids.loc[string_new_rfs])
            source_lookup = load_source_new_rfs_lookup(
                raw_workbook[RAW_SOURCE_SHEET_NAME],
                required_quote_ids,
            )
            has_source_value = string_new_rfs & quote_ids.isin(source_lookup)
            raw.loc[has_source_value, NEW_RFS_INITIAL_HEADER] = quote_ids.loc[
                has_source_value
            ].map(source_lookup)
    finally:
        raw_workbook.close()

    quote_ids = raw[QUOTE_ID_HEADER].map(normalize_quote_id)
    raw = raw.loc[quote_ids.ne("")].copy()
    raw[QUOTE_ID_HEADER] = quote_ids.loc[raw.index]
    return raw.reset_index(drop=True)


def read_previous_dataframe(previous_path: Path) -> pd.DataFrame:
    previous = pd.read_excel(
        previous_path,
        sheet_name=ALL_ORDER_SHEET_NAME,
        dtype=object,
    )
    previous.columns = [normalize_header(column) for column in previous.columns]
    if QUOTE_ID_HEADER not in previous.columns:
        raise ValueError(f"Previous workbook is missing required column: {QUOTE_ID_HEADER}")
    previous[QUOTE_ID_HEADER] = previous[QUOTE_ID_HEADER].map(normalize_quote_id)
    return previous


def occurrence_keys(dataframe: pd.DataFrame) -> list[tuple[str, int]]:
    quote_ids = dataframe[QUOTE_ID_HEADER].map(normalize_quote_id)
    occurrences = quote_ids.groupby(quote_ids, sort=False).cumcount()
    return list(zip(quote_ids, occurrences, strict=False))


def occurrence_lookup(
    previous: pd.DataFrame,
    value_header: str,
) -> dict[tuple[str, int], object]:
    if value_header not in previous.columns:
        return {}
    return dict(
        zip(occurrence_keys(previous), previous[value_header], strict=False)
    )


def first_quote_lookup(previous: pd.DataFrame, value_header: str) -> dict[str, object]:
    if value_header not in previous.columns:
        return {}
    lookup: dict[str, object] = {}
    for quote_id, value in zip(
        previous[QUOTE_ID_HEADER].map(normalize_quote_id),
        previous[value_header],
        strict=False,
    ):
        if quote_id and quote_id not in lookup:
            lookup[quote_id] = value
    return lookup


def apply_previous_field_fallbacks(
    current: pd.DataFrame,
    previous: pd.DataFrame,
    headers: Iterable[str],
) -> int:
    fallback_count = 0
    for header in headers:
        if header not in current.columns or header not in previous.columns:
            continue
        previous_lookup = first_quote_lookup(previous, header)
        values: list[object] = []
        for quote_id, current_value in zip(
            current[QUOTE_ID_HEADER].map(normalize_quote_id),
            current[header],
            strict=False,
        ):
            previous_value = previous_lookup.get(quote_id)
            if is_missing(current_value) and not is_missing(previous_value):
                values.append(previous_value)
                fallback_count += 1
            elif is_missing(current_value):
                values.append(current_value)
            else:
                values.append(current_value)
        current[header] = pd.Series(values, index=current.index, dtype=object)
    return fallback_count


def numeric_value(value: object) -> int | float | None:
    if is_missing(value):
        return None
    if isinstance(value, str):
        value = value.strip().replace(",", "").replace(" ", "")
    number = pd.to_numeric(value, errors="coerce")
    if pd.isna(number):
        return None
    number = float(number)
    return int(number) if number.is_integer() else number


def load_collabs_lookup(
    collabs_path: Path,
    required_quote_ids: set[str],
) -> dict[str, tuple[int | float | None, int | float | None]]:
    """Load only OTC/MRC rows needed as fallback from the large collabs CSV."""
    if not required_quote_ids:
        return {}

    lookup: dict[str, tuple[int | float | None, int | float | None]] = {}
    required_headers = {"quote_num", "otc", "mrc"}
    reader = pd.read_csv(
        collabs_path,
        usecols=lambda column: str(column).strip().lower() in required_headers,
        dtype="string",
        chunksize=200_000,
        low_memory=False,
    )
    with reader:
        for chunk in reader:
            chunk.columns = [str(column).strip().lower() for column in chunk.columns]
            missing_headers = required_headers.difference(chunk.columns)
            if missing_headers:
                raise ValueError(
                    "Collabs CSV is missing required columns: "
                    + ", ".join(sorted(missing_headers))
                )

            chunk["quote_num"] = chunk["quote_num"].map(normalize_quote_id)
            matches = chunk.loc[chunk["quote_num"].isin(required_quote_ids)]
            for row in matches.itertuples(index=False):
                quote_id = normalize_quote_id(row.quote_num)
                if quote_id and quote_id not in lookup:
                    lookup[quote_id] = (numeric_value(row.otc), numeric_value(row.mrc))
            if required_quote_ids.issubset(lookup):
                break

    return lookup


def resolve_financial_values(
    previous_otc: object,
    previous_mrc: object,
    collabs_values: tuple[object, object] = (None, None),
) -> tuple[int | float, int | float, bool, bool]:
    """Resolve OTC/MRC using previous workbook, collabs CSV, then zero."""
    otc_value = numeric_value(previous_otc)
    mrc_value = numeric_value(previous_mrc)
    collabs_otc = numeric_value(collabs_values[0])
    collabs_mrc = numeric_value(collabs_values[1])
    used_collabs = False
    used_zero = False

    if otc_value is None and collabs_otc is not None:
        otc_value = collabs_otc
        used_collabs = True
    if mrc_value is None and collabs_mrc is not None:
        mrc_value = collabs_mrc
        used_collabs = True
    if otc_value is None:
        otc_value = 0
        used_zero = True
    if mrc_value is None:
        mrc_value = 0
        used_zero = True

    return otc_value, mrc_value, used_collabs, used_zero


def parse_excel_datetime(value: object, *, dayfirst: bool = True) -> datetime | None:
    if is_missing(value) or isinstance(value, time):
        return None
    if isinstance(value, pd.Timestamp):
        parsed = value.to_pydatetime()
        return parsed if MIN_VALID_DATE_YEAR <= parsed.year <= MAX_VALID_DATE_YEAR else None
    if isinstance(value, datetime):
        return value if MIN_VALID_DATE_YEAR <= value.year <= MAX_VALID_DATE_YEAR else None
    if isinstance(value, date):
        parsed = datetime.combine(value, time.min)
        return parsed if MIN_VALID_DATE_YEAR <= parsed.year <= MAX_VALID_DATE_YEAR else None

    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if 1 <= float(value) <= 2_958_465:
            try:
                parsed = datetime(1899, 12, 30) + timedelta(days=float(value))
                return parsed if MIN_VALID_DATE_YEAR <= parsed.year <= MAX_VALID_DATE_YEAR else None
            except (OverflowError, ValueError):
                return None

    text = str(value).strip()
    if text.lower() in MISSING_TEXT or text in {"-", "00:00:00"}:
        return None
    text = re.sub(r"/{2,}", "/", text)
    text = re.sub(r"\s+UTC$", "", text, flags=re.IGNORECASE)

    day_first_formats = (
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y",
    )
    month_first_formats = (
        "%m/%d/%Y %H:%M:%S",
        "%m/%d/%Y %H:%M",
        "%m/%d/%Y",
    )
    formats = (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
        *(day_first_formats if dayfirst else month_first_formats),
        *(month_first_formats if dayfirst else day_first_formats),
    )
    for fmt in formats:
        try:
            parsed = datetime.strptime(text, fmt)
            return parsed if MIN_VALID_DATE_YEAR <= parsed.year <= MAX_VALID_DATE_YEAR else None
        except ValueError:
            continue

    try:
        parsed = date_parser.parse(text, dayfirst=dayfirst, fuzzy=False)
        return parsed if MIN_VALID_DATE_YEAR <= parsed.year <= MAX_VALID_DATE_YEAR else None
    except (OverflowError, TypeError, ValueError):
        return None


def load_source_new_rfs_lookup(worksheet, required_quote_ids: set[str]) -> dict[str, object]:
    """Reproduce the raw workbook's VLOOKUP(D:AL, 35, FALSE) first-match rule."""
    if not required_quote_ids:
        return {}

    lookup: dict[str, object] = {}
    for row in worksheet.iter_rows(
        min_col=RAW_SOURCE_QUOTE_ID_COLUMN,
        max_col=RAW_SOURCE_NEW_RFS_COLUMN,
        values_only=True,
    ):
        quote_id = normalize_quote_id(row[0])
        if not quote_id or quote_id not in required_quote_ids or quote_id in lookup:
            continue

        source_value = row[RAW_SOURCE_NEW_RFS_COLUMN - RAW_SOURCE_QUOTE_ID_COLUMN]
        parsed = parse_excel_datetime(source_value, dayfirst=False)
        lookup[quote_id] = (
            datetime.combine(parsed.date(), time.min)
            if parsed is not None
            else source_value
        )
        if len(lookup) == len(required_quote_ids):
            break

    return lookup


def normalize_month_first_dashboard_dates(df: pd.DataFrame) -> None:
    """Resolve region-sensitive dashboard text dates before other lookups."""
    for header in MONTH_FIRST_TEXT_DATE_HEADERS:
        if header not in df.columns:
            continue

        values: list[object] = []
        for value in df[header]:
            if not isinstance(value, str) or is_missing(value) or value.strip() == "-":
                values.append(value)
                continue

            parsed = parse_excel_datetime(value, dayfirst=False)
            values.append(parsed if parsed is not None else value)
        df[header] = pd.Series(values, index=df.index, dtype=object)


def apply_current_date_first_match(df: pd.DataFrame) -> None:
    """Match manual VLOOKUP behavior for nonblank duplicate date values."""
    quote_ids = df[QUOTE_ID_HEADER].map(normalize_quote_id)
    duplicate_mask = quote_ids.duplicated(keep=False)
    if not duplicate_mask.any():
        return

    for header in FIRST_MATCH_CURRENT_DATE_HEADERS:
        if header not in df.columns:
            continue
        first_values = first_quote_lookup(df, header)
        nonblank_duplicate = duplicate_mask & ~df[header].map(is_missing)
        df.loc[nonblank_duplicate, header] = quote_ids.loc[nonblank_duplicate].map(first_values)


def normalize_date_columns(df: pd.DataFrame) -> tuple[pd.DataFrame, dict[str, list[str]]]:
    result = df.copy()
    invalid_dates: dict[str, list[str]] = {}
    for header in DATE_HEADERS:
        if header not in result.columns:
            continue
        normalized_values: list[object] = []
        invalid_values: list[str] = []
        for value in result[header]:
            parsed = parse_excel_datetime(value)
            if is_excel_zero_date(value):
                normalized_values.append(EXCEL_ZERO_DATE_SERIAL)
            elif parsed is not None:
                normalized_values.append(parsed)
            elif is_missing(value) or str(value).strip() == "-":
                normalized_values.append(None)
            else:
                normalized_values.append(str(value).strip())
                invalid_values.append(str(value).strip())
        result[header] = pd.Series(normalized_values, index=result.index, dtype=object)
        if invalid_values:
            invalid_dates[header] = sorted(set(invalid_values))
    return result, invalid_dates


def target_period(report_date: date) -> date:
    """Return the business month used by the target buckets.

    The first calendar-day snapshot still closes the preceding month. From the
    second day onward, the report uses the current month.
    """
    if report_date.day == 1:
        return report_date - timedelta(days=1)
    return report_date


def target_value(value: object, period: date) -> str:
    parsed = parse_excel_datetime(value)
    month_name = calendar.month_name[period.month]
    if parsed is None:
        return "Target Not Yet Inputted"
    source_month = (parsed.year, parsed.month)
    target_month = (period.year, period.month)
    if source_month < target_month:
        return f"Before {month_name}"
    if source_month > target_month:
        return f"After {month_name}"
    return f"Target {month_name}"


def target_item_kind(value: object) -> str | None:
    normalized = " ".join(str(value).strip().casefold().split())
    if normalized in {"(all)", "all"}:
        return "all"
    if "not yet inputted" in normalized:
        return "missing"
    if normalized.startswith("before "):
        return "before"
    if normalized.startswith("after "):
        return "after"
    if normalized.startswith("target "):
        return "target"
    return None


def target_item_label(kind: str, period: date) -> str:
    month_name = calendar.month_name[period.month]
    labels = {
        "before": f"Before {month_name}",
        "target": f"Target {month_name}",
        "after": f"After {month_name}",
        "missing": "Target Not Yet Inputted",
    }
    try:
        return labels[kind]
    except KeyError as exc:
        raise ValueError(f"Unknown target item kind: {kind}") from exc


def target_value_with_previous(
    current_value: object,
    previous_target_value: object,
    period: date,
) -> str:
    """Match the manual carry-forward behavior within one target period."""
    previous_kind = target_item_kind(previous_target_value)
    period_month = calendar.month_name[period.month].casefold()
    previous_target_text = str(previous_target_value).strip().casefold()
    same_period = previous_kind == "missing" or previous_target_text.endswith(period_month)

    if same_period and previous_kind not in {None, "missing", "all"}:
        return target_item_label(previous_kind, period)

    return target_value(current_value, period)


def is_plausible_operational_date(value: datetime | None, report_date: date) -> bool:
    if value is None:
        return False
    if value.year == MIN_VALID_DATE_YEAR:
        return True
    return MIN_VALID_DATE_YEAR < value.year <= report_date.year + MAX_OPERATIONAL_FUTURE_YEARS


def differs_only_by_year(current_value: datetime | None, previous_value: datetime | None) -> bool:
    if current_value is None or previous_value is None:
        return False
    return (
        current_value.year != previous_value.year
        and (current_value.month, current_value.day)
        == (previous_value.month, previous_value.day)
    )


def reconciled_date_value(
    current_value: object,
    previous_value: object,
    report_date: date,
    *,
    prefer_previous: bool = False,
) -> tuple[object, bool]:
    if is_excel_zero_date(current_value):
        return EXCEL_ZERO_DATE_SERIAL, False

    current_date = parse_excel_datetime(current_value)
    previous_date = parse_excel_datetime(previous_value)

    if prefer_previous:
        if is_excel_zero_date(previous_value):
            return EXCEL_ZERO_DATE_SERIAL, not is_excel_zero_date(current_value)
        if is_plausible_operational_date(previous_date, report_date):
            return previous_date, current_date != previous_date

    if is_missing(current_value) or str(current_value).strip() == "-":
        return current_value, False

    if differs_only_by_year(current_date, previous_date) and is_plausible_operational_date(
        previous_date,
        report_date,
    ):
        return previous_date, True
    if is_plausible_operational_date(current_date, report_date):
        return current_date, False
    if is_excel_zero_date(previous_value):
        return EXCEL_ZERO_DATE_SERIAL, True
    if is_plausible_operational_date(previous_date, report_date):
        return previous_date, True
    return current_value, False


def reconcile_date_column(
    current: pd.DataFrame,
    previous: pd.DataFrame,
    header: str,
    report_date: date,
) -> tuple[pd.Series, int]:
    """Reconcile one date column with the lookup semantics used by Excel."""
    month_opening_carry = report_date.day == 1 and header in MONTH_OPENING_CARRY_DATE_HEADERS
    use_first_match = header in FIRST_MATCH_CURRENT_DATE_HEADERS and not month_opening_carry
    previous_lookup = (
        first_quote_lookup(previous, header)
        if use_first_match
        else occurrence_lookup(previous, header)
    )
    reconciled_values: list[object] = []
    fallback_count = 0
    prefer_previous = month_opening_carry

    for current_value, occurrence_key in zip(
        current[header],
        occurrence_keys(current),
        strict=False,
    ):
        lookup_key = occurrence_key[0] if use_first_match else occurrence_key
        reconciled_value, used_fallback = reconciled_date_value(
            current_value,
            previous_lookup.get(lookup_key),
            report_date,
            prefer_previous=prefer_previous,
        )
        reconciled_values.append(reconciled_value)
        fallback_count += int(used_fallback)

    return pd.Series(reconciled_values, index=current.index, dtype=object), fallback_count


def status_order_value(aging: object) -> object:
    if is_missing(aging):
        return pd.NA
    aging_number = numeric_value(aging)
    if aging_number is None:
        return pd.NA
    if aging_number < -5:
        return "On Track"
    if aging_number < 0:
        return "Potential Delay"
    return "Delay"


def pre_installation_range_value(aging: object) -> object:
    aging_number = numeric_value(aging)
    if aging_number is None:
        return pd.NA
    if aging_number <= 14:
        return "1-7 Days"
    if aging_number <= 30:
        return "8-14 Days"
    if aging_number <= 60:
        return "15-30 Days"
    return "> 30 Days"


def financial_value_or_zero(value: object) -> int | float:
    parsed = numeric_value(value)
    return 0 if parsed is None else parsed


def insert_after(df: pd.DataFrame, after_header: str, header: str, values: Iterable[object]) -> None:
    if after_header not in df.columns:
        raise ValueError(f"Could not insert {header}; missing anchor column: {after_header}")
    position = df.columns.get_loc(after_header) + 1
    df.insert(position, header, list(values))


def build_all_order(
    raw: pd.DataFrame,
    previous: pd.DataFrame,
    collabs_path: Path,
    report_date: date,
    previous_path: Path,
) -> BuildResult:
    result = raw.copy().reset_index(drop=True)
    previous = previous.copy().reset_index(drop=True)
    normalize_month_first_dashboard_dates(result)
    apply_current_date_first_match(result)
    quote_ids = result[QUOTE_ID_HEADER].map(normalize_quote_id)
    duplicate_mask = quote_ids.duplicated(keep=False)
    duplicate_quote_ids = tuple(sorted(quote_ids.loc[duplicate_mask].unique()))
    duplicate_rows_preserved = int(quote_ids.duplicated(keep="first").sum())
    previous_keys = set(occurrence_keys(previous))
    field_fallback_count = apply_previous_field_fallbacks(
        result,
        previous,
        PREVIOUS_FALLBACK_HEADERS,
    )
    date_fallback_count = 0

    for header in DATE_HEADERS:
        if header not in result.columns:
            continue
        result[header], fallback_count = reconcile_date_column(
            result,
            previous,
            header,
            report_date,
        )
        date_fallback_count += fallback_count

    order_type_lookup = first_quote_lookup(previous, ORDER_TYPE_HEADER)
    otc_lookup = first_quote_lookup(previous, OTC_HEADER)
    mrc_lookup = first_quote_lookup(previous, MRC_HEADER)
    previous_status_lookup = first_quote_lookup(previous, STATUS_DELIVERY_HEADER)
    previous_quote_ids = set(previous[QUOTE_ID_HEADER].map(normalize_quote_id))
    pre_installation_start_lookup = first_quote_lookup(previous, PRE_INSTALLATION_START_HEADER)

    missing_financial_ids = {
        quote_id
        for quote_id in quote_ids
        if numeric_value(otc_lookup.get(quote_id)) is None
        or numeric_value(mrc_lookup.get(quote_id)) is None
    }
    collabs_lookup = load_collabs_lookup(collabs_path, missing_financial_ids)

    order_types: list[object] = []
    otc_values: list[object] = []
    mrc_values: list[object] = []
    previous_status_values: list[object] = []
    pre_installation_start_values: list[object] = []
    collabs_fallback_count = 0
    zero_fallback_count = 0

    current_status_values = result[STATUS_DELIVERY_HEADER].map(normalize_status)
    for quote_id, current_status in zip(
        quote_ids,
        current_status_values,
        strict=False,
    ):
        previous_order_type = order_type_lookup.get(quote_id)
        order_types.append(
            "New Registration" if is_missing(previous_order_type) else str(previous_order_type).strip()
        )

        previous_otc, previous_mrc, used_collabs, used_zero = resolve_financial_values(
            otc_lookup.get(quote_id),
            mrc_lookup.get(quote_id),
            collabs_lookup.get(quote_id, (None, None)),
        )

        otc_values.append(previous_otc)
        mrc_values.append(previous_mrc)
        collabs_fallback_count += int(used_collabs)
        zero_fallback_count += int(used_zero)

        if quote_id not in previous_quote_ids:
            previous_status_values.append(EXCEL_NA_ERROR)
        else:
            previous_status = previous_status_lookup.get(quote_id)
            previous_status_values.append(0 if is_missing(previous_status) else previous_status)

        previous_start = pre_installation_start_lookup.get(quote_id)
        is_active_order = (
            not is_missing(current_status)
            and str(current_status).strip().casefold() not in {"so complete", "cancel"}
        )
        if quote_id not in previous_quote_ids or (is_missing(previous_start) and is_active_order):
            pre_installation_start_values.append(
                datetime.combine(report_date - timedelta(days=1), time.min)
            )
        elif is_excel_zero_date(previous_start):
            pre_installation_start_values.append(EXCEL_ZERO_DATE_SERIAL)
        else:
            pre_installation_start_values.append(parse_excel_datetime(previous_start))

    insert_after(result, "QUOTE NAME", ORDER_TYPE_HEADER, order_types)
    insert_after(result, "PARTNERS", OTC_HEADER, otc_values)
    insert_after(result, OTC_HEADER, MRC_HEADER, mrc_values)

    fab_dates = result[FAB_UPLOAD_HEADER].map(parse_excel_datetime)
    year_values = [str(value.year) if not is_missing(value) else UNKNOWN_YEAR for value in fab_dates]
    insert_after(result, FAB_UPLOAD_HEADER, YEAR_FAB_UPLOAD_HEADER, year_values)

    dynamic_target_header = target_header(report_date)
    target_bucket_period = target_period(report_date)
    previous_target_header = next(
        (
            column
            for column in previous.columns
            if normalize_header(column).casefold().startswith("target ")
        ),
        None,
    )
    previous_target_lookup = (
        occurrence_lookup(previous, previous_target_header)
        if previous_target_header is not None
        else {}
    )
    target_values = [
        target_value_with_previous(
            current_value,
            previous_target_lookup.get(occurrence_key),
            target_bucket_period,
        )
        for current_value, occurrence_key in zip(
            result[NEW_RFS_INITIAL_HEADER],
            occurrence_keys(result),
            strict=False,
        )
    ]
    insert_after(result, NEW_RFS_INITIAL_HEADER, dynamic_target_header, target_values)

    actual_rfs_dates = result[ACTUAL_RFS_DATE_HEADER].map(parse_excel_datetime)
    aging_values = [
        (report_date - value.date()).days if not is_missing(value) else pd.NA
        for value in actual_rfs_dates
    ]
    insert_after(result, ACTUAL_RFS_DATE_HEADER, AGING_OF_RFS_HEADER, aging_values)
    insert_after(
        result,
        AGING_OF_RFS_HEADER,
        STATUS_ORDER_HEADER,
        [status_order_value(value) for value in aging_values],
    )

    insert_after(result, AOS_SENT_DATE_HEADER, ON_TIME_DELIVERY_HEADER, [pd.NA] * len(result))

    dynamic_previous_status_header = previous_status_header(previous_path, report_date)
    insert_after(
        result,
        STATUS_DELIVERY_HEADER,
        dynamic_previous_status_header,
        previous_status_values,
    )
    insert_after(
        result,
        dynamic_previous_status_header,
        PRE_INSTALLATION_START_HEADER,
        pre_installation_start_values,
    )

    pre_installation_aging_values = [
        (report_date - value.date()).days if isinstance(value, datetime) else pd.NA
        for value in pre_installation_start_values
    ]
    insert_after(
        result,
        PRE_INSTALLATION_START_HEADER,
        PRE_INSTALLATION_AGING_HEADER,
        pre_installation_aging_values,
    )
    insert_after(
        result,
        PRE_INSTALLATION_AGING_HEADER,
        PRE_INSTALLATION_RANGE_HEADER,
        [pre_installation_range_value(value) for value in pre_installation_aging_values],
    )

    result[STATUS_DELIVERY_HEADER] = result[STATUS_DELIVERY_HEADER].map(normalize_status)
    result, invalid_dates = normalize_date_columns(result)
    result = result.reset_index(drop=True)

    if len(result.columns) != 61:
        raise ValueError(
            f"ALL ORDER must contain 61 columns, found {len(result.columns)}. "
            "The raw workbook structure may have changed."
        )

    return BuildResult(
        dataframe=result,
        invalid_dates=invalid_dates,
        collabs_fallback_count=collabs_fallback_count,
        zero_fallback_count=zero_fallback_count,
        duplicate_quote_ids=duplicate_quote_ids,
        duplicate_rows_preserved=duplicate_rows_preserved,
        date_fallback_count=date_fallback_count,
        field_fallback_count=field_fallback_count,
    )


def com_value(value: object) -> object:
    if value is None or value is pd.NA:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(value, pd.Timestamp):
        value = value.to_pydatetime()
    if isinstance(value, datetime):
        delta = value - datetime(1899, 12, 30)
        return delta.days + (delta.seconds + delta.microseconds / 1_000_000) / 86_400
    if isinstance(value, date) and not isinstance(value, datetime):
        delta = datetime.combine(value, time.min) - datetime(1899, 12, 30)
        return delta.days
    if isinstance(value, bool):
        return value
    if isinstance(value, numbers.Integral):
        return int(value)
    if isinstance(value, numbers.Real):
        return float(value)
    return value


def find_sheet(workbook, sheet_name: str):
    for sheet in workbook.Worksheets:
        if str(sheet.Name) == sheet_name:
            return sheet
    return None


def apply_date_formats(worksheet, df: pd.DataFrame) -> None:
    def contiguous_ranges(rows: list[int]) -> Iterable[tuple[int, int]]:
        if not rows:
            return
        start = previous = rows[0]
        for row_number in rows[1:]:
            if row_number != previous + 1:
                yield start, previous
                start = row_number
            previous = row_number
        yield start, previous

    for header in DATE_HEADERS:
        if header not in df.columns:
            continue
        column_number = df.columns.get_loc(header) + 1
        column_letter = get_column_letter(column_number)
        date_only_rows: list[int] = []
        date_time_rows: list[int] = []
        for row_number, value in enumerate(df[header], start=2):
            if is_excel_zero_date(value):
                date_time_rows.append(row_number)
                continue
            if not isinstance(value, datetime):
                continue
            if value.time() == time.min:
                date_only_rows.append(row_number)
            else:
                date_time_rows.append(row_number)

        for rows, number_format in (
            (date_only_rows, DATE_ONLY_FORMAT),
            (date_time_rows, DATE_TIME_FORMAT),
        ):
            for start_row, end_row in contiguous_ranges(rows):
                worksheet.Range(
                    f"{column_letter}{start_row}:{column_letter}{end_row}"
                ).NumberFormat = number_format


def apply_na_errors(excel, worksheet, df: pd.DataFrame) -> None:
    """Restore #N/A sentinels as real Excel errors before PivotTable refresh."""
    for header in df.columns:
        error_rows = [
            row_number
            for row_number, value in enumerate(df[header], start=2)
            if str(value).strip().upper() == EXCEL_NA_ERROR
        ]
        if not error_rows:
            continue

        column_number = df.columns.get_loc(header) + 1
        worksheet.Range(
            worksheet.Cells(2, column_number),
            worksheet.Cells(len(df) + 1, column_number),
        ).NumberFormat = "General"

        start = previous = error_rows[0]
        ranges: list[tuple[int, int]] = []
        for row_number in error_rows[1:]:
            if row_number != previous + 1:
                ranges.append((start, previous))
                start = row_number
            previous = row_number
        ranges.append((start, previous))

        auto_fill_setting = None
        try:
            auto_fill_setting = excel.AutoCorrect.AutoFillFormulasInLists
            excel.AutoCorrect.AutoFillFormulasInLists = False
        except Exception:
            auto_fill_setting = None
        try:
            for start_row, end_row in ranges:
                target = worksheet.Range(
                    worksheet.Cells(start_row, column_number),
                    worksheet.Cells(end_row, column_number),
                )
                target.Formula = "=NA()"
                target.Calculate()
                target.Copy()
                target.PasteSpecial(Paste=-4163)  # xlPasteValues
                excel.CutCopyMode = False
        finally:
            if auto_fill_setting is not None:
                try:
                    excel.AutoCorrect.AutoFillFormulasInLists = auto_fill_setting
                except Exception:
                    pass


def autofit_with_limits(worksheet, column_count: int) -> None:
    for column_number in range(1, column_count + 1):
        column = worksheet.Columns(column_number)
        column.AutoFit()
        width = float(column.ColumnWidth)
        if width > MAX_COLUMN_WIDTH:
            column.ColumnWidth = MAX_COLUMN_WIDTH
        elif width < MIN_COLUMN_WIDTH:
            column.ColumnWidth = MIN_COLUMN_WIDTH


def pivot_header_columns(worksheet, table_range, header_row: int) -> dict[int, str]:
    headers: dict[int, str] = {}
    start_column = int(table_range.Column)
    end_column = start_column + int(table_range.Columns.Count) - 1
    for column_number in range(start_column, end_column + 1):
        value = worksheet.Cells(header_row, column_number).Value
        if value is None:
            continue
        normalized = " ".join(str(value).strip().casefold().split())
        if normalized:
            headers[column_number] = normalized
    return headers


def find_target_pivot_field(pivot_table, expected_header: str | None = None):
    fields = pivot_table.PivotFields()
    expected = normalize_header(expected_header).casefold() if expected_header else None
    fallback = None
    for field_index in range(1, fields.Count + 1):
        field = fields(field_index)
        name = normalize_header(field.Name)
        if expected is not None and name.casefold() == expected:
            return field
        if fallback is None and name.casefold().startswith("target "):
            fallback = field
    return fallback


def capture_target_pivot_filters(worksheet) -> list[TargetPivotPageFilter]:
    """Capture target page filters before the monthly source header changes."""
    states: list[TargetPivotPageFilter] = []
    pivot_tables = worksheet.PivotTables()
    for pivot_index in range(1, pivot_tables.Count + 1):
        pivot_table = pivot_tables(pivot_index)
        field = find_target_pivot_field(pivot_table)
        if field is None or int(field.Orientation) != 3:  # xlPageField
            continue

        current_page = str(field.CurrentPage)
        show_all = target_item_kind(current_page) == "all"
        selection_kinds: list[str] = []
        try:
            multiple_items = bool(field.EnableMultiplePageItems)
        except Exception:
            multiple_items = False

        if multiple_items and not show_all:
            items = field.PivotItems()
            for item_index in range(1, items.Count + 1):
                item = items(item_index)
                if not bool(item.Visible):
                    continue
                kind = target_item_kind(item.Name)
                if kind is not None and kind != "all" and kind not in selection_kinds:
                    selection_kinds.append(kind)
        elif not show_all:
            kind = target_item_kind(current_page)
            if kind is None:
                raise RuntimeError(
                    f"Unsupported target filter item {current_page!r} in "
                    f"PivotTable '{pivot_table.Name}'."
                )
            selection_kinds.append(kind)

        states.append(
            TargetPivotPageFilter(
                pivot_name=str(pivot_table.Name),
                position=int(field.Position),
                selection_kinds=tuple(selection_kinds),
                show_all=show_all,
            )
        )
    return states


def restore_target_pivot_filters(
    worksheet,
    states: list[TargetPivotPageFilter],
    report_date: date,
) -> None:
    """Rebind monthly target filters after the PivotCache schema refresh."""
    expected_header = target_header(report_date)
    period = target_period(report_date)
    for state in states:
        pivot_table = worksheet.PivotTables(state.pivot_name)
        field = find_target_pivot_field(pivot_table, expected_header)
        if field is None:
            raise RuntimeError(
                f"PivotTable '{state.pivot_name}' is missing target field "
                f"{expected_header!r} after refresh."
            )

        field.Orientation = 3  # xlPageField
        field.Position = state.position
        field.ClearAllFilters()
        if state.show_all:
            field.CurrentPage = "(All)"
            pivot_table.RefreshTable()
            continue

        desired_labels = {
            target_item_label(kind, period).casefold(): kind
            for kind in state.selection_kinds
        }
        available_items: dict[str, object] = {}
        items = field.PivotItems()
        for item_index in range(1, items.Count + 1):
            item = items(item_index)
            available_items[str(item.Name).strip().casefold()] = item

        missing_labels = sorted(set(desired_labels) - set(available_items))
        if missing_labels:
            raise RuntimeError(
                f"PivotTable '{state.pivot_name}' cannot apply target filter; "
                f"missing item(s): {', '.join(missing_labels)}."
            )

        if len(desired_labels) == 1:
            selected_label = next(iter(desired_labels))
            field.CurrentPage = str(available_items[selected_label].Name)
        else:
            field.EnableMultiplePageItems = True
            for label in desired_labels:
                available_items[label].Visible = True
            for label, item in available_items.items():
                item.Visible = label in desired_labels
        pivot_table.RefreshTable()


def update_pivot_period_titles(worksheet, report_date: date) -> None:
    month_name = calendar.month_name[report_date.month].upper()
    year = report_date.year
    replacements = {
        "target complete": (
            f"TARGET COMPLETE {month_name} {year} "
            f"(Based on Dashboard 1 {month_name} {year})"
        ),
        "order aging start from pre-installation status": (
            f"ORDER AGING START FROM PRE-INSTALLATION STATUS {month_name} {year % 100:02d}"
        ),
        "status order with aging from rfs commit date": (
            f"STATUS ORDER WITH AGING FROM RFS COMMIT DATE {month_name} {year % 100:02d}"
        ),
        "delay completion": (
            "DELAY COMPLETION (SHOULD BE COMPLETE BEFORE "
            f"{month_name} {year})"
        ),
        "target after": f"TARGET AFTER {month_name} {year}",
    }

    used_range = worksheet.UsedRange
    used_values = used_range.Value2
    if used_values is None:
        return
    for row_offset, row in enumerate(used_values):
        for column_offset, value in enumerate(row):
            if not isinstance(value, str):
                continue
            normalized = " ".join(value.strip().casefold().split())
            for prefix, replacement in replacements.items():
                if normalized.startswith(prefix):
                    worksheet.Cells(
                        used_range.Row + row_offset,
                        used_range.Column + column_offset,
                    ).Value = replacement
                    break


def hidden_row_ranges_between_sections(
    section_bottom_row: int,
    next_title_row: int,
    content_rows: Iterable[int] = (),
) -> list[tuple[int, int]]:
    """Return blank row runs to hide while preserving useful section spacing."""
    gap_start = section_bottom_row + 1
    gap_end = next_title_row - 1
    if gap_start > gap_end:
        return []

    protected_rows = {gap_start, gap_end, gap_end - 1}
    for row_number in content_rows:
        if gap_start <= row_number <= gap_end:
            protected_rows.add(row_number)
            protected_rows.add(max(gap_start, row_number - 1))

    hidden_rows = [
        row_number
        for row_number in range(gap_start, gap_end + 1)
        if row_number not in protected_rows
    ]
    if not hidden_rows:
        return []

    ranges: list[tuple[int, int]] = []
    range_start = previous_row = hidden_rows[0]
    for row_number in hidden_rows[1:]:
        if row_number != previous_row + 1:
            ranges.append((range_start, previous_row))
            range_start = row_number
        previous_row = row_number
    ranges.append((range_start, previous_row))
    return ranges


def normalize_pivot_section_row_visibility(worksheet) -> None:
    """Hide vacated Pivot rows using the refreshed section boundaries."""
    used_range = worksheet.UsedRange
    used_values = used_range.Value2
    if used_values is None:
        return
    if not isinstance(used_values, tuple):
        value_rows = ((used_values,),)
    elif used_values and not isinstance(used_values[0], tuple):
        value_rows = (used_values,)
    else:
        value_rows = used_values

    title_markers = tuple(
        normalize_layout_text(marker) for marker in PIVOT_SECTION_TITLE_MARKERS
    )
    title_rows: set[int] = set()
    content_rows: set[int] = set()
    for row_offset, row_values in enumerate(value_rows, start=int(used_range.Row)):
        if not isinstance(row_values, tuple):
            row_values = (row_values,)
        nonblank_values = [value for value in row_values if value not in (None, "")]
        if nonblank_values:
            content_rows.add(row_offset)
        if any(
            marker in normalize_layout_text(value)
            for value in nonblank_values
            for marker in title_markers
        ):
            title_rows.add(row_offset)

    sorted_title_rows = sorted(title_rows)
    if len(sorted_title_rows) < 2:
        return

    pivot_ranges: list[tuple[int, int]] = []
    pivot_tables = worksheet.PivotTables()
    for pivot_index in range(1, int(pivot_tables.Count) + 1):
        table_range = pivot_tables(pivot_index).TableRange2
        pivot_start_row = int(table_range.Row)
        pivot_ranges.append(
            (
                pivot_start_row,
                pivot_start_row + int(table_range.Rows.Count) - 1,
            )
        )

    for title_row, next_title_row in zip(
        sorted_title_rows,
        sorted_title_rows[1:],
        strict=False,
    ):
        section_bottoms = [
            bottom_row
            for pivot_start_row, bottom_row in pivot_ranges
            if title_row < pivot_start_row < next_title_row
        ]
        if not section_bottoms:
            continue

        section_bottom_row = max(section_bottoms)
        gap_start = section_bottom_row + 1
        gap_end = next_title_row - 1
        if gap_start > gap_end:
            continue

        worksheet.Range(
            worksheet.Cells(gap_start, 1),
            worksheet.Cells(gap_end, 1),
        ).EntireRow.Hidden = False
        gap_content_rows = {
            row_number
            for row_number in content_rows
            if gap_start <= row_number <= gap_end
        }
        for hidden_start, hidden_end in hidden_row_ranges_between_sections(
            section_bottom_row,
            next_title_row,
            gap_content_rows,
        ):
            worksheet.Range(
                worksheet.Cells(hidden_start, 1),
                worksheet.Cells(hidden_end, 1),
            ).EntireRow.Hidden = True


def capture_pivot_percentage_blocks(workbook, worksheet):
    """Move percentage blocks aside so dynamic PivotTable columns can expand."""
    scratch = workbook.Worksheets.Add(After=workbook.Worksheets(workbook.Worksheets.Count))
    scratch.Name = "__PERCENTAGE_LAYOUT__"

    used_range = worksheet.UsedRange
    used_values = used_range.Value2
    percentage_headers: list[tuple[int, int]] = []
    if used_values is not None:
        for row_offset, row in enumerate(used_values):
            for column_offset, value in enumerate(row):
                if isinstance(value, str) and "percentage" in value.casefold():
                    percentage_headers.append(
                        (used_range.Row + row_offset, used_range.Column + column_offset)
                    )

    blocks: list[PivotPercentageBlock] = []
    scratch_row = 1
    pivot_tables = worksheet.PivotTables()
    for header_row, percentage_column in percentage_headers:
        candidates = []
        for pivot_index in range(1, pivot_tables.Count + 1):
            pivot_table = pivot_tables(pivot_index)
            table_range = pivot_table.TableRange2
            table_bottom = table_range.Row + table_range.Rows.Count - 1
            table_right = table_range.Column + table_range.Columns.Count - 1
            if table_range.Row <= header_row <= table_bottom and table_right < percentage_column:
                candidates.append((table_right, pivot_table))
        if not candidates:
            continue

        _, pivot_table = max(candidates, key=lambda item: item[0])
        table_range = pivot_table.TableRange2
        formula_row = header_row + 1
        formula = str(worksheet.Cells(formula_row, percentage_column).Formula or "")
        if not formula.startswith("="):
            continue

        end_row = formula_row
        while end_row + 1 <= used_range.Row + used_range.Rows.Count - 1:
            next_formula = str(
                worksheet.Cells(end_row + 1, percentage_column).Formula or ""
            )
            if not next_formula.startswith("="):
                break
            end_row += 1

        source_range = worksheet.Range(
            worksheet.Cells(header_row, percentage_column),
            worksheet.Cells(end_row, percentage_column),
        )
        source_range.Copy(Destination=scratch.Cells(scratch_row, 1))
        blocks.append(
            PivotPercentageBlock(
                pivot_name=str(pivot_table.Name),
                header_row_offset=header_row - int(table_range.Row),
                original_header_row=header_row,
                original_end_row=end_row,
                scratch_start_row=scratch_row,
                formula=formula,
                formula_row=formula_row,
                pivot_header_columns=pivot_header_columns(
                    worksheet,
                    table_range,
                    header_row,
                ),
            )
        )
        source_range.Clear()
        scratch_row += end_row - header_row + 2

    scratch.Visible = 2  # xlSheetVeryHidden
    return scratch, blocks


def remap_percentage_formula(
    formula: str,
    source_row: int,
    target_row: int,
    column_mapping: dict[int, int],
) -> str:
    cell_reference = re.compile(r"(?<![A-Z0-9_])(\$?)([A-Z]{1,3})(\$?)(\d+)")
    row_delta = target_row - source_row

    def replace(match: re.Match[str]) -> str:
        column_absolute, column_text, row_absolute, row_text = match.groups()
        source_column = column_index_from_string(column_text)
        target_column = column_mapping.get(source_column, source_column)
        row_number = int(row_text)
        if not row_absolute:
            row_number += row_delta
        return (
            f"{column_absolute}{get_column_letter(target_column)}"
            f"{row_absolute}{row_number}"
        )

    return cell_reference.sub(replace, formula)


def restore_pivot_percentage_blocks(worksheet, scratch, blocks: list[PivotPercentageBlock]) -> None:
    for block in blocks:
        pivot_table = worksheet.PivotTables(block.pivot_name)
        table_range = pivot_table.TableRange2
        if int(table_range.Rows.Count) <= 1 or int(table_range.Columns.Count) <= 1:
            raise RuntimeError(
                f"PivotTable '{block.pivot_name}' did not render after refresh."
            )

        header_row = int(table_range.Row) + block.header_row_offset
        percentage_column = int(table_range.Column) + int(table_range.Columns.Count)
        end_row = int(table_range.Row) + int(table_range.Rows.Count) - 1
        source_height = block.original_end_row - block.original_header_row + 1
        source_range = scratch.Range(
            scratch.Cells(block.scratch_start_row, 1),
            scratch.Cells(block.scratch_start_row + source_height - 1, 1),
        )
        destination_range = worksheet.Range(
            worksheet.Cells(header_row, percentage_column),
            worksheet.Cells(header_row + source_height - 1, percentage_column),
        )
        source_range.Copy(Destination=destination_range)

        formula_start_row = header_row + 1
        if end_row < formula_start_row:
            continue

        if end_row != header_row + source_height - 1:
            body_template = scratch.Cells(block.scratch_start_row + 1, 1)
            grand_total_template = scratch.Cells(
                block.scratch_start_row + source_height - 1,
                1,
            )
            if end_row > formula_start_row:
                body_template.Copy(
                    Destination=worksheet.Range(
                        worksheet.Cells(formula_start_row, percentage_column),
                        worksheet.Cells(end_row - 1, percentage_column),
                    )
                )
            grand_total_template.Copy(
                Destination=worksheet.Cells(end_row, percentage_column)
            )
            if end_row < header_row + source_height - 1:
                worksheet.Range(
                    worksheet.Cells(end_row + 1, percentage_column),
                    worksheet.Cells(header_row + source_height - 1, percentage_column),
                ).Clear()

        current_headers = pivot_header_columns(worksheet, table_range, header_row)
        current_columns_by_header = {
            header: column for column, header in current_headers.items()
        }
        column_mapping = {
            old_column: current_columns_by_header[header]
            for old_column, header in block.pivot_header_columns.items()
            if header in current_columns_by_header
        }
        formulas = tuple(
            (
                remap_percentage_formula(
                    block.formula,
                    block.formula_row,
                    row_number,
                    column_mapping,
                ),
            )
            for row_number in range(formula_start_row, end_row + 1)
        )
        worksheet.Range(
            worksheet.Cells(formula_start_row, percentage_column),
            worksheet.Cells(end_row, percentage_column),
        ).Formula = formulas


def adjust_pivot_column_widths(worksheet) -> None:
    pivot_columns: set[int] = set()
    pivot_tables = worksheet.PivotTables()
    for pivot_index in range(1, pivot_tables.Count + 1):
        table_range = pivot_tables(pivot_index).TableRange2
        pivot_columns.update(
            range(table_range.Column, table_range.Column + table_range.Columns.Count)
        )

    percentage_columns: set[int] = set()
    used_range = worksheet.UsedRange
    used_values = used_range.Value2
    if used_values is not None:
        for row in used_values:
            for offset, value in enumerate(row):
                if isinstance(value, str) and "percentage" in value.casefold():
                    percentage_columns.add(used_range.Column + offset)

    for column_number in sorted(pivot_columns | percentage_columns):
        column = worksheet.Columns(column_number)
        column.AutoFit()
        width = float(column.ColumnWidth)
        if width > MAX_COLUMN_WIDTH:
            column.ColumnWidth = MAX_COLUMN_WIDTH
        elif width < MIN_COLUMN_WIDTH:
            column.ColumnWidth = MIN_COLUMN_WIDTH

    for column_number in percentage_columns:
        column = worksheet.Columns(column_number)
        if float(column.ColumnWidth) < PIVOT_PERCENTAGE_MIN_WIDTH:
            column.ColumnWidth = PIVOT_PERCENTAGE_MIN_WIDTH


def repair_pivot_percentage_ranges(worksheet) -> None:
    used_range = worksheet.UsedRange
    used_values = used_range.Value2
    percentage_headers: list[tuple[int, int]] = []
    if used_values is not None:
        for row_offset, row in enumerate(used_values):
            for column_offset, value in enumerate(row):
                if isinstance(value, str) and "percentage" in value.casefold():
                    percentage_headers.append(
                        (used_range.Row + row_offset, used_range.Column + column_offset)
                    )

    pivot_tables = worksheet.PivotTables()
    for header_row, percentage_column in percentage_headers:
        related_bottom_rows: list[int] = []
        for pivot_index in range(1, pivot_tables.Count + 1):
            table_range = pivot_tables(pivot_index).TableRange2
            table_bottom = table_range.Row + table_range.Rows.Count - 1
            table_right = table_range.Column + table_range.Columns.Count - 1
            if table_range.Row <= header_row <= table_bottom and table_right < percentage_column:
                related_bottom_rows.append(table_bottom)
        if not related_bottom_rows:
            continue

        target_end_row = max(related_bottom_rows)
        formula_start_row = header_row + 1
        first_formula_cell = worksheet.Cells(formula_start_row, percentage_column)
        formula_r1c1 = str(first_formula_cell.FormulaR1C1 or "")
        if not formula_r1c1.startswith("="):
            continue

        old_end_row = formula_start_row
        while old_end_row + 1 <= used_range.Row + used_range.Rows.Count - 1:
            next_formula = str(
                worksheet.Cells(old_end_row + 1, percentage_column).FormulaR1C1 or ""
            )
            if not next_formula.startswith("="):
                break
            old_end_row += 1

        old_grand_total_cell = worksheet.Cells(old_end_row, percentage_column)
        target_grand_total_cell = worksheet.Cells(target_end_row, percentage_column)
        if target_end_row != old_end_row:
            old_grand_total_cell.Copy(Destination=target_grand_total_cell)

        if target_end_row > old_end_row:
            body_style_source = worksheet.Cells(max(formula_start_row, old_end_row - 1), percentage_column)
            body_style_source.Copy(
                Destination=worksheet.Range(
                    worksheet.Cells(old_end_row, percentage_column),
                    worksheet.Cells(target_end_row - 1, percentage_column),
                )
            )
        elif target_end_row < old_end_row:
            worksheet.Range(
                worksheet.Cells(target_end_row + 1, percentage_column),
                worksheet.Cells(old_end_row, percentage_column),
            ).Clear()

        worksheet.Range(
            worksheet.Cells(formula_start_row, percentage_column),
            worksheet.Cells(target_end_row, percentage_column),
        ).FormulaR1C1 = formula_r1c1


def normalize_pivot_percentage_borders(
    worksheet,
    blocks: list[PivotPercentageBlock],
) -> None:
    """Keep red outlines on percentage blocks without boxing every data row."""
    xl_edge_left = 7
    xl_edge_top = 8
    xl_edge_bottom = 9
    xl_edge_right = 10
    xl_inside_horizontal = 12
    xl_continuous = 1
    xl_line_style_none = -4142
    xl_medium = -4138
    xl_red = 255

    def apply_red_outline(cell) -> None:
        for border_index in (
            xl_edge_left,
            xl_edge_top,
            xl_edge_bottom,
            xl_edge_right,
        ):
            border = cell.Borders(border_index)
            border.LineStyle = xl_continuous
            border.Color = xl_red
            border.Weight = xl_medium

    for block in blocks:
        pivot_table = worksheet.PivotTables(block.pivot_name)
        table_range = pivot_table.TableRange2
        header_row = int(table_range.Row) + block.header_row_offset
        end_row = int(table_range.Row) + int(table_range.Rows.Count) - 1
        percentage_column = int(table_range.Column) + int(table_range.Columns.Count)

        first_body_row = header_row + 1
        last_body_row = end_row - 1
        if end_row <= header_row:
            continue

        if last_body_row >= first_body_row:
            body_range = worksheet.Range(
                worksheet.Cells(first_body_row, percentage_column),
                worksheet.Cells(last_body_row, percentage_column),
            )
            for border_index in (xl_edge_left, xl_edge_right):
                border = body_range.Borders(border_index)
                border.LineStyle = xl_continuous
                border.Color = xl_red
                border.Weight = xl_medium
            body_range.Borders(xl_inside_horizontal).LineStyle = xl_line_style_none
            body_range.Borders(xl_edge_bottom).LineStyle = xl_line_style_none

            first_body_cell = worksheet.Cells(first_body_row, percentage_column)
            first_body_border = first_body_cell.Borders(xl_edge_top)
            first_body_border.LineStyle = xl_continuous
            first_body_border.Color = xl_red
            first_body_border.Weight = xl_medium

        # Apply section boundaries last because Excel treats adjacent cell
        # borders as one shared edge when ranges are reformatted.
        apply_red_outline(worksheet.Cells(header_row, percentage_column))
        apply_red_outline(worksheet.Cells(end_row, percentage_column))


def completion_thresholds(report_date: date) -> CompletionThresholds:
    """Return the capped weekly completion thresholds for a report date."""
    week = min(((report_date.day - 1) // 7) + 1, 4)
    weekly_increment = (week - 1) * 10
    return CompletionThresholds(
        week=week,
        green_percent=30 + weekly_increment,
        yellow_percent=30 + weekly_increment,
        red_percent=20 + weekly_increment,
    )


def apply_weekly_completion_thresholds(
    worksheet,
    blocks: list[PivotPercentageBlock],
    report_date: date,
) -> None:
    """Update percentage legends and traffic-light rules for the report week."""
    thresholds = completion_thresholds(report_date)
    legend_values = {
        "green": f"Green : >{thresholds.green_percent}%",
        "yellow": f"Yellow : >={thresholds.yellow_percent}%",
        "red": f"Red : <{thresholds.red_percent}%",
    }

    used_range = worksheet.UsedRange
    used_values = used_range.Value2
    if used_values is not None:
        for row_offset, row in enumerate(used_values):
            for column_offset, value in enumerate(row):
                if not isinstance(value, str):
                    continue
                normalized = value.strip().casefold()
                for prefix, replacement in legend_values.items():
                    if normalized.startswith(prefix):
                        worksheet.Cells(
                            used_range.Row + row_offset,
                            used_range.Column + column_offset,
                        ).Value = replacement
                        break

    for block in blocks:
        pivot_table = worksheet.PivotTables(block.pivot_name)
        table_range = pivot_table.TableRange2
        header_row = int(table_range.Row) + block.header_row_offset
        formula_start_row = header_row + 1
        end_row = int(table_range.Row) + int(table_range.Rows.Count) - 1
        percentage_column = int(table_range.Column) + int(table_range.Columns.Count)
        if end_row < formula_start_row:
            continue

        target_range = worksheet.Range(
            worksheet.Cells(formula_start_row, percentage_column),
            worksheet.Cells(end_row, percentage_column),
        )
        updated_rules = 0
        format_conditions = target_range.FormatConditions
        for condition_index in range(1, int(format_conditions.Count) + 1):
            condition = format_conditions(condition_index)
            try:
                if int(condition.Type) != 6 or int(condition.IconSet.ID) != 4:
                    continue
                criteria = condition.IconCriteria
                criteria(2).Type = 0  # xlConditionValueNumber
                criteria(2).Value = thresholds.red_percent / 100
                criteria(2).Operator = 7  # xlGreaterEqual
                criteria(3).Type = 0  # xlConditionValueNumber
                criteria(3).Value = thresholds.green_percent / 100
                criteria(3).Operator = 5  # xlGreater
                updated_rules += 1
            except Exception as exc:
                raise RuntimeError(
                    f"Could not update percentage color rule for '{block.pivot_name}': {exc}"
                ) from exc

        if updated_rules == 0:
            raise RuntimeError(
                f"No traffic-light rule found for percentage block '{block.pivot_name}'."
            )


def identify_target_complete_pivots(worksheet):
    pivot_tables = worksheet.PivotTables()
    pivots = [pivot_tables(index) for index in range(1, pivot_tables.Count + 1)]
    summary_candidates = [
        pivot for pivot in pivots if int(pivot.TableRange2.Column) == 1
    ]
    summary_pivot = min(
        summary_candidates,
        key=lambda pivot: int(pivot.TableRange2.Row),
        default=None,
    )

    status_pivot = None
    if summary_pivot is not None:
        summary_range = summary_pivot.TableRange2
        summary_start = int(summary_range.Row)
        summary_end = summary_start + int(summary_range.Rows.Count) - 1
        status_candidates = []
        for pivot in pivots:
            table_range = pivot.TableRange2
            if int(table_range.Column) <= 1:
                continue
            status_start = int(table_range.Row)
            status_end = status_start + int(table_range.Rows.Count) - 1
            if status_start <= summary_end and summary_start <= status_end:
                status_candidates.append(pivot)
        status_pivot = min(
            status_candidates,
            key=lambda pivot: (
                int(pivot.TableRange2.Row),
                int(pivot.TableRange2.Column),
            ),
            default=None,
        )

    if summary_pivot is None or status_pivot is None:
        raise RuntimeError("Could not identify both TARGET COMPLETE PivotTables.")

    return summary_pivot, status_pivot


def validate_target_complete_pivots(worksheet, dataframe: pd.DataFrame) -> None:
    summary_pivot, status_pivot = identify_target_complete_pivots(worksheet)

    mask = pd.Series(True, index=dataframe.index)
    for header in (YEAR_FAB_UPLOAD_HEADER, next(
        (column for column in dataframe.columns if str(column).startswith("TARGET")),
        "",
    )):
        if not header:
            continue
        field = summary_pivot.PivotFields(header)
        visible_items = {
            str(item.Name).strip().casefold()
            for item in field.PivotItems()
            if item.Visible
        }
        if visible_items:
            normalized_values = dataframe[header].map(
                lambda value: "(blank)"
                if is_missing(value)
                else str(value).strip().casefold()
            )
            mask &= normalized_values.isin(visible_items)

    expected_count = int(
        dataframe.loc[mask, QUOTE_ID_HEADER].map(normalize_quote_id).ne("").sum()
    )
    summary_count = int(summary_pivot.GetPivotData(f"Count of {QUOTE_ID_HEADER}").Value)
    status_grand_total_row = status_pivot.TableRange2.Rows.Count
    status_values: list[int] = []
    for column in range(2, status_pivot.TableRange2.Columns.Count + 1):
        value = status_pivot.TableRange2.Cells(status_grand_total_row, column).Value
        if isinstance(value, numbers.Number):
            status_values.append(int(value))
    status_count = expected_count if expected_count in status_values else sum(status_values)

    if expected_count != summary_count or expected_count != status_count:
        raise RuntimeError(
            "TARGET COMPLETE count mismatch: "
            f"source={expected_count}, summary_pivot={summary_count}, "
            f"status_pivot={status_count}."
        )


def update_workbook_via_com(
    workbook_path: Path,
    df: pd.DataFrame,
    report_date: date,
) -> int:
    try:
        import pythoncom  # type: ignore
        import win32com.client  # type: ignore
    except ImportError as exc:
        raise RuntimeError(
            "pywin32 and Microsoft Excel are required to preserve and refresh IDE PivotTables."
        ) from exc

    pythoncom.CoInitialize()
    excel = None
    workbook = None
    previous_calculation = None
    percentage_scratch = None
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.ScreenUpdating = False
        excel.EnableEvents = False
        try:
            previous_calculation = excel.Calculation
            excel.Calculation = -4135  # xlCalculationManual
        except Exception:
            previous_calculation = None

        workbook = excel.Workbooks.Open(str(workbook_path.resolve()), UpdateLinks=0, ReadOnly=False)
        all_order_sheet = find_sheet(workbook, ALL_ORDER_SHEET_NAME)
        pivot_sheet = find_sheet(workbook, PIVOT_SHEET_NAME)
        if all_order_sheet is None or pivot_sheet is None:
            raise RuntimeError("Previous workbook must contain ALL ORDER and PIVOT sheets.")
        target_filter_states = capture_target_pivot_filters(pivot_sheet)
        if not target_filter_states:
            raise RuntimeError("Previous workbook has no target PivotTable page filters.")

        row_count = len(df) + 1
        column_count = len(df.columns)
        table = all_order_sheet.ListObjects(TABLE_NAME)
        old_last_row = max(
            int(all_order_sheet.UsedRange.Rows.Count),
            int(table.Range.Rows.Count),
        )
        clear_last_row = max(old_last_row, row_count)
        all_order_sheet.Range(
            all_order_sheet.Cells(2, 1),
            all_order_sheet.Cells(clear_last_row, column_count),
        ).ClearContents()

        target_range = all_order_sheet.Range(
            all_order_sheet.Cells(1, 1),
            all_order_sheet.Cells(row_count, column_count),
        )
        table.Resize(target_range)
        table.TableStyle = TABLE_STYLE_NAME
        year_fab_upload_column = df.columns.get_loc(YEAR_FAB_UPLOAD_HEADER) + 1
        all_order_sheet.Range(
            all_order_sheet.Cells(2, year_fab_upload_column),
            all_order_sheet.Cells(row_count, year_fab_upload_column),
        ).NumberFormat = YEAR_TEXT_FORMAT
        body_values = tuple(
            tuple(com_value(value) for value in row)
            for row in df.itertuples(index=False, name=None)
        )
        all_order_sheet.Range(
            all_order_sheet.Cells(2, 1),
            all_order_sheet.Cells(row_count, column_count),
        ).Value = body_values
        for column_number, header in enumerate(df.columns, start=1):
            all_order_sheet.Cells(1, column_number).Value = str(header)
        apply_na_errors(excel, all_order_sheet, df)
        if old_last_row > row_count:
            all_order_sheet.Rows(f"{row_count + 1}:{old_last_row}").Delete()

        target_range.Font.Name = FONT_NAME
        target_range.Font.Size = FONT_SIZE
        target_range.HorizontalAlignment = -4108  # xlCenter
        target_range.VerticalAlignment = -4108  # xlCenter

        otc_column = df.columns.get_loc(OTC_HEADER) + 1
        mrc_column = df.columns.get_loc(MRC_HEADER) + 1
        all_order_sheet.Range(
            all_order_sheet.Cells(2, otc_column),
            all_order_sheet.Cells(row_count, mrc_column),
        ).NumberFormat = "#,##0"

        all_order_sheet.Range(
            all_order_sheet.Cells(2, year_fab_upload_column),
            all_order_sheet.Cells(row_count, year_fab_upload_column),
        ).NumberFormat = YEAR_TEXT_FORMAT

        apply_date_formats(all_order_sheet, df)

        formula_column = df.columns.get_loc(ON_TIME_DELIVERY_HEADER) + 1
        actual_rfs_column = df.columns.get_loc(ACTUAL_RFS_DATE_HEADER) + 1
        initial_rfs_column = df.columns.get_loc(NEW_RFS_INITIAL_HEADER) + 1
        actual_offset = actual_rfs_column - formula_column
        initial_offset = initial_rfs_column - formula_column
        formula_values = []
        formula_rows: list[bool] = []
        for actual_rfs, initial_rfs in zip(
            df[ACTUAL_RFS_DATE_HEADER],
            df[NEW_RFS_INITIAL_HEADER],
            strict=False,
        ):
            has_required_dates = isinstance(actual_rfs, datetime) and isinstance(
                initial_rfs, datetime
            )
            formula_rows.append(has_required_dates)
            if has_required_dates:
                formula_values.append(
                    (
                        f'=IF(RC[{actual_offset}]>RC[{initial_offset}],'
                        '"DELAY","ONTIME")',
                    )
                )
            else:
                formula_values.append(("",))
        formula_range = all_order_sheet.Range(
            all_order_sheet.Cells(2, formula_column),
            all_order_sheet.Cells(row_count, formula_column),
        )
        formula_range.FormulaR1C1 = tuple(formula_values)
        blank_formula_rows = [
            row_number
            for row_number, has_formula in enumerate(formula_rows, start=2)
            if not has_formula
        ]
        if blank_formula_rows:
            start_row = previous_row = blank_formula_rows[0]
            for row_number in [*blank_formula_rows[1:], None]:
                if row_number is None or row_number != previous_row + 1:
                    all_order_sheet.Range(
                        all_order_sheet.Cells(start_row, formula_column),
                        all_order_sheet.Cells(previous_row, formula_column),
                    ).ClearContents()
                    if row_number is None:
                        break
                    start_row = row_number
                previous_row = row_number

        autofit_with_limits(all_order_sheet, column_count)
        all_order_sheet.Activate()
        excel.ActiveWindow.FreezePanes = False
        excel.ActiveWindow.SplitColumn = 2
        excel.ActiveWindow.SplitRow = 0
        excel.ActiveWindow.FreezePanes = True

        percentage_scratch, percentage_blocks = capture_pivot_percentage_blocks(
            workbook,
            pivot_sheet,
        )

        refreshed_count = 0
        pivot_tables = pivot_sheet.PivotTables()
        for pivot_index in range(1, pivot_tables.Count + 1):
            pivot_table = pivot_tables(pivot_index)
            try:
                pivot_cache = pivot_table.PivotCache()
                source_data = str(pivot_cache.SourceData)
                if TABLE_NAME.casefold() not in source_data.casefold():
                    raise RuntimeError(
                        f"PivotTable source is {source_data!r}, expected {TABLE_NAME}."
                    )
                pivot_cache.Refresh()
                pivot_table.RefreshTable()
                refreshed_count += 1
            except Exception as exc:
                raise RuntimeError(
                    f"Could not refresh PivotTable '{pivot_table.Name}': {exc}"
                ) from exc

        restore_target_pivot_filters(
            pivot_sheet,
            target_filter_states,
            report_date,
        )
        restore_pivot_percentage_blocks(
            pivot_sheet,
            percentage_scratch,
            percentage_blocks,
        )
        percentage_scratch.Visible = -1  # xlSheetVisible
        percentage_scratch.Delete()
        percentage_scratch = None
        repair_pivot_percentage_ranges(pivot_sheet)
        apply_weekly_completion_thresholds(
            pivot_sheet,
            percentage_blocks,
            report_date,
        )
        normalize_pivot_percentage_borders(
            pivot_sheet,
            percentage_blocks,
        )
        update_pivot_period_titles(pivot_sheet, report_date)
        resize_dynamic_pivot_section_banners(
            pivot_sheet,
            title_markers=PIVOT_SECTION_TITLE_MARKERS,
            side_header_markers=PIVOT_SECTION_SIDE_HEADER_MARKERS,
        )
        normalize_pivot_section_row_visibility(pivot_sheet)
        adjust_pivot_column_widths(pivot_sheet)
        validate_target_complete_pivots(pivot_sheet, df)

        try:
            excel.Calculation = -4105  # xlCalculationAutomatic
        except Exception:
            pass
        excel.Calculate()
        try:
            formula_errors = pivot_sheet.UsedRange.SpecialCells(-4123, 16)  # formulas, errors
            error_addresses = [
                area.Address for area in formula_errors.Areas
            ]
        except Exception:
            error_addresses = []
        if error_addresses:
            raise RuntimeError(
                "PIVOT contains formula errors after refresh: " + ", ".join(error_addresses)
            )
        workbook.Save()
        return refreshed_count
    finally:
        if workbook is not None:
            try:
                workbook.Close(SaveChanges=False)
            except Exception:
                pass
        if excel is not None:
            try:
                if previous_calculation is not None:
                    excel.Calculation = previous_calculation
            except Exception:
                pass
            try:
                excel.Quit()
            except Exception:
                pass
        pythoncom.CoUninitialize()


def validate_output(
    output_path: Path,
    expected_quote_ids: list[str],
    expected_pivot_count: int,
) -> dict[str, object]:
    workbook = load_workbook(output_path, read_only=False, data_only=False, keep_links=False)
    try:
        missing_sheets = {
            ALL_ORDER_SHEET_NAME,
            PIVOT_SHEET_NAME,
        } - set(workbook.sheetnames)
        if missing_sheets:
            raise ValueError(f"Output is missing sheets: {', '.join(sorted(missing_sheets))}")

        worksheet = workbook[ALL_ORDER_SHEET_NAME]
        if worksheet.max_column != 61:
            raise ValueError(f"Output ALL ORDER has {worksheet.max_column} columns, expected 61.")
        actual_rows = worksheet.max_row - 1
        expected_rows = len(expected_quote_ids)
        if actual_rows != expected_rows:
            raise ValueError(f"Output has {actual_rows} data rows, expected {expected_rows}.")
        if worksheet.freeze_panes != "C1":
            raise ValueError(f"Output freeze pane is {worksheet.freeze_panes}, expected C1.")
        if TABLE_NAME not in worksheet.tables:
            raise ValueError(f"Output is missing Excel table: {TABLE_NAME}")
        expected_ref = f"A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"
        if worksheet.tables[TABLE_NAME].ref != expected_ref:
            raise ValueError(
                f"{TABLE_NAME} range is {worksheet.tables[TABLE_NAME].ref}, expected {expected_ref}."
            )

        headers = {worksheet.cell(1, column).value: column for column in range(1, 62)}
        quote_column = headers[QUOTE_ID_HEADER]
        otc_column = headers[OTC_HEADER]
        mrc_column = headers[MRC_HEADER]
        year_fab_upload_column = headers[YEAR_FAB_UPLOAD_HEADER]
        actual_quote_ids: list[str] = []
        invalid_financial_values = 0
        invalid_year_values = 0
        for row_number in range(2, worksheet.max_row + 1):
            actual_quote_ids.append(
                normalize_quote_id(worksheet.cell(row_number, quote_column).value)
            )
            for column in (otc_column, mrc_column):
                if numeric_value(worksheet.cell(row_number, column).value) is None:
                    invalid_financial_values += 1
            year_cell = worksheet.cell(row_number, year_fab_upload_column)
            year_text = str(year_cell.value).strip()
            if (
                not re.fullmatch(r"\d{4}", year_text)
                or not MIN_VALID_DATE_YEAR <= int(year_text) <= MAX_VALID_DATE_YEAR
                or year_cell.number_format != YEAR_TEXT_FORMAT
            ):
                invalid_year_values += 1
        if actual_quote_ids != expected_quote_ids:
            mismatch_index = next(
                (
                    index
                    for index, (actual, expected) in enumerate(
                        zip(actual_quote_ids, expected_quote_ids, strict=False),
                        start=2,
                    )
                    if actual != expected
                ),
                None,
            )
            raise ValueError(
                "Output Quote ID sequence does not match the filtered raw data"
                + (f" at Excel row {mismatch_index}." if mismatch_index else ".")
            )
        if invalid_financial_values:
            raise ValueError(
                f"Output contains {invalid_financial_values} invalid OTC/MRC values."
            )
        if invalid_year_values:
            raise ValueError(
                f"Output contains {invalid_year_values} invalid or date-formatted "
                f"{YEAR_FAB_UPLOAD_HEADER} values."
            )

        pivot_count = len(workbook[PIVOT_SHEET_NAME]._pivots)
        if pivot_count != expected_pivot_count:
            raise ValueError(
                f"Output contains {pivot_count} PivotTables, expected {expected_pivot_count}."
            )
        return {
            "rows": actual_rows,
            "columns": worksheet.max_column,
            "pivot_tables": pivot_count,
            "table_ref": expected_ref,
        }
    finally:
        workbook.close()


def generate_ide_tracking(
    files: InputFiles,
    output_path: Path,
    report_date: date,
) -> dict[str, object]:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    raw = read_raw_dataframe(files.raw)
    previous = read_previous_dataframe(files.previous)
    build = build_all_order(raw, previous, files.collabs, report_date, files.previous)

    temp_path: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
            prefix="ide_tracking_",
            suffix=".xlsx",
            dir=output_path.parent,
            delete=False,
        ) as temp_file:
            temp_path = Path(temp_file.name)
        shutil.copy2(files.previous, temp_path)
        refreshed_count = update_workbook_via_com(temp_path, build.dataframe, report_date)
        validation = validate_output(
            temp_path,
            build.dataframe[QUOTE_ID_HEADER].map(normalize_quote_id).tolist(),
            refreshed_count,
        )
        try:
            os.replace(temp_path, output_path)
        except PermissionError as exc:
            raise PermissionError(
                f"Cannot replace output workbook because it is open or locked: {output_path}. "
                "Close the workbook in Excel and run the command again."
            ) from exc
        temp_path = None
    finally:
        if temp_path is not None:
            try:
                temp_path.unlink(missing_ok=True)
            except OSError:
                pass

    return {
        **validation,
        "refreshed_pivots": refreshed_count,
        "collabs_fallback_rows": build.collabs_fallback_count,
        "zero_fallback_rows": build.zero_fallback_count,
        "duplicate_quote_ids": build.duplicate_quote_ids,
        "duplicate_rows_preserved": build.duplicate_rows_preserved,
        "date_fallback_rows": build.date_fallback_count,
        "field_fallback_rows": build.field_fallback_count,
        "invalid_dates": build.invalid_dates,
    }


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate Daily Tracking IDE from the IDE dashboard and previous workbook.",
    )
    parser.add_argument(
        "input",
        nargs="?",
        type=Path,
        default=Path(DEFAULT_INPUT_DIR),
        help=f"Raw IDE workbook or directory. Defaults to .\\{DEFAULT_INPUT_DIR}.",
    )
    parser.add_argument(
        "-r",
        "--reference",
        type=Path,
        default=Path(DEFAULT_REFERENCE_DIR),
        help=f"Previous Daily Tracking IDE workbook or directory. Defaults to .\\{DEFAULT_REFERENCE_DIR}.",
    )
    parser.add_argument(
        "-c",
        "--collabs",
        type=Path,
        default=Path(DEFAULT_COLLABS_DIR),
        help=f"Collabs CSV fallback or directory. Defaults to .\\{DEFAULT_COLLABS_DIR}.",
    )
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=Path(DEFAULT_OUTPUT_DIR),
        help=f"Output workbook or directory. Defaults to .\\{DEFAULT_OUTPUT_DIR}.",
    )
    parser.add_argument(
        "--report-date",
        type=parse_report_date,
        help="Explicit report date using YYYY-MM-DD. Auto-detected when omitted.",
    )
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv or sys.argv[1:])
    try:
        files = resolve_inputs(args)
        report_date = determine_report_date(args.report_date, files.raw)
        output_path = resolve_output_path(args.output, report_date)
        print(f"[1/5] Reading raw data: {files.raw.name}")
        print(f"[2/5] Loading previous mapping: {files.previous.name}")
        print(f"[3/5] Loading OTC/MRC fallback: {files.collabs.name}")
        print("[4/5] Updating ALL ORDER and refreshing PivotTables in Excel")
        result = generate_ide_tracking(files, output_path, report_date)
        print("[5/5] Validating output")
        print(
            f"Wrote {output_path} | rows={result['rows']} | columns={result['columns']} | "
            f"pivots={result['pivot_tables']}"
        )
        print(
            f"Fallbacks: collabs={result['collabs_fallback_rows']}, "
            f"OTC/MRC zero={result['zero_fallback_rows']}, "
            f"dates={result['date_fallback_rows']}, "
            f"fields={result['field_fallback_rows']}"
        )
        duplicate_quote_ids = result["duplicate_quote_ids"]
        if duplicate_quote_ids:
            print(
                f"Preserved {result['duplicate_rows_preserved']} repeated rows across "
                f"{len(duplicate_quote_ids)} Quote IDs; lookups are matched by occurrence."
            )
        invalid_dates = result["invalid_dates"]
        if invalid_dates:
            print("[warn] Invalid source date values were preserved as text:", file=sys.stderr)
            for header, values in invalid_dates.items():
                print(f"  - {header}: {', '.join(values)}", file=sys.stderr)
        return 0
    except Exception as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())

"""Capture and email the Report 4 IDE PIVOT sheet."""

from __future__ import annotations

import argparse
from pathlib import Path

from send_report_1_email import excel_range_to_png, send_email_images

REPORT_4_IMAGE_IDS = (
    "report-4-complete-table",
    "report-4-complete-legend",
    "report-4-after-table",
    "report-4-after-legend",
)


def report_4_ranges(workbook_path: Path) -> list[tuple[str, str]]:
    from openpyxl import load_workbook

    sheet = load_workbook(workbook_path, read_only=True, data_only=False)["PIVOT"]
    rows = {}
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            value = str(cell.value).strip().upper()
            if "TARGET COMPLETE" in value:
                rows.setdefault("complete", cell.row)
            elif "TARGET AFTER" in value:
                rows.setdefault("after", cell.row)
            elif "TARGET NOT INPUTTED" in value:
                rows.setdefault("not_inputted", cell.row)
    if "complete" not in rows or "after" not in rows:
        raise RuntimeError("Could not find TARGET COMPLETE and TARGET AFTER on the Report 4 PIVOT sheet.")
    return [
        (f"A{rows['complete']}:L{rows['after'] - 1}", REPORT_4_IMAGE_IDS[0]),
        ("M10:M12", REPORT_4_IMAGE_IDS[1]),
        (f"A{rows['after']}:Q{rows.get('not_inputted', sheet.max_row + 1) - 1}", REPORT_4_IMAGE_IDS[2]),
        ("R307:R309", REPORT_4_IMAGE_IDS[3]),
    ]


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, required=True)
    parser.add_argument("--image", type=Path)
    parser.add_argument("--endpoint", default=None)
    parser.add_argument("--to", default="ilhaam.akmal@lintasarta.co.id")
    parser.add_argument("--subject", default="Daily Tracking Report 4 IDE")
    parser.add_argument("--message", default='<p>Daily Tracking Report 4 IDE</p><p><img src="cid:report-4-pivot" alt="Report 4 IDE PIVOT"></p>')
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()
    workbook = args.workbook.resolve()
    if not workbook.exists():
        parser.error(f"Workbook not found: {workbook}")
    image_dir = (args.image or workbook.with_name(f"{workbook.stem} - report-4-images")).resolve()
    ranges = report_4_ranges(workbook)
    images = []
    for index, (cell_range, content_id) in enumerate(ranges, start=1):
        image_path = image_dir.parent / f"{image_dir.name}-{index}.png"
        excel_range_to_png(workbook, image_path, "PIVOT", cell_range, scale=3.0 if "legend" in content_id else 1.0)
        images.append((image_path, content_id))
        print(f"Created image: {image_path} ({cell_range})")
    if args.dry_run:
        print("Dry run: SMTP request skipped.")
        return 0
    send_email_images(args.endpoint or "http://10.34.144.197/secm-portal/smtp/api_send_email", args.to, args.subject, args.message, images, timeout=60)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

#!/usr/bin/env python3
"""Generate ongoing tracking workbook from input-ongoing sources."""

from __future__ import annotations

import argparse
import csv
import os
import re
import shutil
import sys
import time
from datetime import date, datetime, timedelta
from pathlib import Path

import pandas as pd
from charset_normalizer import from_path
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries


DEFAULT_INPUT_DIR = "input-ongoing"
DEFAULT_OUTPUT_DIR = "output-outgoing"
DEFAULT_EXAMPLE_DIR = "example-ongoing"
DEFAULT_VALIDATE_DIR = "validate"
TRACKING_GLOB = "Daily Tracking *.xlsx"
LOG_GLOB = "LogUpdateStatusOrderSD-*.csv"
EXAMPLE_TRACKING_GLOB = "Daily Tracking *.xlsx"
SHEET_NAME = "ALL ORDER"
PIVOT_SHEET_NAME = "PIVOT"
ON_PROGRESS_SHEET_PREFIX = "ALL ORDER ON PROGRESS"

CATEGORY_MLD_HEADER = "Category MLD"
QUO_HEADER = "Quo"
YEAR_FAB_UPLOAD_HEADER = "YEAR FAB UPLOAD"
PROCESS_ADJUSTMENT_HEADER = "Process Adjustment"
PHASE_HEADER_PREFIX = "Phase"
STATUS_ORDER_HEADER = "STATUS ORDER"
PRE_INSTALLATION_START_DATE_HEADER = "Pre-Installation Start Date"
PRE_INSTALLATION_AGING_HEADER = "Pre-Installation Aging"
RANGE_AGING_PRE_INSTALLATION_HEADER = "Range Aging on Pre-Installation"
EXCEL_RED_FONT_COLOR = "FF0000"
EXCEL_YELLOW_HEADER_FILL = "FFFF00"
EXCEL_ON_PROGRESS_TAB_COLOR = 15128749  # RGB(173,216,230) light blue for Excel COM

NULL_LIKE_VALUES = {"", "na", "n/a", "null", "none", "nan", "-", "(blank)"}
COMMON_ENCODINGS = ("utf-8-sig", "utf-8", "cp1252", "latin1")
MIN_REPORTING_YEAR_FAB_UPLOAD = 2023


def profile_log(label: str, start: float) -> None:
    if os.getenv("ONGOING_PROFILE", "").strip().lower() in {"1", "true", "yes"}:
        print(f"[profile] {label}: {time.perf_counter() - start:.2f}s", file=sys.stderr)


def normalize_header_key(header_value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(header_value).lower())


DATE_LIKE_HEADER_KEY_TERMS = (
    "quotedate",
    "sodate",
    "mlddate",
    "fabupload",
    "fabsign",
    "rfscustomer",
    "initialrfs",
    "rfscommmit",
    "rfscommit",
    "surveydate",
    "estimateendsurveydate",
    "allocationdate",
    "bacreate",
    "basign",
    "installdate",
    "canceldate",
    "compackdate",
    "linkready",
    "completeddate",
    "filecommpackuploaddate",
    "targetsocompletedate",
    "targetsocompletiondate",
    "targetonholdduration",
)


def should_preserve_excel_date_values(normalized_header: str) -> bool:
    if normalized_header in {
        normalize_header_key(YEAR_FAB_UPLOAD_HEADER),
        normalize_header_key("Aging Of RFS"),
        normalize_header_key(STATUS_ORDER_HEADER),
    }:
        return False

    return "date" in normalized_header or any(term in normalized_header for term in DATE_LIKE_HEADER_KEY_TERMS)


def excel_date_number_format_for_header(normalized_header: str, source_number_format: object = None) -> str:
    source_format = str(source_number_format or "").strip()
    if source_format and source_format.lower() != "general":
        return source_format

    if any(term in normalized_header for term in {"preinstallationstartdate", "allocationdate", "completeddate"}):
        return "dd/mm/yyyy hh:mm"

    return "dd/mm/yyyy"


def normalize_all_order_date_column_formats(worksheet) -> None:
    try:
        used_rows = worksheet.UsedRange.Rows.Count
        used_cols = worksheet.UsedRange.Columns.Count
        headers_raw = worksheet.Range(worksheet.Cells(1, 1), worksheet.Cells(1, used_cols)).Value
        headers = list(headers_raw[0] if isinstance(headers_raw, tuple) and isinstance(headers_raw[0], tuple) else headers_raw)
    except Exception:
        return

    for column_index, header_value in enumerate(headers, start=1):
        normalized_header = normalize_header_key(header_value)
        if not should_preserve_excel_date_values(normalized_header):
            continue
        try:
            worksheet.Range(
                worksheet.Cells(2, column_index),
                worksheet.Cells(used_rows, column_index),
            ).NumberFormat = excel_date_number_format_for_header(normalized_header)
            if normalized_header == normalize_header_key(PRE_INSTALLATION_START_DATE_HEADER):
                worksheet.Columns(column_index).ColumnWidth = max(
                    worksheet.Columns(column_index).ColumnWidth,
                    18,
                )
        except Exception:
            pass


def is_excluded_year_fab_upload_value(value: object) -> bool:
    normalized = str(value).strip()
    if normalized.lower() in {"", "1900", "1900.0", "nan", "none", "n/a", "(blank)", "null"}:
        return True

    if re.fullmatch(r"\d{4}(?:\.0+)?", normalized):
        return int(float(normalized)) < MIN_REPORTING_YEAR_FAB_UPLOAD

    return False


def normalize_quo_value(value: object) -> str:
    normalized = re.sub(r"\s+", " ", str(value).strip())
    if not normalized or normalized.lower() in NULL_LIKE_VALUES:
        return ""
    return normalized


def detect_encoding(path: Path) -> str:
    result = from_path(path).best()
    if result and result.encoding:
        return result.encoding

    for encoding in COMMON_ENCODINGS:
        try:
            path.read_text(encoding=encoding)
            return encoding
        except UnicodeDecodeError:
            continue

    return "latin1"


def detect_delimiter(path: Path, encoding: str) -> str:
    sample = path.read_text(encoding=encoding, errors="replace")[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        return dialect.delimiter
    except csv.Error:
        return ","


def resolve_ongoing_input_files(input_path: Path) -> tuple[Path, Path]:
    if not input_path.is_dir():
        raise FileNotFoundError(f"Input path does not exist: {input_path}")

    tracking_files = sorted(
        (
            path
            for path in input_path.glob(TRACKING_GLOB)
            if path.is_file() and not path.name.startswith("~$")
        ),
        key=lambda path: path.stat().st_mtime,
    )
    log_files = sorted(
        (path for path in input_path.glob(LOG_GLOB) if path.is_file()),
        key=lambda path: path.stat().st_mtime,
    )

    if not tracking_files:
        raise FileNotFoundError(f"No tracking workbook found with pattern {TRACKING_GLOB} in {input_path}")
    if not log_files:
        raise FileNotFoundError(f"No log CSV found with pattern {LOG_GLOB} in {input_path}")

    return tracking_files[-1], log_files[-1]


def resolve_example_source_workbook(base_dir: Path) -> Path:
    example_dir = base_dir / DEFAULT_EXAMPLE_DIR
    if not example_dir.is_dir():
        raise FileNotFoundError(f"Example source directory not found: {example_dir}")

    files = sorted(
        (
            path
            for path in example_dir.glob(EXAMPLE_TRACKING_GLOB)
            if path.is_file() and not path.name.startswith("~$")
        ),
        key=lambda path: path.stat().st_mtime,
    )
    if not files:
        raise FileNotFoundError(f"No example workbook found in: {example_dir}")

    return files[-1]


def resolve_validate_source_workbook(base_dir: Path) -> Path:
    validate_dir = base_dir / DEFAULT_VALIDATE_DIR
    if not validate_dir.is_dir():
        raise FileNotFoundError(f"Validate source directory not found: {validate_dir}")

    files = sorted(
        (
            path
            for path in validate_dir.glob(EXAMPLE_TRACKING_GLOB)
            if path.is_file() and not path.name.startswith("~$")
        ),
        key=lambda path: path.stat().st_mtime,
    )
    if not files:
        raise FileNotFoundError(f"No validate workbook found in: {validate_dir}")
    return files[-1]


def reference_date_from_tracking_workbook_path(tracking_workbook_path: Path) -> date:
    match = re.search(r"(\d{1,2})\s+([A-Za-z]+)\s+(\d{4})", tracking_workbook_path.stem, flags=re.IGNORECASE)
    if not match:
        return date.today()

    try:
        return datetime.strptime(" ".join(match.groups()), "%d %B %Y").date()
    except ValueError:
        return date.today()


def output_filename_for_tracking(tracking_workbook_path: Path) -> str:
    reference_date = reference_date_from_tracking_workbook_path(tracking_workbook_path)
    return f"Daily Tracking {reference_date.day} {reference_date.strftime('%B %Y')} On Going.xlsx"


def expected_on_progress_sheet_name(reference_date: date) -> str:
    return f"{ON_PROGRESS_SHEET_PREFIX} {reference_date.day} {reference_date.strftime('%B')}"


def fallback_target_determined_header(reference_date: date) -> str:
    return f"TARGET  Detemined as 1 {reference_date.strftime('%b %y')}"


def fallback_dated_phase_header(reference_date: date) -> str:
    phase_date = reference_date - timedelta(days=1)
    return f"Phase {phase_date.day} {phase_date.strftime('%B %Y')}"


def resolve_output_path(input_path: Path, tracking_workbook_path: Path, requested_output: Path | None) -> Path:
    default_filename = output_filename_for_tracking(tracking_workbook_path)
    if requested_output is None:
        return input_path.parent / DEFAULT_OUTPUT_DIR / default_filename

    if requested_output.suffix.lower() == ".xlsx":
        return requested_output

    return requested_output / default_filename


def match_column_name(columns: list[object], accepted_normalized_names: set[str]) -> str | None:
    for column in columns:
        if normalize_header_key(column) in accepted_normalized_names:
            return str(column)
    return None


def load_pre_installation_lookup(log_csv_path: Path) -> dict[str, str]:
    encoding = detect_encoding(log_csv_path)
    delimiter = detect_delimiter(log_csv_path, encoding)
    lookup_df = pd.read_csv(
        log_csv_path,
        dtype="string",
        encoding=encoding,
        sep=delimiter,
        keep_default_na=False,
    )

    quo_column = match_column_name(lookup_df.columns.tolist(), {"quo", "#quo"})
    start_date_column = match_column_name(lookup_df.columns.tolist(), {"startdate", "preinstallationstartdate"})
    phase_column = match_column_name(lookup_df.columns.tolist(), {"phase"})
    if quo_column is None or start_date_column is None:
        raise ValueError(
            f"Required columns missing in {log_csv_path.name}. Expected Quo/#Quo and Start Date columns."
        )

    if phase_column is not None:
        phase_values = lookup_df[phase_column].astype("string").map(normalize_header_key)
        lookup_df = lookup_df.loc[phase_values.eq("preinstallation")]

    lookup: dict[str, str] = {}
    for _, row in lookup_df[[quo_column, start_date_column]].iterrows():
        # Preserve the log #Quo exactly. The manual workbook uses an exact lookup,
        # so log keys with trailing spaces intentionally do not match ALL ORDER.
        quo_value = str(row[quo_column]).replace("\ufeff", "")
        start_date_value = str(row[start_date_column]).strip()
        if (
            quo_value
            and quo_value not in lookup
            and start_date_value
            and start_date_value.lower() not in NULL_LIKE_VALUES
        ):
            lookup[quo_value] = start_date_value

    return lookup


def build_pre_installation_range(aging_value: int | None) -> str | None:
    if aging_value is None:
        return None
    if aging_value < 15:
        return "1-7 Days"
    if aging_value < 31:
        return "8-14 Days"
    if aging_value < 61:
        return "15-30 Days"
    return "> 30 Days"


def normalized_pivot_item_name(value: object) -> str:
    return re.sub(r"\s+", " ", str(value).strip().lower())


def phase_code_from_pivot_item_name(value: object) -> str:
    normalized = normalized_pivot_item_name(value)
    digit_prefix_match = re.match(r"^(\d{1,2})\b", normalized)
    if digit_prefix_match:
        return digit_prefix_match.group(1).zfill(2)

    anywhere_match = re.search(r"\b(\d{1,2})\b", normalized)
    if anywhere_match:
        return anywhere_match.group(1).zfill(2)
    return ""


def find_pivot_field(pivot_table, accepted_normalized_names: set[str]):
    try:
        pivot_fields = pivot_table.PivotFields()
        for field_index in range(1, pivot_fields.Count + 1):
            field = pivot_fields(field_index)
            if normalize_header_key(field.Name) in accepted_normalized_names:
                return field
    except Exception:
        return None
    return None


def set_pivot_field_visibility(field, visible_predicate) -> None:
    visible_items = []
    hidden_items = []
    for item in field.PivotItems():
        try:
            if visible_predicate(item):
                visible_items.append(item)
            else:
                hidden_items.append(item)
        except Exception:
            hidden_items.append(item)

    if not visible_items:
        for item in field.PivotItems():
            try:
                item.Visible = True
            except Exception:
                pass
        return

    for item in visible_items:
        try:
            item.Visible = True
        except Exception:
            pass
    for item in hidden_items:
        try:
            item.Visible = False
        except Exception:
            pass


def apply_range_aging_column_order(pivot_table) -> None:
    """Keep aging buckets in business order after Excel rebuilds the pivot cache."""
    range_field = find_pivot_field(
        pivot_table,
        {normalize_header_key(RANGE_AGING_PRE_INSTALLATION_HEADER)},
    )
    if range_field is None:
        return

    ordered_items = ["1-7 Days", "8-14 Days", "15-30 Days", "> 30 Days", "#N/A"]
    for position, item_name in enumerate(ordered_items, start=1):
        try:
            range_field.PivotItems(item_name).Position = position
        except Exception:
            pass


def apply_common_pivot_filters(pivot_table) -> None:
    xl_page_field = 3

    year_field = find_pivot_field(pivot_table, {normalize_header_key(YEAR_FAB_UPLOAD_HEADER)})
    if year_field is not None:
        try:
            hidden_year_items = set()
            try:
                for item in year_field.PivotItems():
                    if not item.Visible:
                        hidden_year_items.add(normalized_pivot_item_name(item.Name))
            except Exception:
                hidden_year_items = set()
            year_field.Orientation = xl_page_field
            year_field.Position = 1
            year_field.EnableMultiplePageItems = True
            set_pivot_field_visibility(
                year_field,
                lambda item: (
                    normalized_pivot_item_name(item.Name) not in hidden_year_items
                    and normalized_pivot_item_name(item.Name) not in {"1900", "1900.0"}
                ),
            )
        except Exception:
            pass

    process_adjustment_field = find_pivot_field(pivot_table, {normalize_header_key(PROCESS_ADJUSTMENT_HEADER)})
    if process_adjustment_field is not None:
        try:
            process_adjustment_field.Orientation = xl_page_field
            process_adjustment_field.Position = 2
            process_adjustment_field.EnableMultiplePageItems = True
            set_pivot_field_visibility(process_adjustment_field, lambda _item: True)
        except Exception:
            pass


def apply_phase_filter_excluding_cancel_so_complete(pivot_table) -> None:
    xl_page_field = 3
    excluded_values = {"cancel", "so complete"}
    phase_field = find_pivot_field(pivot_table, {normalize_header_key(PHASE_HEADER_PREFIX)})
    if phase_field is None:
        return

    try:
        phase_field.Orientation = xl_page_field
        phase_field.Position = 3
        phase_field.EnableMultiplePageItems = True
        set_pivot_field_visibility(
            phase_field,
            lambda item: all(excluded not in normalized_pivot_item_name(item.Name) for excluded in excluded_values),
        )
    except Exception:
        pass


def apply_phase_filter_04_to_07(pivot_table) -> None:
    xl_page_field = 3
    phase_field = find_pivot_field(pivot_table, {normalize_header_key(PHASE_HEADER_PREFIX)})
    if phase_field is None:
        return

    try:
        phase_field.Orientation = xl_page_field
        phase_field.Position = 3
        phase_field.EnableMultiplePageItems = True
        set_pivot_field_visibility(
            phase_field,
            lambda item: phase_code_from_pivot_item_name(item.Name) in {"04", "05", "06", "07"},
        )
    except Exception:
        pass


def is_target_determined_header(header_value: object) -> bool:
    normalized = normalize_header_key(header_value)
    return normalized.startswith("target") and (
        "deteminedas1" in normalized or "determinedas1" in normalized
    )


def is_dated_phase_header(header_value: object) -> bool:
    text = str(header_value).strip()
    if normalize_header_key(text) == normalize_header_key(PHASE_HEADER_PREFIX):
        return False
    return bool(re.match(r"^phase\s+\d{1,2}\s+[A-Za-z]+\s+\d{4}$", text, flags=re.IGNORECASE))


def date_from_dated_phase_header(header_value: object) -> date | None:
    match = re.match(r"^phase\s+(\d{1,2}\s+[A-Za-z]+\s+\d{4})$", str(header_value).strip(), flags=re.IGNORECASE)
    if not match:
        return None
    try:
        return datetime.strptime(match.group(1), "%d %B %Y").date()
    except ValueError:
        return None


def find_source_dynamic_headers(source_headers: list[object], reference_date: date) -> tuple[str, str]:
    target_header = None
    phase_header = None
    expected_phase_date = reference_date - timedelta(days=1)
    for header_value in source_headers:
        if header_value is None:
            continue
        if target_header is None and is_target_determined_header(header_value):
            target_header = str(header_value)
        if (
            phase_header is None
            and is_dated_phase_header(header_value)
            and date_from_dated_phase_header(header_value) == expected_phase_date
        ):
            phase_header = str(header_value)
        if target_header is not None and phase_header is not None:
            break
    return (
        target_header or fallback_target_determined_header(reference_date),
        phase_header or fallback_dated_phase_header(reference_date),
    )


def synchronize_dynamic_headers_on_com_sheet(target_ws, target_table_cols: int, target_header: str, phase_header: str) -> None:
    try:
        target_headers_raw = target_ws.Range(target_ws.Cells(1, 1), target_ws.Cells(1, target_table_cols)).Value
        if isinstance(target_headers_raw, tuple):
            target_headers = list(
                target_headers_raw[0]
                if target_headers_raw and isinstance(target_headers_raw[0], tuple)
                else target_headers_raw
            )
        else:
            target_headers = [target_headers_raw]
    except Exception:
        return

    for column_index, header_value in enumerate(target_headers, start=1):
        if is_target_determined_header(header_value):
            target_ws.Cells(1, column_index).Value = target_header
        elif is_dated_phase_header(header_value):
            target_ws.Cells(1, column_index).Value = phase_header


def apply_sales_hierarchy_header_style_on_com_sheet(worksheet) -> None:
    header_labels = {
        "groupsales": "group_sales",
        "divisionsales": "Division_Sales",
        "segmentsales": "segment_sales",
        "sales": "Sales",
    }
    try:
        used_cols = worksheet.UsedRange.Columns.Count
        headers_raw = worksheet.Range(worksheet.Cells(1, 1), worksheet.Cells(1, used_cols)).Value
        if isinstance(headers_raw, tuple):
            headers = list(headers_raw[0] if headers_raw and isinstance(headers_raw[0], tuple) else headers_raw)
        else:
            headers = [headers_raw]
    except Exception:
        return

    for column_index, header_value in enumerate(headers, start=1):
        normalized = normalize_header_key(header_value)
        replacement = header_labels.get(normalized)
        if replacement is None:
            continue

        try:
            cell = worksheet.Cells(1, column_index)
            cell.Value = replacement
            cell.Interior.Color = 255  # red
            cell.Font.Color = 0  # black
            cell.Font.Bold = True
            cell.HorizontalAlignment = -4108  # center
            cell.VerticalAlignment = -4108
        except Exception:
            pass


def count_non_empty_column_values(data_sheet, column_index: int) -> int:
    xl_up = -4162
    try:
        last_row = data_sheet.Cells(data_sheet.Rows.Count, 1).End(xl_up).Row
    except Exception:
        return 0
    if last_row < 2:
        return 0
    try:
        values = data_sheet.Range(data_sheet.Cells(2, column_index), data_sheet.Cells(last_row, column_index)).Value
    except Exception:
        return 0

    if isinstance(values, tuple):
        rows = values
    else:
        rows = (values,)

    non_empty_count = 0
    for row_value in rows:
        value = row_value[0] if isinstance(row_value, tuple) else row_value
        if value not in (None, ""):
            non_empty_count += 1
    return non_empty_count


def resolve_top_pivot_row_field_for_reference_date(data_sheet, reference_date: date) -> str | None:
    try:
        used_cols = data_sheet.UsedRange.Columns.Count
        headers_raw = data_sheet.Range(data_sheet.Cells(1, 1), data_sheet.Cells(1, used_cols)).Value
        if isinstance(headers_raw, tuple):
            headers = list(headers_raw[0] if headers_raw and isinstance(headers_raw[0], tuple) else headers_raw)
        else:
            headers = [headers_raw]
    except Exception:
        return None

    candidates: list[tuple[str, int]] = []
    fallback_candidates: list[tuple[str, int]] = []
    for column_index, header_value in enumerate(headers, start=1):
        if header_value is None:
            continue
        if is_target_determined_header(header_value):
            candidates.append((str(header_value), column_index))
        normalized = normalize_header_key(header_value)
        if normalized == "targetinweek":
            fallback_candidates.append((str(header_value), column_index))
    if not candidates:
        # No TARGET Determined column; try fallback target column that has values.
        if fallback_candidates:
            best_name = None
            best_count = -1
            for header_name, col_idx in fallback_candidates:
                cnt = count_non_empty_column_values(data_sheet, col_idx)
                if cnt > best_count:
                    best_count = cnt
                    best_name = header_name
            return best_name
        return None

    expected_tokens = {
        f"1 {reference_date.strftime('%b %y')}".lower(),
        f"1 {reference_date.strftime('%B %y')}".lower(),
        f"1 {reference_date.strftime('%b %Y')}".lower(),
        f"1 {reference_date.strftime('%B %Y')}".lower(),
    }
    for header_name, col_idx in candidates:
        header_lower = header_name.lower()
        if any(token in header_lower for token in expected_tokens) and count_non_empty_column_values(data_sheet, col_idx) > 0:
            return header_name

    # Fallback 1: choose TARGET Determined candidate with most non-empty values.
    best_header = candidates[0][0]
    best_count = -1
    for header_name, column_index in candidates:
        non_empty_count = count_non_empty_column_values(data_sheet, column_index)
        if non_empty_count > best_count:
            best_count = non_empty_count
            best_header = header_name

    if best_count > 0:
        return best_header

    # Fallback 2: if all TARGET Determined columns are empty, use Target In Week.
    if fallback_candidates:
        fallback_best_name = None
        fallback_best_count = -1
        for header_name, col_idx in fallback_candidates:
            cnt = count_non_empty_column_values(data_sheet, col_idx)
            if cnt > fallback_best_count:
                fallback_best_count = cnt
                fallback_best_name = header_name
        if fallback_best_name:
            return fallback_best_name

    return best_header


def apply_target_row_field_for_top_pivot(pivot_table, row_field_name: str | None) -> None:
    if not row_field_name:
        return
    xl_hidden = 0
    xl_row_field = 1

    target_field = None
    target_key = normalize_header_key(row_field_name)
    try:
        pivot_fields = pivot_table.PivotFields()
        for field_index in range(1, pivot_fields.Count + 1):
            field = pivot_fields(field_index)
            if normalize_header_key(field.Name) == target_key:
                target_field = field
                break
    except Exception:
        target_field = None
    if target_field is None:
        return

    try:
        pivot_fields = pivot_table.PivotFields()
        for field_index in range(1, pivot_fields.Count + 1):
            field = pivot_fields(field_index)
            if field.Orientation == xl_row_field:
                try:
                    field.Orientation = xl_hidden
                except Exception:
                    pass
    except Exception:
        pass

    try:
        target_field.Orientation = xl_row_field
        target_field.Position = 1
    except Exception:
        pass


def clear_row_field_item_filters(pivot_table) -> None:
    try:
        pivot_fields = pivot_table.PivotFields()
        for field_index in range(1, pivot_fields.Count + 1):
            field = pivot_fields(field_index)
            try:
                if field.Orientation != 1:
                    continue
            except Exception:
                continue
            try:
                field.ClearAllFilters()
            except Exception:
                pass
            try:
                for item in field.PivotItems():
                    try:
                        item.Visible = True
                    except Exception:
                        pass
            except Exception:
                pass
    except Exception:
        pass


def pivot_has_non_empty_grand_total(pivot_table) -> bool:
    try:
        table_range = pivot_table.TableRange1
        value = table_range.Cells(table_range.Rows.Count, table_range.Columns.Count).Value
        return value not in (None, "")
    except Exception:
        return False


def rebuild_top_pivot_layout(pivot_table, row_field_name: str | None) -> None:
    xl_row_field = 1
    xl_count = -4112
    try:
        pivot_table.ClearTable()
    except Exception:
        pass
    apply_target_row_field_for_top_pivot(pivot_table, row_field_name)
    clear_row_field_item_filters(pivot_table)

    quo_field = find_pivot_field(pivot_table, {normalize_header_key(QUO_HEADER)})
    if quo_field is not None:
        try:
            data_field = pivot_table.AddDataField(quo_field, "Count of quo", xl_count)
            data_field.NumberFormat = "0"
        except Exception:
            pass

    try:
        pivot_fields = pivot_table.PivotFields()
        for field_index in range(1, pivot_fields.Count + 1):
            field = pivot_fields(field_index)
            try:
                if field.Orientation == xl_row_field and normalize_header_key(field.Name) != normalize_header_key(
                    row_field_name or ""
                ):
                    field.Orientation = 0
            except Exception:
                pass
    except Exception:
        pass


def ensure_target_month_items_visible(pivot_table, reference_date: date) -> None:
    month_full = reference_date.strftime("%B").lower()
    month_short = reference_date.strftime("%b").lower()
    preferred_items = {
        "target not yet inputted",
        f"before {month_full}",
        f"target {month_full}",
        f"after {month_full}",
        f"before {month_short}",
        f"target {month_short}",
        f"after {month_short}",
    }
    try:
        pivot_fields = pivot_table.PivotFields()
        row_field = None
        for field_index in range(1, pivot_fields.Count + 1):
            field = pivot_fields(field_index)
            try:
                if field.Orientation == 1 and is_target_determined_header(field.Name):
                    row_field = field
                    break
            except Exception:
                continue
        if row_field is None:
            return
        for item in row_field.PivotItems():
            try:
                item_name = str(item.Name).strip().lower()
            except Exception:
                continue
            if item_name in preferred_items:
                try:
                    item.Visible = True
                except Exception:
                    pass
    except Exception:
        pass


def mirror_row_field_visibility(source_pivot, target_pivot) -> None:
    try:
        source_row_field = None
        target_row_field = None

        source_fields = source_pivot.PivotFields()
        for field_index in range(1, source_fields.Count + 1):
            field = source_fields(field_index)
            try:
                if field.Orientation == 1:
                    source_row_field = field
                    break
            except Exception:
                continue

        target_fields = target_pivot.PivotFields()
        for field_index in range(1, target_fields.Count + 1):
            field = target_fields(field_index)
            try:
                if field.Orientation == 1:
                    target_row_field = field
                    break
            except Exception:
                continue

        if source_row_field is None or target_row_field is None:
            return

        visible_names = set()
        for item in source_row_field.PivotItems():
            try:
                if item.Visible:
                    visible_names.add(str(item.Name))
            except Exception:
                pass

        if not visible_names:
            return

        try:
            target_row_field.ClearAllFilters()
        except Exception:
            pass

        for item in target_row_field.PivotItems():
            name = str(item.Name)
            if name in visible_names:
                try:
                    item.Visible = True
                except Exception:
                    pass
    except Exception:
        pass


def find_first_pivot_row_field(pivot_table):
    try:
        pivot_fields = pivot_table.PivotFields()
        for field_index in range(1, pivot_fields.Count + 1):
            field = pivot_fields(field_index)
            try:
                if field.Orientation == 1:
                    return field
            except Exception:
                continue
    except Exception:
        pass
    return None


def set_row_field_visible_items_only(row_field, allowed_names: set[str]) -> None:
    allowed_keys = {str(name).strip().lower() for name in allowed_names}
    try:
        pivot_items = row_field.PivotItems()
    except Exception:
        return

    item_names = []
    try:
        for item in pivot_items:
            try:
                item_names.append(str(item.Name).strip())
            except Exception:
                pass
    except Exception:
        return

    if not any(item_name.lower() in allowed_keys for item_name in item_names):
        return

    try:
        row_field.ClearAllFilters()
    except Exception:
        pass

    try:
        for item in row_field.PivotItems():
            try:
                item.Visible = True
            except Exception:
                pass
        for item in row_field.PivotItems():
            try:
                item_name = str(item.Name).strip()
            except Exception:
                continue
            if item_name.lower() not in allowed_keys:
                try:
                    item.Visible = False
                except Exception:
                    pass
    except Exception:
        pass


def find_on_progress_pivot(progress_sheet, preferred_name: str, fallback_position: int | None = None):
    try:
        return progress_sheet.PivotTables(preferred_name)
    except Exception:
        pass

    if fallback_position is None:
        return None

    try:
        candidates = []
        pivot_tables = progress_sheet.PivotTables()
        for pivot_index in range(1, pivot_tables.Count + 1):
            pt = pivot_tables(pivot_index)
            try:
                if str(pt.Name) == "PivotTable_StatusOrder_PreInstall_UAT":
                    continue
                candidates.append((pt.TableRange2.Row, pt.TableRange2.Column, str(pt.Name)))
            except Exception:
                continue
        candidates.sort()
        if len(candidates) > fallback_position:
            return progress_sheet.PivotTables(candidates[fallback_position][2])
    except Exception:
        pass
    return None


def enforce_on_progress_row_label_filters(progress_sheet, reference_date: date) -> None:
    month_name = reference_date.strftime("%B")
    target_not_yet = "Target Not Yet Inputted"
    before_month = f"Before {month_name}"
    target_month = f"Target {month_name}"
    after_month = f"After {month_name}"

    first_pivot = None
    second_pivot = None
    status_pivot = find_on_progress_pivot(progress_sheet, "PivotTable_StatusOrder_PreInstall_UAT")
    try:
        top_candidates = []
        pivot_tables = progress_sheet.PivotTables()
        for pivot_index in range(1, pivot_tables.Count + 1):
            pt = pivot_tables(pivot_index)
            try:
                if str(pt.Name) == "PivotTable_StatusOrder_PreInstall_UAT":
                    continue
                row = pt.TableRange2.Row
                if row > 10:
                    continue
                top_candidates.append((pt.TableRange2.Column, row, str(pt.Name)))
            except Exception:
                continue
        top_candidates.sort()
        if top_candidates:
            first_pivot = progress_sheet.PivotTables(top_candidates[0][2])
        if len(top_candidates) > 1:
            second_pivot = progress_sheet.PivotTables(top_candidates[1][2])
    except Exception:
        first_pivot = find_on_progress_pivot(progress_sheet, "PivotTable2", 0)
        second_pivot = find_on_progress_pivot(progress_sheet, "PivotTable9", 1)

    status_row_field_name = None
    first_row_field = find_first_pivot_row_field(first_pivot) if first_pivot is not None else None
    second_row_field = find_first_pivot_row_field(second_pivot) if second_pivot is not None else None

    if first_row_field is not None:
        try:
            status_row_field_name = str(first_row_field.Name)
        except Exception:
            status_row_field_name = None
    if status_row_field_name is None and second_row_field is not None:
        try:
            status_row_field_name = str(second_row_field.Name)
        except Exception:
            status_row_field_name = None

    if first_pivot is not None:
        if first_row_field is None and status_row_field_name:
            rebuild_top_pivot_layout(first_pivot, status_row_field_name)
            apply_common_pivot_filters(first_pivot)
            apply_phase_filter_excluding_cancel_so_complete(first_pivot)
            try:
                first_pivot.RefreshTable()
            except Exception:
                pass
            first_row_field = find_first_pivot_row_field(first_pivot)
        if first_row_field is not None:
            set_row_field_visible_items_only(first_row_field, {target_month, after_month})
            try:
                first_pivot.RefreshTable()
            except Exception:
                pass

    if second_pivot is not None:
        if second_row_field is not None:
            if status_row_field_name is None:
                try:
                    status_row_field_name = str(second_row_field.Name)
                except Exception:
                    status_row_field_name = None
            set_row_field_visible_items_only(second_row_field, {target_not_yet, before_month})
            try:
                second_pivot.RefreshTable()
            except Exception:
                pass

    if status_pivot is not None:
        if status_row_field_name:
            apply_target_row_field_for_top_pivot(status_pivot, status_row_field_name)
        status_row_field = find_first_pivot_row_field(status_pivot)
        if status_row_field is not None:
            set_row_field_visible_items_only(
                status_row_field,
                {target_not_yet, before_month, after_month, target_month},
            )
            try:
                status_pivot.RefreshTable()
            except Exception:
                pass


def restore_on_progress_filter_caption_layout(progress_sheet) -> None:
    """Restore the visible page-field caption order used by the ongoing template."""
    caption_cells = {
        "A2": YEAR_FAB_UPLOAD_HEADER,
        "A3": PROCESS_ADJUSTMENT_HEADER,
        "A4": PHASE_HEADER_PREFIX,
        "D1": YEAR_FAB_UPLOAD_HEADER,
        "D2": PROCESS_ADJUSTMENT_HEADER,
        "D3": PHASE_HEADER_PREFIX,
        "A15": YEAR_FAB_UPLOAD_HEADER,
        "A16": PROCESS_ADJUSTMENT_HEADER,
        "A17": PHASE_HEADER_PREFIX,
    }
    for cell_address, value in caption_cells.items():
        try:
            progress_sheet.Range(cell_address).Value = value
        except Exception:
            pass


def find_contiguous_percentage_block_last_row(pivot_sheet, header_row: int, header_col: int) -> int:
    used_last_row = pivot_sheet.UsedRange.Rows.Count
    for row_index in range(header_row + 1, used_last_row + 1):
        has_block_value = False
        for column_index in range(1, header_col + 1):
            try:
                if pivot_sheet.Cells(row_index, column_index).Value not in (None, ""):
                    has_block_value = True
                    break
            except Exception:
                continue
        if not has_block_value:
            return row_index - 1
    return used_last_row


def normalize_percentage_block_formatting(pivot_sheet) -> None:
    percentage_header_key = normalize_header_key("Percentage of Order On going from Pre-Installation to UAT")
    used_rows = pivot_sheet.UsedRange.Rows.Count
    used_cols = pivot_sheet.UsedRange.Columns.Count
    header_cells: list[tuple[int, int]] = []

    for row_index in range(1, used_rows + 1):
        for column_index in range(1, used_cols + 1):
            try:
                value = pivot_sheet.Cells(row_index, column_index).Value
            except Exception:
                continue
            if normalize_header_key(value) == percentage_header_key:
                header_cells.append((row_index, column_index))

    xl_edge_left = 7
    xl_edge_top = 8
    xl_edge_bottom = 9
    xl_edge_right = 10
    xl_inside_horizontal = 12
    xl_continuous = 1
    xl_none = -4142
    xl_medium = -4138
    xl_red = 255

    for index, (header_row, header_col) in enumerate(header_cells):
        last_row = find_contiguous_percentage_block_last_row(pivot_sheet, header_row, header_col)
        next_header_row = header_cells[index + 1][0] if index + 1 < len(header_cells) else used_rows + 1
        cleanup_last_row = min(next_header_row - 1, max(last_row + 8, header_row))

        try:
            cleanup_range = pivot_sheet.Range(
                pivot_sheet.Cells(header_row, header_col),
                pivot_sheet.Cells(cleanup_last_row, header_col),
            )
            for border_id in (xl_edge_left, xl_edge_top, xl_edge_bottom, xl_edge_right, xl_inside_horizontal):
                cleanup_range.Borders(border_id).LineStyle = xl_none
            data_cleanup_start = last_row + 1
            if data_cleanup_start <= cleanup_last_row:
                pivot_sheet.Range(
                    pivot_sheet.Cells(data_cleanup_start, header_col),
                    pivot_sheet.Cells(cleanup_last_row, header_col),
                ).FormatConditions.Delete()
        except Exception:
            pass

        if last_row <= header_row:
            continue

        try:
            block_range = pivot_sheet.Range(
                pivot_sheet.Cells(header_row, header_col),
                pivot_sheet.Cells(last_row, header_col),
            )
            for border_id in (xl_edge_left, xl_edge_right):
                border = block_range.Borders(border_id)
                border.LineStyle = xl_continuous
                border.Color = xl_red
                border.Weight = xl_medium

            header_cell = pivot_sheet.Cells(header_row, header_col)
            header_cell.Font.Bold = True
            header_cell.HorizontalAlignment = -4108
            header_cell.VerticalAlignment = -4108
            header_cell.WrapText = True
            for border_id in (xl_edge_left, xl_edge_top, xl_edge_bottom, xl_edge_right):
                border = header_cell.Borders(border_id)
                border.LineStyle = xl_continuous
                border.Color = xl_red
                border.Weight = xl_medium

            bottom_cell = pivot_sheet.Cells(last_row, header_col)
            bottom_cell.Font.Bold = True
            bottom_cell.HorizontalAlignment = -4108
            bottom_cell.VerticalAlignment = -4108
            for border_id in (xl_edge_left, xl_edge_top, xl_edge_bottom, xl_edge_right):
                border = bottom_cell.Borders(border_id)
                border.LineStyle = xl_continuous
                border.Color = xl_red
                border.Weight = xl_medium

            data_range = pivot_sheet.Range(
                pivot_sheet.Cells(header_row + 1, header_col),
                pivot_sheet.Cells(last_row, header_col),
            )
            data_range.Font.Bold = True
            data_range.HorizontalAlignment = -4108
            data_range.VerticalAlignment = -4108
            data_range.FormatConditions.Delete()
            icon_condition = data_range.FormatConditions.AddIconSetCondition()
            icon_condition.IconSet = pivot_sheet.Parent.IconSets(4)
            icon_condition.ReverseOrder = False
            icon_condition.ShowIconOnly = False
            icon_condition.IconCriteria(2).Type = 0
            icon_condition.IconCriteria(2).Operator = 7
            icon_condition.IconCriteria(2).Value = 0.98
            icon_condition.IconCriteria(3).Type = 0
            icon_condition.IconCriteria(3).Operator = 7
            icon_condition.IconCriteria(3).Value = 0.99
        except Exception:
            pass


def clear_ongoing_percentage_columns(pivot_sheet) -> None:
    percentage_header_key = normalize_header_key("Percentage of Order On going from Pre-Installation to UAT")
    try:
        used_range = pivot_sheet.UsedRange
        first_row = used_range.Row
        first_col = used_range.Column
        row_count = used_range.Rows.Count
        col_count = used_range.Columns.Count
        raw_values = used_range.Value
    except Exception:
        return

    if not isinstance(raw_values, tuple):
        values = [[raw_values]]
    elif raw_values and not isinstance(raw_values[0], tuple):
        values = [list(raw_values)]
    else:
        values = [list(row) for row in raw_values]

    for row_offset, row_values in enumerate(values):
        row_index = first_row + row_offset
        for col_offset, value in enumerate(row_values):
            if normalize_header_key(value) != percentage_header_key:
                continue

            column_index = first_col + col_offset
            cleanup_last_row = find_contiguous_percentage_block_last_row(pivot_sheet, row_index, column_index)
            try:
                pivot_sheet.Range(
                    pivot_sheet.Cells(row_index, column_index),
                    pivot_sheet.Cells(cleanup_last_row, column_index),
                ).Clear()
            except Exception:
                pass


def repair_spilled_pivot_tables(pivot_sheet) -> None:
    try:
        pivot_tables = pivot_sheet.PivotTables()
    except Exception:
        return

    for pivot_index in range(1, pivot_tables.Count + 1):
        try:
            pivot_table = pivot_tables(pivot_index)
            table_range = pivot_table.TableRange2
            top_row = table_range.Row
            left_col = table_range.Column
            top_left_cell = pivot_sheet.Cells(top_row, left_col)
            is_spill_error = str(top_left_cell.Text).strip().upper() == "#SPILL!" or top_left_cell.Value == -2146826243
            if not is_spill_error:
                continue

            last_row = top_row + 1
            for row_index in range(top_row + 1, top_row + 180):
                try:
                    left_label = str(pivot_sheet.Cells(row_index, 1).Text).strip().lower()
                    if left_label == "grand total":
                        last_row = row_index
                        break
                    if row_index > top_row + 1 and not left_label:
                        last_row = row_index - 1
                        break
                except Exception:
                    continue

            blocker_col = left_col + 5
            pivot_sheet.Range(
                pivot_sheet.Cells(top_row + 1, blocker_col),
                pivot_sheet.Cells(last_row, blocker_col),
            ).Clear()
            pivot_table.RefreshTable()
        except Exception:
            continue


def ensure_ongoing_percentage_columns(pivot_sheet) -> None:
    percentage_header = "Percentage of Order On going from Pre-Installation to UAT"
    percentage_header_key = normalize_header_key(percentage_header)
    div_dept_key = normalize_header_key("Div./Dept.")
    count_quo_key = normalize_header_key("Count of quo")
    phase_headers_for_numerator = {
        normalize_header_key("04-Pre Installation"),
        normalize_header_key("05-Customer Preparation"),
        normalize_header_key("06-Installation"),
        normalize_header_key("07-UAT On Hold"),
    }
    try:
        used_range = pivot_sheet.UsedRange
        first_row = used_range.Row
        first_col = used_range.Column
        used_rows = used_range.Rows.Count
        used_cols = used_range.Columns.Count
        raw_values = used_range.Value
    except Exception:
        return

    if used_rows <= 0 or used_cols <= 0:
        return

    if not isinstance(raw_values, tuple):
        values = [[raw_values]]
    elif raw_values and not isinstance(raw_values[0], tuple):
        values = [list(raw_values)]
    else:
        values = [list(row) for row in raw_values]

    def _cell_value(row_index: int, column_index: int):
        row_offset = row_index - first_row
        column_offset = column_index - first_col
        if row_offset < 0 or column_offset < 0:
            return None
        try:
            return values[row_offset][column_offset]
        except Exception:
            return None

    def _column_letter(column_number: int) -> str:
        result = ""
        while column_number:
            column_number, remainder = divmod(column_number - 1, 26)
            result = chr(65 + remainder) + result
        return result

    for header_row in range(first_row, first_row + used_rows):
        right_label_col = None
        phase_columns: list[int] = []
        segments: list[tuple[int, list[int]]] = []

        for column_index in range(first_col, first_col + used_cols + 1):
            value = _cell_value(header_row, column_index)
            normalized_value = normalize_header_key(value)
            if normalized_value == div_dept_key:
                if right_label_col is not None and phase_columns:
                    segments.append((right_label_col, phase_columns))
                right_label_col = column_index
                phase_columns = []
                continue

            if right_label_col is None or column_index <= right_label_col:
                continue

            if normalized_value == percentage_header_key:
                if phase_columns:
                    segments.append((right_label_col, phase_columns))
                break

            if normalized_value in phase_headers_for_numerator:
                phase_columns.append(column_index)

        if right_label_col is not None and phase_columns:
            segments.append((right_label_col, phase_columns))

        if not segments:
            continue

        for segment_label_col, segment_phase_columns in segments:
            denominator_col = None
            for column_index in range(segment_label_col - 1, 0, -1):
                if normalize_header_key(_cell_value(header_row, column_index)) == count_quo_key:
                    denominator_col = column_index
                    break

            if denominator_col is None:
                continue

            percentage_col = max(segment_phase_columns) + 1
            try:
                existing_header = _cell_value(header_row, percentage_col)
                if existing_header not in (None, "") and normalize_header_key(existing_header) != percentage_header_key:
                    continue

                pivot_sheet.Cells(header_row, percentage_col).Value = percentage_header
                pivot_sheet.Cells(header_row, percentage_col).WrapText = True
                pivot_sheet.Cells(header_row, percentage_col).HorizontalAlignment = -4108
                pivot_sheet.Cells(header_row, percentage_col).VerticalAlignment = -4108
                pivot_sheet.Cells(header_row, percentage_col).Font.Bold = True

                last_row = find_contiguous_percentage_block_last_row(pivot_sheet, header_row, percentage_col)
                numerator_range = ":".join(
                    (
                        f"{_column_letter(min(segment_phase_columns))}{{row}}",
                        f"{_column_letter(max(segment_phase_columns))}{{row}}",
                    )
                )
                denominator_letter = _column_letter(denominator_col)
                for row_index in range(header_row + 1, last_row + 1):
                    denominator = _cell_value(row_index, denominator_col)
                    if denominator in (None, ""):
                        pivot_sheet.Cells(row_index, percentage_col).ClearContents()
                        continue

                    pivot_sheet.Cells(row_index, percentage_col).Formula = (
                        f"=(SUM({numerator_range.format(row=row_index)}))/"
                        f"{denominator_letter}{row_index}"
                    )
                    pivot_sheet.Cells(row_index, percentage_col).NumberFormat = "0.00%"

                pivot_sheet.Columns(percentage_col).ColumnWidth = max(
                    pivot_sheet.Columns(percentage_col).ColumnWidth,
                    24,
                )
            except Exception:
                pass


def restore_ongoing_pivot_sheet_column_widths(pivot_sheet, reference_date: date | None = None) -> None:
    """Keep the ongoing PIVOT report width model stable after Excel refreshes pivots."""
    try:
        pivot_tables = pivot_sheet.PivotTables()
        for pivot_index in range(1, pivot_tables.Count + 1):
            apply_range_aging_column_order(pivot_tables(pivot_index))
    except Exception:
        pass

    april_widths = {
        5: 69.55,
        6: 25.82,
        7: 25.36,
        8: 23.09,
        9: 22.0,
        10: 26.55,
        11: 30.36,
        12: 16.27,
        13: 29.27,
    }
    default_widths = {
        5: 69.55,
        6: 25.82,
        7: 25.36,
        8: 31.73,
        11: 32.45,
        12: 24.91,
        13: 18.36,
    }
    column_widths = april_widths if reference_date is not None and reference_date.month == 4 else default_widths
    for column_index, width in column_widths.items():
        try:
            pivot_sheet.Columns(column_index).ColumnWidth = width
        except Exception:
            pass

    if reference_date is not None and reference_date.month == 4:
        try:
            pivot_sheet.Cells(1, 12).Interior.Color = 16777215
        except Exception:
            pass
        try:
            percentage_header = "Percentage of Order On going from Pre-Installation to UAT"
            try:
                pivot_tables = pivot_sheet.PivotTables()
                for pivot_index in range(1, pivot_tables.Count + 1):
                    pivot_table = pivot_tables(pivot_index)
                    fields = pivot_table.PivotFields()
                    for field_index in range(1, fields.Count + 1):
                        field = fields(field_index)
                        if field.Orientation == 2 and normalize_header_key(field.Name) == normalize_header_key(PHASE_HEADER_PREFIX):
                            try:
                                field.PivotItems(percentage_header).Caption = "06-Installation"
                            except Exception:
                                pass
            except Exception:
                pass
            pivot_sheet.Cells(9, 10).Value = "07-UAT On Hold"
            pivot_sheet.Cells(9, 11).Value = percentage_header
            pivot_sheet.Cells(9, 12).ClearContents()
            for row_index in range(10, 22):
                denominator = pivot_sheet.Cells(row_index, 3).Value
                if isinstance(denominator, (int, float)) and denominator:
                    pivot_sheet.Cells(row_index, 11).Formula = f"=(SUM(H{row_index}:J{row_index}))/C{row_index}"
                    pivot_sheet.Cells(row_index, 11).NumberFormat = "0.00%"
                else:
                    pivot_sheet.Cells(row_index, 11).ClearContents()
                pivot_sheet.Cells(row_index, 12).ClearContents()
        except Exception:
            pass

    try:
        used_range = pivot_sheet.UsedRange
        used_range.Replace(What=-2146826281, Replacement="", LookAt=1)
    except Exception:
        pass

    ensure_ongoing_percentage_columns(pivot_sheet)
    repair_spilled_pivot_tables(pivot_sheet)
    ensure_ongoing_percentage_columns(pivot_sheet)
    normalize_percentage_block_formatting(pivot_sheet)


def refresh_reporting_sheets_fast_in_workbook(workbook, reference_date: date) -> bool:
    """Refresh existing template pivots once; return False when rebuild fallback is needed."""
    data_sheet = None
    progress_sheet = None
    pivot_sheet = None
    for sheet in workbook.Sheets:
        if sheet.Name == SHEET_NAME:
            data_sheet = sheet
        elif sheet.Name == PIVOT_SHEET_NAME:
            pivot_sheet = sheet
        elif str(sheet.Name).startswith(ON_PROGRESS_SHEET_PREFIX):
            progress_sheet = sheet

    if data_sheet is None or progress_sheet is None or pivot_sheet is None:
        return False

    try:
        phase_start = time.perf_counter()
        refreshed_cache_indexes: set[int] = set()
        for sheet in (progress_sheet, pivot_sheet):
            pivot_tables = sheet.PivotTables()
            for pivot_index in range(1, pivot_tables.Count + 1):
                pivot_table = pivot_tables(pivot_index)
                cache_index = int(pivot_table.CacheIndex)
                if cache_index in refreshed_cache_indexes:
                    continue
                pivot_table.PivotCache().Refresh()
                refreshed_cache_indexes.add(cache_index)
        profile_log("refresh pivot caches only", phase_start)
        phase_start = time.perf_counter()
        clear_ongoing_percentage_columns(pivot_sheet)
        for sheet in (progress_sheet, pivot_sheet):
            pivot_tables = sheet.PivotTables()
            for pivot_index in range(1, pivot_tables.Count + 1):
                pivot_tables(pivot_index).RefreshTable()
        repair_spilled_pivot_tables(pivot_sheet)
        profile_log("refresh pivot tables", phase_start)
    except Exception:
        try:
            phase_start = time.perf_counter()
            workbook.RefreshAll()
            profile_log("fallback workbook RefreshAll", phase_start)
        except Exception:
            return False

    target_progress_name = expected_on_progress_sheet_name(reference_date)
    try:
        if progress_sheet.Name != target_progress_name:
            progress_sheet.Name = target_progress_name
        progress_sheet.Tab.Color = EXCEL_ON_PROGRESS_TAB_COLOR
    except Exception:
        pass

    try:
        phase_start = time.perf_counter()
        target_row_field_name = resolve_top_pivot_row_field_for_reference_date(data_sheet, reference_date)
        pivot_tables = progress_sheet.PivotTables()
        top_candidates = []
        for pivot_index in range(1, pivot_tables.Count + 1):
            pivot_table = pivot_tables(pivot_index)
            try:
                if str(pivot_table.Name) == "PivotTable_StatusOrder_PreInstall_UAT":
                    continue
                row = pivot_table.TableRange2.Row
                if row > 10:
                    continue
                top_candidates.append((row, pivot_table.TableRange2.Column, str(pivot_table.Name)))
            except Exception:
                continue
        top_candidates.sort()
        for _, _, pivot_name in top_candidates[:2]:
            pivot_table = progress_sheet.PivotTables(pivot_name)
            apply_target_row_field_for_top_pivot(pivot_table, target_row_field_name)
            clear_row_field_item_filters(pivot_table)
            apply_common_pivot_filters(pivot_table)
            apply_phase_filter_excluding_cancel_so_complete(pivot_table)
            ensure_target_month_items_visible(pivot_table, reference_date)

        if len(top_candidates) < 2:
            return False
        first_top = progress_sheet.PivotTables(top_candidates[0][2])
        second_top = progress_sheet.PivotTables(top_candidates[1][2])
        if not pivot_has_non_empty_grand_total(first_top) or not pivot_has_non_empty_grand_total(second_top):
            return False
        profile_log("repair top progress pivots", phase_start)
    except Exception:
        return False

    try:
        phase_start = time.perf_counter()
        restore_on_progress_filter_caption_layout(progress_sheet)
        enforce_on_progress_row_label_filters(progress_sheet, reference_date)
        profile_log("restore progress sheet layout", phase_start)
        phase_start = time.perf_counter()
        restore_ongoing_pivot_sheet_column_widths(pivot_sheet, reference_date)
        profile_log("restore ongoing pivot sheet", phase_start)
    except Exception:
        return False

    return True


def excel_serial_from_datetime(value: datetime) -> float:
    return (value - datetime(1899, 12, 30)).total_seconds() / 86400


def create_named_range_pivot_cache(workbook, data_sheet, source_name: str):
    try:
        try:
            list_objects = data_sheet.ListObjects
            if list_objects.Count > 0:
                source_range = list_objects(1).Range
            else:
                source_range = data_sheet.UsedRange
        except Exception:
            source_range = data_sheet.UsedRange

        start_row = source_range.Row
        start_col = source_range.Column
        end_row = start_row + source_range.Rows.Count - 1
        end_col = start_col + source_range.Columns.Count - 1
        source_address = (
            f"${get_column_letter(start_col)}${start_row}:"
            f"${get_column_letter(end_col)}${end_row}"
        )
        try:
            workbook.Names(source_name).Delete()
        except Exception:
            pass
        workbook.Names.Add(Name=source_name, RefersTo=f"='{SHEET_NAME}'!{source_address}")
        return workbook.PivotCaches().Create(SourceType=1, SourceData=source_name)
    except Exception:
        return None


def update_on_progress_sheet_pivots_via_com(
    output_path: Path,
    reference_date: date,
    xl_app=None,
    workbook=None,
    save_workbook: bool = True,
) -> None:
    try:
        import win32com.client  # type: ignore
    except ImportError:
        raise RuntimeError("pywin32 is required to update pivot tables on the ongoing output workbook.")

    xl = xl_app
    wb = workbook
    owns_excel = xl is None
    owns_workbook = wb is None
    try:
        if xl is None:
            xl = win32com.client.DispatchEx("Excel.Application")
            xl.Visible = False
            xl.DisplayAlerts = False

        if wb is None:
            wb = xl.Workbooks.Open(str(output_path.resolve()))
        data_sheet = None
        progress_sheet = None
        for sheet in wb.Sheets:
            if sheet.Name == SHEET_NAME:
                data_sheet = sheet
            if str(sheet.Name).startswith(ON_PROGRESS_SHEET_PREFIX):
                progress_sheet = sheet

        if data_sheet is None or progress_sheet is None:
            if save_workbook:
                wb.Save()
            return

        pivot_sheet = None
        for sheet in wb.Sheets:
            if sheet.Name == PIVOT_SHEET_NAME:
                pivot_sheet = sheet
                break

        target_progress_name = expected_on_progress_sheet_name(reference_date)
        if progress_sheet.Name != target_progress_name:
            progress_sheet.Name = target_progress_name

        try:
            progress_sheet.Tab.Color = EXCEL_ON_PROGRESS_TAB_COLOR
        except Exception:
            pass

        pivot_tables = progress_sheet.PivotTables()
        top_candidates = []
        for pivot_index in range(1, pivot_tables.Count + 1):
            pt = pivot_tables(pivot_index)
            try:
                top_candidates.append((pt.TableRange2.Row, pt.TableRange2.Column, str(pt.Name)))
            except Exception:
                pass
        top_candidates.sort()
        top_pivot_names = [name for _, _, name in top_candidates[:2]]
        top_pivot_count = len(top_pivot_names)
        target_row_field_name = resolve_top_pivot_row_field_for_reference_date(data_sheet, reference_date)
        for top_name in top_pivot_names:
            pt = progress_sheet.PivotTables(top_name)
            try:
                pt.ManualUpdate = True
            except Exception:
                pass
            apply_target_row_field_for_top_pivot(pt, target_row_field_name)
            clear_row_field_item_filters(pt)
            apply_common_pivot_filters(pt)
            apply_phase_filter_excluding_cancel_so_complete(pt)
            try:
                pt.ManualUpdate = False
                pt.RefreshTable()
                ensure_target_month_items_visible(pt, reference_date)
            except Exception:
                pass
            if find_first_pivot_row_field(pt) is None or not pivot_has_non_empty_grand_total(pt):
                try:
                    pt.ManualUpdate = True
                except Exception:
                    pass
                rebuild_top_pivot_layout(pt, target_row_field_name)
                apply_common_pivot_filters(pt)
                apply_phase_filter_excluding_cancel_so_complete(pt)
                try:
                    pt.ManualUpdate = False
                    pt.RefreshTable()
                    ensure_target_month_items_visible(pt, reference_date)
                except Exception:
                    pass

        # If second top pivot still empty while first has data, recreate second pivot
        # from the same cache in-place (template may carry stale month-specific item filters).
        if top_pivot_count >= 2:
            try:
                first_top = progress_sheet.PivotTables(top_pivot_names[0])
                second_top = progress_sheet.PivotTables(top_pivot_names[1])
                if pivot_has_non_empty_grand_total(first_top) and not pivot_has_non_empty_grand_total(second_top):
                    second_row = second_top.TableRange2.Row
                    second_col = second_top.TableRange2.Column
                    second_name = str(second_top.Name)
                    second_top.TableRange2.Clear()
                    recreated_second = first_top.PivotCache().CreatePivotTable(
                        TableDestination=progress_sheet.Cells(second_row, second_col),
                        TableName=second_name,
                    )
                    try:
                        recreated_second.ManualUpdate = True
                    except Exception:
                        pass
                    rebuild_top_pivot_layout(recreated_second, target_row_field_name)
                    apply_common_pivot_filters(recreated_second)
                    apply_phase_filter_excluding_cancel_so_complete(recreated_second)
                    try:
                        recreated_second.ManualUpdate = False
                        recreated_second.RefreshTable()
                        ensure_target_month_items_visible(recreated_second, reference_date)
                    except Exception:
                        pass
            except Exception:
                pass

        max_bottom = 1
        for top_name in top_pivot_names:
            pt = progress_sheet.PivotTables(top_name)
            try:
                bottom = pt.TableRange2.Row + pt.TableRange2.Rows.Count - 1
                if bottom > max_bottom:
                    max_bottom = bottom
            except Exception:
                continue

        destination_row = max_bottom + 6
        third_pivot_name = "PivotTable_StatusOrder_PreInstall_UAT"
        # Pre-clean extra pivots (like PivotTable3) before creating status pivot
        # to avoid overlap with the target destination area.
        try:
            existing_pivots = progress_sheet.PivotTables()
            existing_candidates = []
            for pivot_index in range(1, existing_pivots.Count + 1):
                pt = existing_pivots(pivot_index)
                try:
                    row = pt.TableRange2.Row
                    col = pt.TableRange2.Column
                    existing_candidates.append((row, col, str(pt.Name)))
                except Exception:
                    continue
            existing_candidates.sort()
            keep_top_names = {name for _, _, name in existing_candidates[:2]}
            keep_top_names.add(third_pivot_name)
            for pivot_index in range(progress_sheet.PivotTables().Count, 0, -1):
                pt = progress_sheet.PivotTables(pivot_index)
                name = str(pt.Name)
                if name in keep_top_names:
                    continue
                try:
                    pt.TableRange2.Clear()
                except Exception:
                    pass
        except Exception:
            pass

        # Force status pivot location to row 15 (TableRange2) on every run.
        # We create it with an offset destination because page fields expand upward.
        status_target_row = 15
        status_create_row = 19
        third_pivot = None
        try:
            existing_third = progress_sheet.PivotTables(third_pivot_name)
            existing_third.TableRange2.Clear()
        except Exception:
            pass

        if pivot_tables.Count:
            shared_cache = pivot_tables(1).PivotCache()
        else:
            shared_cache = wb.PivotCaches().Create(SourceType=1, SourceData=data_sheet.UsedRange)
        third_pivot = shared_cache.CreatePivotTable(
            TableDestination=progress_sheet.Range(f"A{status_create_row}"),
            TableName=third_pivot_name,
        )

        xl_row_field = 1
        xl_count = -4112
        try:
            third_pivot.ManualUpdate = True
        except Exception:
            pass
        try:
            third_pivot.ClearTable()
        except Exception:
            pass

        status_order_field = find_pivot_field(third_pivot, {normalize_header_key(STATUS_ORDER_HEADER)})
        if status_order_field is not None:
            try:
                status_order_field.Orientation = xl_row_field
                status_order_field.Position = 1
            except Exception:
                pass

        quo_field = find_pivot_field(third_pivot, {normalize_header_key(QUO_HEADER)})
        if quo_field is not None:
            try:
                data_field = third_pivot.AddDataField(quo_field, "Count of quo", xl_count)
                data_field.NumberFormat = "0"
            except Exception:
                pass

        try:
            third_pivot.TableStyle2 = "PivotStyleMedium2"
        except Exception:
            pass

        apply_common_pivot_filters(third_pivot)
        apply_phase_filter_04_to_07(third_pivot)

        try:
            third_pivot.ManualUpdate = False
            third_pivot.RefreshTable()
            apply_common_pivot_filters(third_pivot)
            apply_phase_filter_04_to_07(third_pivot)
        except Exception:
            pass

        # If position shifted, adjust once and recreate.
        try:
            current_row = third_pivot.TableRange2.Row
            if current_row != status_target_row:
                adjusted_create_row = status_create_row + (status_target_row - current_row)
                third_pivot.TableRange2.Clear()
                third_pivot = shared_cache.CreatePivotTable(
                    TableDestination=progress_sheet.Range(f"A{adjusted_create_row}"),
                    TableName=third_pivot_name,
                )
                try:
                    third_pivot.ManualUpdate = True
                except Exception:
                    pass
                try:
                    third_pivot.ClearTable()
                except Exception:
                    pass
                status_order_field = find_pivot_field(third_pivot, {normalize_header_key(STATUS_ORDER_HEADER)})
                if status_order_field is not None:
                    try:
                        status_order_field.Orientation = xl_row_field
                        status_order_field.Position = 1
                    except Exception:
                        pass
                quo_field = find_pivot_field(third_pivot, {normalize_header_key(QUO_HEADER)})
                if quo_field is not None:
                    try:
                        data_field = third_pivot.AddDataField(quo_field, "Count of quo", xl_count)
                        data_field.NumberFormat = "0"
                    except Exception:
                        pass
                try:
                    third_pivot.TableStyle2 = "PivotStyleMedium2"
                except Exception:
                    pass
                apply_common_pivot_filters(third_pivot)
                apply_phase_filter_04_to_07(third_pivot)
                try:
                    third_pivot.ManualUpdate = False
                    third_pivot.RefreshTable()
                except Exception:
                    pass
        except Exception:
            pass

        try:
            for row_index in range(1, 40):
                cell = progress_sheet.Range(f"A{row_index}")
                if str(cell.Value) == "Status Order Start From Pre-installation to UAT":
                    cell.ClearContents()
        except Exception:
            pass

        # Keep only the expected three pivots on ON PROGRESS sheet:
        # two upper pivots + status pivot table.
        try:
            latest_pivots = progress_sheet.PivotTables()
            top_candidates = []
            for pivot_index in range(1, latest_pivots.Count + 1):
                pt = latest_pivots(pivot_index)
                try:
                    row = pt.TableRange2.Row
                    col = pt.TableRange2.Column
                    top_candidates.append((row, col, str(pt.Name)))
                except Exception:
                    continue
            top_candidates.sort()
            keep_names = {name for _, _, name in top_candidates[:2]}
            keep_names.add(third_pivot_name)

            for pivot_index in range(progress_sheet.PivotTables().Count, 0, -1):
                pt = progress_sheet.PivotTables(pivot_index)
                name = str(pt.Name)
                if name in keep_names:
                    continue
                try:
                    # Extra pivot appears in middle area (row 13-24); clear it.
                    row = pt.TableRange2.Row
                    if 13 <= row <= 24:
                        pt.TableRange2.Clear()
                    else:
                        pt.TableRange2.Clear()
                except Exception:
                    pass
        except Exception:
            pass

        # Put label after cleanup so it persists in final output.
        try:
            progress_sheet.Range("A13").Value = "Status Order Start From Pre-installation to UAT"
        except Exception:
            pass

        # Final enforce for top-right pivot (commonly PivotTable9) on month rollover.
        try:
            if progress_sheet.PivotTables().Count >= 2:
                first_top = progress_sheet.PivotTables(1)
                second_top = progress_sheet.PivotTables(2)
                ensure_target_month_items_visible(second_top, reference_date)
                if pivot_has_non_empty_grand_total(first_top) and not pivot_has_non_empty_grand_total(second_top):
                    mirror_row_field_visibility(first_top, second_top)
                    second_top.RefreshTable()
                if pivot_has_non_empty_grand_total(first_top) and not pivot_has_non_empty_grand_total(second_top):
                    second_row = second_top.TableRange2.Row
                    second_col = second_top.TableRange2.Column
                    second_name = str(second_top.Name)
                    second_top.TableRange2.Clear()
                    recreated_second = first_top.PivotCache().CreatePivotTable(
                        TableDestination=progress_sheet.Cells(second_row, second_col),
                        TableName=second_name,
                    )
                    try:
                        recreated_second.ManualUpdate = True
                    except Exception:
                        pass
                    rebuild_top_pivot_layout(recreated_second, target_row_field_name)
                    clear_row_field_item_filters(recreated_second)
                    ensure_target_month_items_visible(recreated_second, reference_date)
                    apply_common_pivot_filters(recreated_second)
                    apply_phase_filter_excluding_cancel_so_complete(recreated_second)
                    try:
                        recreated_second.ManualUpdate = False
                        recreated_second.RefreshTable()
                    except Exception:
                        pass
        except Exception:
            pass

        enforce_on_progress_row_label_filters(progress_sheet, reference_date)
        restore_on_progress_filter_caption_layout(progress_sheet)
        if pivot_sheet is not None:
            restore_ongoing_pivot_sheet_column_widths(pivot_sheet, reference_date)

        if save_workbook:
            wb.Save()
    finally:
        if owns_workbook and wb is not None:
            try:
                wb.Close(SaveChanges=False)
            except Exception:
                pass
        if owns_excel and xl is not None:
            try:
                xl.Quit()
            except Exception:
                pass


def clone_and_refresh_reporting_sheets_from_source_via_com(output_path: Path, source_workbook_path: Path) -> None:
    try:
        import win32com.client  # type: ignore
    except ImportError:
        raise RuntimeError("pywin32 is required to clone and refresh PIVOT sheet from source workbook.")

    xl = None
    target_wb = None
    source_wb = None
    try:
        xl = win32com.client.DispatchEx("Excel.Application")
        xl.Visible = False
        xl.DisplayAlerts = False

        target_wb = xl.Workbooks.Open(str(output_path.resolve()))
        source_wb = xl.Workbooks.Open(str(source_workbook_path.resolve()))

        data_sheet = None
        for sheet in target_wb.Sheets:
            if sheet.Name == SHEET_NAME:
                data_sheet = sheet
                break
        if data_sheet is None:
            raise RuntimeError(f"Could not find '{SHEET_NAME}' sheet in target workbook.")

        source_pivot_sheet = None
        source_data_sheet = None
        for sheet in source_wb.Sheets:
            if sheet.Name == PIVOT_SHEET_NAME:
                source_pivot_sheet = sheet
            if sheet.Name == SHEET_NAME:
                source_data_sheet = sheet
        if source_pivot_sheet is None:
            raise RuntimeError(f"Could not find '{PIVOT_SHEET_NAME}' sheet in source workbook.")

        # Sync today's ALL ORDER data into the source template (in-memory only),
        # refresh its pivot cache, then clone the refreshed PIVOT sheet.
        if source_data_sheet is not None:
            xl_up = -4162
            xl_left = -4159
            last_row = data_sheet.Cells(data_sheet.Rows.Count, 1).End(xl_up).Row
            last_col = data_sheet.Cells(1, data_sheet.Columns.Count).End(xl_left).Column
            target_data_range = data_sheet.Range(data_sheet.Cells(1, 1), data_sheet.Cells(last_row, last_col))
            target_values = target_data_range.Value

            try:
                source_data_sheet.UsedRange.ClearContents()
            except Exception:
                pass
            destination_range = source_data_sheet.Range(
                source_data_sheet.Cells(1, 1), source_data_sheet.Cells(last_row, last_col)
            )
            destination_range.Value = target_values

            try:
                list_objects = source_data_sheet.ListObjects
                if list_objects.Count > 0:
                    list_objects(1).Resize(destination_range)
            except Exception:
                pass

            try:
                source_pivots = source_pivot_sheet.PivotTables()
                for pivot_index in range(1, source_pivots.Count + 1):
                    try:
                        source_pivots(pivot_index).RefreshTable()
                    except Exception:
                        pass
            except Exception:
                pass

        for sheet_index in range(target_wb.Sheets.Count, 0, -1):
            sheet = target_wb.Sheets(sheet_index)
            if sheet.Name == PIVOT_SHEET_NAME:
                sheet.Delete()

        source_pivot_sheet.Copy(Before=None, After=target_wb.Sheets(target_wb.Sheets.Count))
        copied_sheet = target_wb.Sheets(target_wb.Sheets.Count)
        copied_sheet.Name = PIVOT_SHEET_NAME

        local_source_data = None
        try:
            list_objects = data_sheet.ListObjects
            if list_objects.Count > 0:
                local_source_data = list_objects(1).Name
        except Exception:
            local_source_data = None

        if not local_source_data:
            local_source_data = data_sheet.UsedRange

        new_cache = target_wb.PivotCaches().Create(SourceType=1, SourceData=local_source_data)
        pivot_tables = copied_sheet.PivotTables()
        for pivot_index in range(1, pivot_tables.Count + 1):
            pivot_table = pivot_tables(pivot_index)
            try:
                if local_source_data:
                    try:
                        pivot_table.PivotCache().SourceData = local_source_data
                    except Exception:
                        pass
                try:
                    pivot_table.ChangePivotCache(new_cache)
                except Exception:
                    pass
                try:
                    pivot_table.ManualUpdate = False
                except Exception:
                    pass
                pivot_table.RefreshTable()
            except Exception:
                pass

        target_wb.Save()
    finally:
        if source_wb is not None:
            try:
                source_wb.Close(SaveChanges=False)
            except Exception:
                pass
        if target_wb is not None:
            try:
                target_wb.Close(SaveChanges=False)
            except Exception:
                pass
        if xl is not None:
            try:
                xl.Quit()
            except Exception:
                pass


def refresh_internal_pivot_sheet_via_com(
    output_path: Path,
    xl_app=None,
    workbook=None,
    save_workbook: bool = True,
) -> None:
    try:
        import win32com.client  # type: ignore
    except ImportError:
        raise RuntimeError("pywin32 is required to refresh internal PIVOT sheet.")

    xl = xl_app
    wb = workbook
    owns_excel = xl is None
    owns_workbook = wb is None
    try:
        if xl is None:
            xl = win32com.client.DispatchEx("Excel.Application")
            xl.Visible = False
            xl.DisplayAlerts = False

        if wb is None:
            wb = xl.Workbooks.Open(str(output_path.resolve()))
        sheet_names = {str(sheet.Name) for sheet in wb.Sheets}
        if PIVOT_SHEET_NAME not in sheet_names or SHEET_NAME not in sheet_names:
            if save_workbook:
                wb.Save()
            return

        data_sheet = wb.Sheets(SHEET_NAME)
        pivot_sheet = wb.Sheets(PIVOT_SHEET_NAME)

        try:
            pivot_tables = pivot_sheet.PivotTables()
            if pivot_tables.Count >= 1:
                try:
                    pivot_tables(1).PivotCache().Refresh()
                except Exception:
                    for pivot_index in range(1, pivot_tables.Count + 1):
                        try:
                            pivot_tables(pivot_index).RefreshTable()
                        except Exception:
                            pass
        except Exception:
            pass

        restore_ongoing_pivot_sheet_column_widths(pivot_sheet, reference_date_from_tracking_workbook_path(output_path))
        if save_workbook:
            wb.Save()
    finally:
        if owns_workbook and wb is not None:
            try:
                wb.Close(SaveChanges=False)
            except Exception:
                pass
        if owns_excel and xl is not None:
            try:
                xl.Quit()
            except Exception:
                pass


def synchronize_worksheet_table_columns(worksheet) -> None:
    for table in worksheet.tables.values():
        min_col, min_row, _, _ = range_boundaries(table.ref)
        last_column = worksheet.max_column
        last_row = worksheet.max_row
        table.ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(last_column)}{last_row}"
        if table.autoFilter is not None:
            table.autoFilter.ref = table.ref

        min_col, min_row, max_col, _ = range_boundaries(table.ref)
        header_values = [
            worksheet.cell(row=min_row, column=column_index).value for column_index in range(min_col, max_col + 1)
        ]
        table.tableColumns = []
        table._initialise_columns()
        for offset, table_column in enumerate(table.tableColumns):
            header_value = header_values[offset]
            table_column.name = str(header_value).strip() if header_value is not None else f"Column{offset + min_col}"


def style_added_headers(worksheet, header_columns: list[int]) -> None:
    fill = PatternFill(fill_type="solid", fgColor=EXCEL_YELLOW_HEADER_FILL)
    font = Font(color=EXCEL_RED_FONT_COLOR, bold=True)
    alignment = Alignment(horizontal="center", vertical="center")
    for column_index in header_columns:
        cell = worksheet.cell(row=1, column=column_index)
        cell.fill = fill
        cell.font = font
        cell.alignment = alignment


def build_start_date_payload_lookup(
    lookup: dict[str, str],
    aging_date: date,
) -> dict[str, tuple[object, object, object]]:
    payload_lookup: dict[str, tuple[object, object, object]] = {}
    aging_reference_date = pd.Timestamp(aging_date).normalize()
    parsed_cache: dict[str, pd.Timestamp | None] = {}

    for quo_value, raw_start_date in lookup.items():
        cleaned = str(raw_start_date).strip()
        if not cleaned:
            payload_lookup[quo_value] = ("#N/A", "#N/A", "#N/A")
            continue

        if cleaned not in parsed_cache:
            parsed_value = pd.to_datetime(cleaned, errors="coerce", format="mixed")
            parsed_cache[cleaned] = None if pd.isna(parsed_value) else parsed_value

        parsed = parsed_cache[cleaned]
        if parsed is None:
            payload_lookup[quo_value] = (cleaned, "#N/A", "#N/A")
            continue

        aging_value = int((aging_reference_date - parsed.normalize()).days)
        payload_lookup[quo_value] = (
            parsed.to_pydatetime(),
            aging_value,
            build_pre_installation_range(aging_value),
        )

    return payload_lookup


def apply_logic_to_workbook_openpyxl(
    tracking_workbook_path: Path,
    log_csv_path: Path,
    output_path: Path,
    with_pivot: bool = False,
    aging_date: date | None = None,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(tracking_workbook_path, output_path)

    reference_date = reference_date_from_tracking_workbook_path(tracking_workbook_path)
    aging_date = aging_date or reference_date
    lookup = load_pre_installation_lookup(log_csv_path)
    payload_lookup = build_start_date_payload_lookup(lookup, aging_date)

    wb = load_workbook(output_path)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Sheet '{SHEET_NAME}' not found in {tracking_workbook_path.name}.")
    ws = wb[SHEET_NAME]

    headers = [ws.cell(row=1, column=column_index).value for column_index in range(1, ws.max_column + 1)]
    normalized_headers = {
        normalize_header_key(header_value): column_index
        for column_index, header_value in enumerate(headers, start=1)
        if header_value is not None and str(header_value).strip()
    }

    anchor_column = normalized_headers.get(normalize_header_key(CATEGORY_MLD_HEADER))
    quo_column = normalized_headers.get(normalize_header_key(QUO_HEADER))
    if anchor_column is None:
        raise ValueError(f"Header '{CATEGORY_MLD_HEADER}' not found in sheet '{SHEET_NAME}'.")
    if quo_column is None:
        raise ValueError(f"Header '{QUO_HEADER}' not found in sheet '{SHEET_NAME}'.")

    requested_headers = [
        PRE_INSTALLATION_START_DATE_HEADER,
        PRE_INSTALLATION_AGING_HEADER,
        RANGE_AGING_PRE_INSTALLATION_HEADER,
    ]
    existing = set(normalized_headers.keys())
    for header in reversed(requested_headers):
        if normalize_header_key(header) not in existing:
            ws.insert_cols(anchor_column + 1)
            ws.cell(row=1, column=anchor_column + 1, value=header)

    headers = [ws.cell(row=1, column=column_index).value for column_index in range(1, ws.max_column + 1)]
    normalized_headers = {
        normalize_header_key(header_value): column_index
        for column_index, header_value in enumerate(headers, start=1)
        if header_value is not None and str(header_value).strip()
    }
    start_date_column = normalized_headers[normalize_header_key(PRE_INSTALLATION_START_DATE_HEADER)]
    aging_column = normalized_headers[normalize_header_key(PRE_INSTALLATION_AGING_HEADER)]
    range_column = normalized_headers[normalize_header_key(RANGE_AGING_PRE_INSTALLATION_HEADER)]
    style_added_headers(ws, [start_date_column, aging_column, range_column])

    for row_index in range(2, ws.max_row + 1):
        quo_value = normalize_quo_value(ws.cell(row=row_index, column=quo_column).value)
        start_value, aging_value, range_value = payload_lookup.get(quo_value, ("#N/A", "#N/A", "#N/A"))
        start_cell = ws.cell(row=row_index, column=start_date_column)
        if isinstance(start_value, datetime):
            start_cell.value = start_value
            start_cell.number_format = "yyyy-mm-dd hh:mm:ss"
        else:
            start_cell.value = start_value

        ws.cell(row=row_index, column=aging_column, value=aging_value)
        ws.cell(row=row_index, column=range_column, value=range_value)

    synchronize_worksheet_table_columns(ws)
    wb.save(output_path)
    if with_pivot:
        refresh_internal_pivot_sheet_via_com(output_path)
        update_on_progress_sheet_pivots_via_com(output_path, reference_date)


def apply_logic_to_workbook_com(
    tracking_workbook_path: Path,
    log_csv_path: Path,
    output_path: Path,
    with_pivot: bool = False,
    clone_pivot_template: bool = False,
    template_workbook_path: Path | None = None,
    aging_date: date | None = None,
) -> None:
    try:
        import win32com.client  # type: ignore
    except ImportError:
        raise RuntimeError("pywin32 is required for --engine com.")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    base_workbook_path = template_workbook_path or tracking_workbook_path
    if clone_pivot_template and template_workbook_path is None:
        try:
            base_workbook_path = resolve_validate_source_workbook(output_path.parent.parent)
        except FileNotFoundError:
            base_workbook_path = resolve_example_source_workbook(output_path.parent.parent)
    shutil.copy2(base_workbook_path, output_path)

    reference_date = reference_date_from_tracking_workbook_path(tracking_workbook_path)
    aging_date = aging_date or reference_date
    lookup = load_pre_installation_lookup(log_csv_path)
    payload_lookup = build_start_date_payload_lookup(lookup, aging_date)

    xl = None
    wb = None
    source_wb = None
    try:
        total_start = time.perf_counter()
        xl = win32com.client.DispatchEx("Excel.Application")
        xl.Visible = False
        xl.DisplayAlerts = False
        try:
            xl.ScreenUpdating = False
        except Exception:
            pass
        try:
            xl.EnableEvents = False
        except Exception:
            pass
        try:
            xl.Calculation = -4135  # xlCalculationManual
        except Exception:
            pass

        phase_start = time.perf_counter()
        wb = xl.Workbooks.Open(str(output_path.resolve()))
        profile_log("open target workbook", phase_start)
        if (clone_pivot_template or template_workbook_path is not None) and base_workbook_path.resolve() != tracking_workbook_path.resolve():
            phase_start = time.perf_counter()
            source_wb = xl.Workbooks.Open(str(tracking_workbook_path.resolve()))
            profile_log("open source workbook", phase_start)
            source_ws = source_wb.Worksheets(SHEET_NAME)
            target_ws = wb.Worksheets(SHEET_NAME)

            xl_up = -4162
            xl_left = -4159
            source_last_row = source_ws.Cells(source_ws.Rows.Count, 1).End(xl_up).Row
            source_last_col = source_ws.Cells(1, source_ws.Columns.Count).End(xl_left).Column
            target_last_row = target_ws.Cells(target_ws.Rows.Count, 1).End(xl_up).Row
            target_table_cols = source_last_col
            try:
                if target_ws.ListObjects.Count > 0:
                    target_table_cols = max(target_table_cols, target_ws.ListObjects(1).Range.Columns.Count)
            except Exception:
                pass

            source_headers_raw = source_ws.Range(source_ws.Cells(1, 1), source_ws.Cells(1, source_last_col)).Value
            if isinstance(source_headers_raw, tuple):
                source_headers = list(
                    source_headers_raw[0] if source_headers_raw and isinstance(source_headers_raw[0], tuple) else source_headers_raw
                )
            else:
                source_headers = [source_headers_raw]

            source_target_header, source_phase_header = find_source_dynamic_headers(source_headers, reference_date)
            synchronize_dynamic_headers_on_com_sheet(
                target_ws,
                target_table_cols,
                source_target_header,
                source_phase_header,
            )

            source_header_to_col: dict[str, int] = {}
            source_header_model_by_target_key: dict[str, int] = {}
            source_target_determined_col = None
            source_target_determined_best_count = -1
            for idx, header_value in enumerate(source_headers, start=1):
                normalized_key = normalize_header_key(header_value)
                if normalized_key and normalized_key not in source_header_to_col:
                    source_header_to_col[normalized_key] = idx
                if normalized_key in {"groupsales", "column1", "divisionsales", "segmentsales", "sales"}:
                    target_model_key = "divisionsales" if normalized_key == "column1" else normalized_key
                    source_header_model_by_target_key[target_model_key] = idx
                if is_target_determined_header(header_value):
                    cnt = count_non_empty_column_values(source_ws, idx)
                    if cnt > source_target_determined_best_count:
                        source_target_determined_best_count = cnt
                        source_target_determined_col = idx

            phase_start = time.perf_counter()
            if target_last_row >= 2:
                target_ws.Range(
                    target_ws.Cells(2, 1), target_ws.Cells(target_last_row, target_table_cols)
                ).ClearContents()
            profile_log("clear target all order rows", phase_start)

            if source_last_row >= 2:
                phase_start = time.perf_counter()
                target_headers_raw = target_ws.Range(target_ws.Cells(1, 1), target_ws.Cells(1, target_table_cols)).Value
                if isinstance(target_headers_raw, tuple):
                    target_headers = list(
                        target_headers_raw[0]
                        if target_headers_raw and isinstance(target_headers_raw[0], tuple)
                        else target_headers_raw
                    )
                else:
                    target_headers = [target_headers_raw]

                for target_col in range(1, target_table_cols + 1):
                    target_header_value = target_headers[target_col - 1] if target_col - 1 < len(target_headers) else ""
                    target_key = normalize_header_key(target_header_value)
                    source_col = source_header_to_col.get(target_key)
                    if source_col is None and target_key == "divisionsales":
                        source_col = source_header_to_col.get("column1")
                    if source_col is None and is_target_determined_header(target_header_value):
                        source_col = source_target_determined_col
                    if source_col is None:
                        continue
                    original_source_col = source_col
                    source_column_range = source_ws.Range(
                        source_ws.Cells(2, source_col), source_ws.Cells(source_last_row, source_col)
                    )

                    # Month rollover fallback:
                    # when TARGET Determined column is empty in source, use Target In Week values.
                    if is_target_determined_header(target_header_value):
                        source_column_values_for_check = source_column_range.Value2
                        fallback_col = source_header_to_col.get("targetinweek")
                        if fallback_col is not None:
                            is_empty = True
                            if isinstance(source_column_values_for_check, tuple):
                                for row_value in source_column_values_for_check:
                                    value = row_value[0] if isinstance(row_value, tuple) else row_value
                                    if value not in (None, ""):
                                        is_empty = False
                                        break
                            else:
                                is_empty = source_column_values_for_check in (None, "")
                            if is_empty:
                                source_col = fallback_col
                                source_column_range = source_ws.Range(
                                    source_ws.Cells(2, source_col), source_ws.Cells(source_last_row, source_col)
                                )

                    target_column_range = target_ws.Range(
                        target_ws.Cells(2, target_col), target_ws.Cells(source_last_row, target_col)
                    )
                    if target_key == normalize_header_key(YEAR_FAB_UPLOAD_HEADER):
                        source_column_values = source_column_range.Value2
                        normalized_year_values = []
                        for raw_value in source_column_values:
                            value = raw_value[0] if isinstance(raw_value, tuple) else raw_value
                            normalized_year_values.append((str(value).strip() if value not in (None, "") else "",))
                        target_column_range.NumberFormat = "@"
                        try:
                            target_column_range.Value = tuple(normalized_year_values)
                        except Exception:
                            source_column_range.Copy()
                            target_column_range.PasteSpecial(-4163)  # xlPasteValues
                    else:
                        try:
                            if should_preserve_excel_date_values(target_key):
                                try:
                                    target_column_range.NumberFormat = excel_date_number_format_for_header(
                                        target_key,
                                        source_ws.Cells(2, source_col).NumberFormat,
                                    )
                                except Exception:
                                    pass
                                target_column_range.Value = source_column_range.Value
                            elif source_col != target_col or source_col != original_source_col:
                                target_column_range.Value2 = source_column_range.Value2
                            else:
                                target_column_range.Value2 = source_column_range.Value2
                        except Exception:
                            source_column_range.Copy()
                            target_column_range.PasteSpecial(-4163)  # xlPasteValues
                        if target_key in {"groupsales", "divisionsales", "segmentsales"}:
                            try:
                                target_column_range.Replace(
                                    What=-2146826246,
                                    Replacement="#N/A",
                                    LookAt=1,
                                )
                            except Exception:
                                pass
                try:
                    xl.CutCopyMode = False
                except Exception:
                    pass
                profile_log("copy all order columns", phase_start)

                # Final safeguard for month rollover:
                # if TARGET Determined column is still empty, mirror from Target In Week.
                target_determined_col = None
                target_in_week_col = None
                for column_index in range(1, target_table_cols + 1):
                    header_value = target_headers[column_index - 1] if column_index - 1 < len(target_headers) else ""
                    if target_determined_col is None and is_target_determined_header(header_value):
                        target_determined_col = column_index
                    if target_in_week_col is None and normalize_header_key(header_value) == "targetinweek":
                        target_in_week_col = column_index
                if target_determined_col is not None and target_in_week_col is not None:
                    try:
                        det_range = target_ws.Range(
                            target_ws.Cells(2, target_determined_col), target_ws.Cells(source_last_row, target_determined_col)
                        )
                        det_values = det_range.Value2
                        det_has_value = False
                        if isinstance(det_values, tuple):
                            for row_value in det_values:
                                value = row_value[0] if isinstance(row_value, tuple) else row_value
                                if value not in (None, ""):
                                    det_has_value = True
                                    break
                        else:
                            det_has_value = det_values not in (None, "")
                        if not det_has_value:
                            in_week_range = target_ws.Range(
                                target_ws.Cells(2, target_in_week_col), target_ws.Cells(source_last_row, target_in_week_col)
                            )
                            det_range.Value2 = in_week_range.Value2
                    except Exception:
                        pass

            try:
                if target_ws.ListObjects.Count > 0:
                    target_table = target_ws.ListObjects(1)
                    target_table.Resize(
                        target_ws.Range(target_ws.Cells(1, 1), target_ws.Cells(source_last_row, target_table_cols))
                    )
            except Exception:
                pass
        if SHEET_NAME not in [str(sheet.Name) for sheet in wb.Worksheets]:
            raise ValueError(f"Sheet '{SHEET_NAME}' not found in {tracking_workbook_path.name}.")
        ws = wb.Worksheets(SHEET_NAME)

        used_cols = ws.UsedRange.Columns.Count
        header_values = ws.Range(ws.Cells(1, 1), ws.Cells(1, used_cols)).Value
        if isinstance(header_values, tuple):
            header_row = list(header_values[0] if isinstance(header_values[0], tuple) else header_values)
        else:
            header_row = [header_values]

        normalized_headers = {
            normalize_header_key(header_value): column_index
            for column_index, header_value in enumerate(header_row, start=1)
            if header_value is not None and str(header_value).strip()
        }
        anchor_column = normalized_headers.get(normalize_header_key(CATEGORY_MLD_HEADER))
        quo_column = normalized_headers.get(normalize_header_key(QUO_HEADER))
        if anchor_column is None:
            raise ValueError(f"Header '{CATEGORY_MLD_HEADER}' not found in sheet '{SHEET_NAME}'.")
        if quo_column is None:
            raise ValueError(f"Header '{QUO_HEADER}' not found in sheet '{SHEET_NAME}'.")

        requested_headers = [
            PRE_INSTALLATION_START_DATE_HEADER,
            PRE_INSTALLATION_AGING_HEADER,
            RANGE_AGING_PRE_INSTALLATION_HEADER,
        ]
        existing = set(normalized_headers.keys())
        for header in reversed(requested_headers):
            if normalize_header_key(header) not in existing:
                ws.Columns(anchor_column + 1).Insert()
                ws.Cells(1, anchor_column + 1).Value = header

        used_cols = ws.UsedRange.Columns.Count
        header_values = ws.Range(ws.Cells(1, 1), ws.Cells(1, used_cols)).Value
        if isinstance(header_values, tuple):
            header_row = list(header_values[0] if isinstance(header_values[0], tuple) else header_values)
        else:
            header_row = [header_values]
        normalized_headers = {
            normalize_header_key(header_value): column_index
            for column_index, header_value in enumerate(header_row, start=1)
            if header_value is not None and str(header_value).strip()
        }

        start_date_column = normalized_headers[normalize_header_key(PRE_INSTALLATION_START_DATE_HEADER)]
        aging_column = normalized_headers[normalize_header_key(PRE_INSTALLATION_AGING_HEADER)]
        range_column = normalized_headers[normalize_header_key(RANGE_AGING_PRE_INSTALLATION_HEADER)]

        header_range = ws.Range(ws.Cells(1, start_date_column), ws.Cells(1, range_column))
        header_range.HorizontalAlignment = -4108  # xlCenter
        header_range.VerticalAlignment = -4108  # xlCenter
        header_range.Interior.Color = 65535  # yellow
        header_range.Font.Bold = True
        header_range.Font.Color = 255  # red

        xl_up = -4162
        last_row = ws.Cells(ws.Rows.Count, quo_column).End(xl_up).Row
        if last_row >= 2:
            phase_start = time.perf_counter()
            quo_values = ws.Range(ws.Cells(2, quo_column), ws.Cells(last_row, quo_column)).Value
            if isinstance(quo_values, tuple):
                if quo_values and isinstance(quo_values[0], tuple):
                    quo_rows = list(quo_values)
                else:
                    quo_rows = [(value,) for value in quo_values]
            else:
                quo_rows = [(quo_values,)]

            start_values = []
            aging_values = []
            range_values = []
            for row_value in quo_rows:
                quo_value = normalize_quo_value(row_value[0])
                start_value, aging_value, range_value = payload_lookup.get(quo_value, ("#N/A", "#N/A", "#N/A"))
                if isinstance(start_value, datetime):
                    start_value = excel_serial_from_datetime(start_value)
                start_values.append((start_value,))
                aging_values.append((aging_value,))
                range_values.append((range_value,))

            ws.Range(ws.Cells(2, start_date_column), ws.Cells(last_row, start_date_column)).Value = tuple(start_values)
            ws.Range(ws.Cells(2, aging_column), ws.Cells(last_row, aging_column)).Value = tuple(aging_values)
            ws.Range(ws.Cells(2, range_column), ws.Cells(last_row, range_column)).Value = tuple(range_values)
            ws.Range(ws.Cells(2, start_date_column), ws.Cells(last_row, start_date_column)).NumberFormat = (
                "m/d/yy h:mm AM/PM"
            )
            profile_log("write pre-installation fields", phase_start)

        normalize_all_order_date_column_formats(ws)
        apply_sales_hierarchy_header_style_on_com_sheet(ws)
        if clone_pivot_template and not with_pivot:
            phase_start = time.perf_counter()
            fast_refresh_ok = refresh_reporting_sheets_fast_in_workbook(wb, reference_date)
            profile_log("fast refresh reporting sheets", phase_start)
            if not fast_refresh_ok:
                phase_start = time.perf_counter()
                refresh_internal_pivot_sheet_via_com(
                    output_path,
                    xl_app=xl,
                    workbook=wb,
                    save_workbook=False,
                )
                profile_log("refresh pivot sheet", phase_start)
                phase_start = time.perf_counter()
                update_on_progress_sheet_pivots_via_com(
                    output_path,
                    reference_date,
                    xl_app=xl,
                    workbook=wb,
                    save_workbook=False,
                )
                profile_log("refresh on progress sheet", phase_start)
        phase_start = time.perf_counter()
        wb.Save()
        profile_log("save workbook", phase_start)
        profile_log("total com processing", total_start)
    finally:
        if source_wb is not None:
            try:
                source_wb.Close(SaveChanges=False)
            except Exception:
                pass
        if wb is not None:
            try:
                wb.Close(SaveChanges=False)
            except Exception:
                pass
        if xl is not None:
            try:
                xl.Calculation = -4105  # xlCalculationAutomatic
            except Exception:
                pass
            try:
                xl.Quit()
            except Exception:
                pass

    if with_pivot:
        refresh_internal_pivot_sheet_via_com(output_path)
        update_on_progress_sheet_pivots_via_com(output_path, reference_date)


def apply_logic_to_workbook(
    tracking_workbook_path: Path,
    log_csv_path: Path,
    output_path: Path,
    with_pivot: bool = False,
    engine: str = "com",
    clone_pivot_template: bool = False,
    template_workbook_path: Path | None = None,
    aging_date: date | None = None,
) -> None:
    if engine == "openpyxl":
        apply_logic_to_workbook_openpyxl(
            tracking_workbook_path,
            log_csv_path,
            output_path,
            with_pivot=with_pivot,
            aging_date=aging_date,
        )
        return

    try:
        apply_logic_to_workbook_com(
            tracking_workbook_path,
            log_csv_path,
            output_path,
            with_pivot=with_pivot,
            clone_pivot_template=clone_pivot_template,
            template_workbook_path=template_workbook_path,
            aging_date=aging_date,
        )
    except RuntimeError:
        apply_logic_to_workbook_openpyxl(
            tracking_workbook_path,
            log_csv_path,
            output_path,
            with_pivot=with_pivot,
            aging_date=aging_date,
        )


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build ongoing tracking workbook from ongoing input files.")
    parser.add_argument(
        "input",
        nargs="?",
        type=Path,
        default=Path(DEFAULT_INPUT_DIR),
        help=f"Input directory containing {TRACKING_GLOB} and {LOG_GLOB}. Defaults to .\\{DEFAULT_INPUT_DIR}.",
    )
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        help=f"Output .xlsx file or output directory. Defaults to .\\{DEFAULT_OUTPUT_DIR}.",
    )
    parser.add_argument(
        "--with-pivot",
        action="store_true",
        help="Run slower Excel COM pivot/sheet updates after data generation.",
    )
    parser.add_argument(
        "--engine",
        choices=("com", "openpyxl"),
        default="com",
        help="Processing engine. 'com' is faster and preserves workbook internals better.",
    )
    parser.add_argument(
        "--clone-pivot-template",
        action="store_true",
        default=True,
        help="Use validate/example workbook as base to preserve PIVOT template format, then refresh from internal ALL ORDER.",
    )
    parser.add_argument(
        "--skip-template-refresh",
        action="store_false",
        dest="clone_pivot_template",
        help="Skip template cloning/refresh and write a faster workbook based on the input tracking file.",
    )
    parser.add_argument(
        "--aging-date",
        type=lambda value: datetime.strptime(value, "%Y-%m-%d").date(),
        default=None,
        help="Date used to calculate Pre-Installation Aging. Defaults to report date from the input workbook filename. Format: YYYY-MM-DD.",
    )
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv or sys.argv[1:])
    input_path = args.input.resolve()
    try:
        tracking_workbook_path, log_csv_path = resolve_ongoing_input_files(input_path)
        output_path = resolve_output_path(input_path, tracking_workbook_path, args.output).resolve()
        apply_logic_to_workbook(
            tracking_workbook_path,
            log_csv_path,
            output_path,
            with_pivot=args.with_pivot,
            engine=args.engine,
            clone_pivot_template=args.clone_pivot_template,
            aging_date=args.aging_date,
        )
        print(f"Wrote {output_path}")
    except Exception as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 1

    return 0


if __name__ == "__main__":
    raise SystemExit(main())

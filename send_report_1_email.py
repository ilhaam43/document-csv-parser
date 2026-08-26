"""Capture an Excel report range and send it as an inline email image.

The SMTP API must preserve Content-ID and Content-Disposition=inline for
Outlook to render the image in the HTML message body. The image is sent as
raw multipart data, matching curl's ``attachment[]=@...`` behavior.
"""

from __future__ import annotations

import argparse
import mimetypes
import os
import sys
import time
import uuid
from pathlib import Path
from urllib import request


DEFAULT_ENDPOINT = "http://10.34.144.197/secm-portal/smtp/api_send_email"
DEFAULT_RECIPIENT = "ilhaam.akmal@lintasarta.co.id"
INLINE_IMAGE_IDS = (
    "target-complete-table",
    "target-complete-legend",
    "target-after-table",
    "target-after-legend",
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, required=True, help="Excel workbook to capture")
    parser.add_argument("--sheet", help="Worksheet name; defaults to the active sheet")
    parser.add_argument("--range", dest="cell_range", help="Range such as A1:K40; defaults to UsedRange")
    parser.add_argument("--image", type=Path, help="PNG output path; defaults beside the workbook")
    parser.add_argument("--endpoint", default=os.getenv("SMTP_API_URL", DEFAULT_ENDPOINT))
    parser.add_argument("--to", default=os.getenv("SMTP_TO", DEFAULT_RECIPIENT))
    parser.add_argument("--subject", default="Daily Tracking Report")
    parser.add_argument("--message", default="<p>Daily Tracking Report</p><img src=\"cid:daily-tracking-image\" alt=\"Daily Tracking Report\">")
    parser.add_argument("--message-file", type=Path, help="UTF-8 HTML file used instead of --message")
    parser.add_argument("--timeout", type=int, default=30)
    parser.add_argument("--dry-run", action="store_true", help="Create the image and print the request fields without calling the API")
    return parser.parse_args()


def excel_range_to_png(
    workbook_path: Path,
    image_path: Path,
    sheet_name: str | None,
    cell_range: str | None,
    scale: float = 1.0,
) -> None:
    try:
        import pythoncom  # type: ignore
        import win32com.client  # type: ignore
    except ImportError as exc:
        raise RuntimeError("Install Windows Excel COM dependencies with: pip install pywin32") from exc

    excel = None
    workbook = None
    chart = None
    pythoncom.CoInitialize()
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        # Excel's CopyPicture requires an interactive desktop window. Run this
        # workflow under a signed-in Windows user on the server.
        excel.Visible = True
        excel.ScreenUpdating = True
        excel.DisplayAlerts = False
        workbook = excel.Workbooks.Open(str(workbook_path.resolve()), ReadOnly=True, UpdateLinks=0)
        worksheet = workbook.Worksheets(sheet_name) if sheet_name else workbook.ActiveSheet
        selected_range = worksheet.Range(cell_range) if cell_range else worksheet.UsedRange
        workbook.Activate()
        worksheet.Activate()
        selected_range.Select()
        selected_range.CopyPicture(Appearance=1, Format=2)
        time.sleep(0.75)

        width = max(1, int(selected_range.Width * scale))
        height = max(1, int(selected_range.Height * scale))
        chart = worksheet.ChartObjects().Add(0, 0, width, height)
        chart.Activate()
        for attempt in range(3):
            try:
                chart.Chart.Paste()
                break
            except Exception:
                if attempt == 2:
                    raise
                time.sleep(0.75)
        image_path.parent.mkdir(parents=True, exist_ok=True)
        chart.Chart.Export(str(image_path.resolve()), "PNG")
    finally:
        if chart is not None:
            try:
                chart.Delete()
            except Exception:
                pass
        if workbook is not None:
            workbook.Close(SaveChanges=False)
        if excel is not None:
            excel.Quit()
        pythoncom.CoUninitialize()

    if not image_path.exists() or image_path.stat().st_size == 0:
        raise RuntimeError(f"Excel did not create an image at {image_path}")


def discover_pivot_ranges(workbook_path: Path) -> list[tuple[str, str]]:
    """Find the four requested PIVOT regions from their visible headers."""
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    sheet = workbook["PIVOT"]
    headers: dict[str, tuple[int, int]] = {}
    legends: list[tuple[int, int]] = []
    for row in sheet.iter_rows():
        for cell in row:
            value = str(cell.value or "").strip().upper()
            if "TARGET COMPLETE" in value and "TARGET AFTER" not in value and "target_complete" not in headers:
                headers["target_complete"] = (cell.row, cell.column)
            if "TARGET AFTER" in value and "target_after" not in headers:
                headers["target_after"] = (cell.row, cell.column)
            if "DELAY COMPLETION" in value and "delay_completion" not in headers:
                headers["delay_completion"] = (cell.row, cell.column)
            if "TARGET NOT INPUTTED" in value and "target_not_inputted" not in headers:
                headers["target_not_inputted"] = (cell.row, cell.column)
            if value.startswith(("GREEN", "YELLOW", "RED")):
                legends.append((cell.row, cell.column))

    missing = [name for name in ("target_complete", "target_after") if name not in headers]
    if missing:
        raise RuntimeError(f"Could not find PIVOT headers: {', '.join(missing)}")

    complete_row = headers["target_complete"][0]
    after_row = headers["target_after"][0]
    after_end = headers.get("target_not_inputted", (sheet.max_row + 1, 1))[0] - 1
    table_end_column = max(
        (cell.column for row in sheet.iter_rows(min_row=complete_row, max_row=max(complete_row, after_end - 1)) for cell in row if cell.value is not None),
        default=18,
    )
    legend_columns = [column for _, column in legends]
    if legend_columns:
        table_end_column = min(table_end_column, min(legend_columns) - 1)
    table_end_column = max(1, min(table_end_column, 38))
    table_end_letter = get_column_letter(table_end_column)

    complete_legend = min((item for item in legends if item[0] < after_row), default=(complete_row + 9, 16))
    after_legend = min((item for item in legends if item[0] > after_row), default=(after_row + 9, 16))

    def legend_range(anchor: tuple[int, int]) -> str:
        row, column = anchor
        return f"{get_column_letter(column)}{row}:{get_column_letter(column)}{row + 2}"

    return [
        (f"A{complete_row}:{table_end_letter}{after_row - 1}", INLINE_IMAGE_IDS[0]),
        (legend_range(complete_legend), INLINE_IMAGE_IDS[1]),
        (f"A{after_row}:{table_end_letter}{after_end}", INLINE_IMAGE_IDS[2]),
        (legend_range(after_legend), INLINE_IMAGE_IDS[3]),
    ]


def excel_pivot_regions_to_png(workbook_path: Path, image_dir: Path) -> list[tuple[Path, str, str]]:
    """Capture all requested PIVOT regions in one Excel session."""
    ranges = discover_pivot_ranges(workbook_path)
    image_dir.mkdir(parents=True, exist_ok=True)
    outputs: list[tuple[Path, str, str]] = []
    for index, (cell_range, content_id) in enumerate(ranges, start=1):
        image_path = image_dir / f"report-1-pivot-{index}.png"
        scale = 3.0 if "legend" in content_id else 1.0
        excel_range_to_png(workbook_path, image_path, "PIVOT", cell_range, scale=scale)
        outputs.append((image_path, content_id, cell_range))
    return outputs


def multipart_form(fields: dict[str, str], image_path: Path, content_id: str) -> tuple[bytes, str]:
    boundary = f"----ReportEmailBoundary{uuid.uuid4().hex}"
    chunks: list[bytes] = []

    for name, value in fields.items():
        chunks.extend([
            f"--{boundary}\r\n".encode(),
            f'Content-Disposition: form-data; name="{name}"\r\n\r\n'.encode(),
            value.encode("utf-8"),
            b"\r\n",
        ])

    filename = image_path.name
    content_type = mimetypes.guess_type(filename)[0] or "image/png"
    chunks.extend([
        f"--{boundary}\r\n".encode(),
        (
            'Content-Disposition: form-data; name="attachment[]"; '
            f'filename="{filename}"\r\n'
            f"Content-Type: {content_type}\r\n"
            f"Content-ID: <{content_id}>\r\n"
            "\r\n"
        ).encode(),
        image_path.read_bytes(),
        b"\r\n",
        f"--{boundary}--\r\n".encode(),
    ])
    return b"".join(chunks), f"multipart/form-data; boundary={boundary}"


def multipart_form_images(fields: dict[str, str], images: list[tuple[Path, str]]) -> tuple[bytes, str]:
    boundary = f"----ReportEmailBoundary{uuid.uuid4().hex}"
    chunks: list[bytes] = []
    for name, value in fields.items():
        chunks.extend([
            f"--{boundary}\r\n".encode(),
            f'Content-Disposition: form-data; name="{name}"\r\n\r\n'.encode(),
            value.encode("utf-8"),
            b"\r\n",
        ])
    for image_path, content_id in images:
        chunks.extend([
            f"--{boundary}\r\n".encode(),
            (
                'Content-Disposition: form-data; name="attachment[]"; '
                f'filename="{image_path.name}"\r\n'
                "Content-Type: image/png\r\n"
                f"Content-ID: <{content_id}>\r\n\r\n"
            ).encode(),
            image_path.read_bytes(),
            b"\r\n",
        ])
    chunks.append(f"--{boundary}--\r\n".encode())
    return b"".join(chunks), f"multipart/form-data; boundary={boundary}"


def send_email(endpoint: str, recipient: str, subject: str, message: str, image_path: Path, timeout: int, content_id: str = "daily-tracking-image") -> None:
    fields = {
        "to": recipient,
        "subject": subject,
        "message": message,
        # These fields help APIs that expose inline metadata separately.
        "attachment_content_id[]": content_id,
        "attachment_disposition[]": "inline",
    }
    body, content_type = multipart_form(fields, image_path, content_id)
    req = request.Request(endpoint, data=body, method="POST", headers={"Content-Type": content_type})
    with request.urlopen(req, timeout=timeout) as response:
        response_body = response.read().decode("utf-8", errors="replace")
        print(f"SMTP API responded {response.status}: {response_body}")


def send_email_images(endpoint: str, recipient: str, subject: str, message: str, images: list[tuple[Path, str]], timeout: int) -> None:
    content_ids = [content_id for _, content_id in images]
    fields = {
        "to": recipient,
        "subject": subject,
        "message": message,
        "attachment_content_id[]": "\n".join(content_ids),
        "attachment_disposition[]": "\n".join("inline" for _ in images),
    }
    body, content_type = multipart_form_images(fields, images)
    req = request.Request(endpoint, data=body, method="POST", headers={"Content-Type": content_type})
    with request.urlopen(req, timeout=timeout) as response:
        response_body = response.read().decode("utf-8", errors="replace")
        print(f"SMTP API responded {response.status}: {response_body}")


def main() -> int:
    args = parse_args()
    workbook = args.workbook.resolve()
    if not workbook.exists():
        print(f"Workbook not found: {workbook}", file=sys.stderr)
        return 2

    image_path = (args.image or workbook.with_suffix(".png")).resolve()
    message = args.message_file.read_text(encoding="utf-8") if args.message_file else args.message
    if "cid:daily-tracking-image" not in message:
        print("Warning: HTML message does not reference cid:daily-tracking-image", file=sys.stderr)

    excel_range_to_png(workbook, image_path, args.sheet, args.cell_range)
    print(f"Created image: {image_path}")

    if args.dry_run:
        print(f"Dry run: would POST multipart data to {args.endpoint}")
        print(f"Recipient: {args.to}")
        print(f"Subject: {args.subject}")
        return 0

    try:
        send_email(args.endpoint, args.to, args.subject, message, image_path, args.timeout)
    except Exception as exc:
        print(f"SMTP API request failed: {exc}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
